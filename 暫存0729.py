import os
import platform
import subprocess
import sys
import unicodedata
import json
import math
import pandas as pd
import customtkinter as ctk
import tkinter as tk
import tkinter.font as tkfont
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta
from tkcalendar import DateEntry, Calendar
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

# 初始化 GUI 風格
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

global start_date, end_date

class ScheduleApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("工單排程工具")
        self.geometry("500x300")

        self.label = ctk.CTkLabel(self, text="碳帶工單排程", font=("Arial", 20))
        self.label.pack(pady=20)

        self.run_button = ctk.CTkButton(self, text="執行排程", command=self.run_schedule)
        self.run_button.pack(pady=10)

        self.output_label = ctk.CTkLabel(self, text="")
        self.output_label.pack(pady=10)

        # 綁定關閉事件
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def run_schedule(self):
        self.run_button.configure(state="disabled")
        self.output_label.configure(text="⏳ 處理中，請稍候...")
        self.update_idletasks()  # 強制更新畫面
        try:
            output_path = main() # 排程主程式
            self.output_label.configure(text="✅ 排程完成！結果已輸出至 Excel")
            self.after(500, lambda: open_file(output_path))

        except Exception as e:
            messagebox.showerror("錯誤", f"執行失敗，請確認資料格式或 Excel 是否已開啟。\n詳細錯誤：{str(e)[:200]}")
            print(str(e)[:200])
            self.output_label.configure(text="❌ 排程失敗")

        finally:
            self.run_button.configure(state="normal")

    def on_closing(self):
        # 這邊可以放你要結束前要做的事情
        self.destroy()   # 關閉視窗
        sys.exit()       # 強制結束程式

# 選擇日期區間彈出視窗
class DateFilterDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("選擇日期區間")
        self.geometry("500x360")  # 放大視窗

        self.start_date_choose = None
        self.end_date_choose = None

        # 標題文字
        ctk.CTkLabel(self, text="📅 請選擇日期區間", font=("微軟正黑體", 20, "bold")).pack(pady=(15, 10))

        # 設定共用樣式
        label_font = ("微軟正黑體", 14)
        dateentry_style = {
            "font": ("微軟正黑體", 14),
            "width": 20,
            "background": "#004080",
            "foreground": "white",
            "borderwidth": 2,
            "date_pattern": "yyyy-mm-dd"
        }

        # 起始日期
        ctk.CTkLabel(self, text="起始日期", font=label_font).pack()
        self.start_entry = DateEntry(self, **dateentry_style)
        self.start_entry.pack(pady=(5, 15))

        # 結束日期
        ctk.CTkLabel(self, text="結束日期", font=label_font).pack()
        self.end_entry = DateEntry(self, **dateentry_style)
        self.end_entry.pack(pady=(5, 15))

        # 錯誤訊息
        self.error_label = ctk.CTkLabel(self, text="", text_color="red", font=("微軟正黑體", 12))
        self.error_label.pack(pady=5)

        # 確定按鈕
        ctk.CTkButton(self, text="✅ 確定", font=("微軟正黑體", 14), command=self.confirm).pack(pady=15)

        self.grab_set()  # 鎖定視窗
        self.focus()

    def confirm(self): 
        try:
            self.start_date_choose = datetime.strptime(self.start_entry.get(), "%Y-%m-%d")
            self.end_date_choose = datetime.strptime(self.end_entry.get(), "%Y-%m-%d")
            self.destroy()
        except ValueError:
            self.error_label.configure(text="⚠️ 請選擇有效的日期")


def process_schedule_data():
    
    # 讓使用者選擇排程資料 Excel
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="請選擇排程資料 Excel 檔案",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not file_path:
        print("❌ 沒有選擇任何檔案，程式結束。")
        return  # 改用 return 而不是 exit()

    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得腳本所在目錄
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    # 基本資料路徑
    #base_path = r"E:\ribbon_schedule\\data\\基本資料-20250617-All.xlsx"
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return

    # 讀取資料
    order_df = pd.read_excel(file_path, dtype=str)
    base_df = pd.read_excel(base_path, dtype=str)

    # 過濾條件 1. 忽略81R, 81T  2. 狀態為Released   3. 根據選擇日期範圍去篩選開工時間資料
    order_df = order_df[~order_df["工單號碼"].str.startswith(("81R", "81T"))]
    order_df = order_df[order_df["狀態"].str.lower() == "released"] # complete released
    order_df = choose_date(order_df)

    # 數值處理
    base_df["寬度Cm"] = base_df["寬度Cm"].astype(float)
    base_df["車數"] = base_df["車數"].astype(float)
    base_df["搭1產出車數"] = base_df["搭1產出車數"].fillna(0).astype(float)
    order_df["料號"] = order_df["料號"].str.strip()
    order_df["開工數量"] = order_df["開工數量"].astype(int)
    

    # 合併資料
    merged = pd.merge(order_df, base_df, how="left", on="料號")
    merged["總長度(cm)"] = merged["寬度Cm"] * merged["車數"]
    merged["刀次"] = (merged["開工數量"] / merged["車數"]).round(2)

    # 無須配對條件
    '''
    mask_no_pair = (
        merged["工單號碼"].str.startswith(("81A", "81B")) &
        (merged["刀次"] % 1 == 0) &
        ((merged["總長度(cm)"] == 88) | (merged["總長度(cm)"] == 89) | (merged["總長度(cm)"] == 68))
    )
    '''
    mask_no_pair = (
        merged["工單號碼"].str.startswith(("81A", "81B")) &
        (merged["搭1產出車數"] == 0) &
        (merged["刀次"] % 1 == 0)
    )
    
    no_pair_df = merged[mask_no_pair].copy()
    remaining_df = merged[~mask_no_pair].copy()

    # 88cm 配對（備註配對）
    pair_rows_88 = []
    df_81c_remarks = merged[(merged["工單號碼"].str.startswith("81C")) & (merged["備註"].notna())]

    for _, c_order in df_81c_remarks.iterrows():
        pair_order_id = c_order["備註"]
        matched_rows = remaining_df[remaining_df["工單號碼"] == pair_order_id]
        if matched_rows.empty:
            continue
        a_order = matched_rows.iloc[0]
        for cars in range(1, 5):
            total_cm = a_order["總長度(cm)"] + c_order["寬度Cm"] * cars
            if abs(total_cm - 88) <= 1:
                a_copy = a_order.copy()
                c_copy = c_order.copy()
                c_copy["車數"] = cars
                c_copy["總長度(cm)"] = c_copy["寬度Cm"] * cars
                c_copy["刀次"] = 0
                pair_rows_88.extend([a_copy, c_copy])
                remaining_df = remaining_df[remaining_df["工單號碼"] != pair_order_id]
                remaining_df = remaining_df[remaining_df["工單號碼"] != c_order["工單號碼"]]
                break

    # 68cm 配對（備註配對）
    pair_rows_68 = []
    df_81c_remain = remaining_df[(remaining_df["工單號碼"].str.startswith("81C")) & (remaining_df["備註"].notna())]

    for _, c_order in df_81c_remain.iterrows():
        pair_order_id = c_order["備註"]
        matched_rows = remaining_df[remaining_df["工單號碼"] == pair_order_id]
        if matched_rows.empty:
            continue
        a_order = matched_rows.iloc[0]
        for cars in range(1, 5):
            total_cm = a_order["總長度(cm)"] + c_order["寬度Cm"] * cars
            if abs(total_cm - 68) <= 1:
                a_copy = a_order.copy()
                c_copy = c_order.copy()
                c_copy["車數"] = cars
                c_copy["總長度(cm)"] = c_copy["寬度Cm"] * cars
                c_copy["刀次"] = 0
                pair_rows_68.extend([a_copy, c_copy])
                remaining_df = remaining_df[remaining_df["工單號碼"] != pair_order_id]
                remaining_df = remaining_df[remaining_df["工單號碼"] != c_order["工單號碼"]]
                break

    # ➕ 搭1料號配對（修正版）
    pair_rows_d1 = []

    # 重新建立剩餘工單索引
    remaining_df = remaining_df.reset_index(drop=True)
    

    # 建立料號到工單列表的映射
    remaining_by_item = {}
    for _, row in remaining_df.iterrows():
        remaining_by_item.setdefault(row["料號"], []).append(row)

    for _, main_order in remaining_df.iterrows():
        main_order_id = main_order["工單號碼"]
        main_item_code = main_order["料號"]

        # 找基本資料搭1料號
        base_info = base_df[base_df["料號"] == main_item_code]
        if base_info.empty:
            continue
        d1_code = base_info.iloc[0].get("搭1料號")
        d1_car_count = base_info.iloc[0].get("搭1產出車數")
        if pd.isna(d1_code) or d1_code == "":
            continue

        d1_base_info = base_df[base_df["料號"] == d1_code]
        if d1_base_info.empty:
            continue
        d1_width = d1_base_info.iloc[0].get("寬度Cm")
        if pd.isna(d1_car_count) or pd.isna(d1_width):
            continue

        if d1_code not in remaining_by_item:
            continue

        d1_candidates = remaining_by_item[d1_code]

        for d1_order in d1_candidates:
            d1_order_id = d1_order["工單號碼"]

            d1_copy = d1_order.copy()
            d1_copy["車數"] = d1_car_count
            d1_copy["刀次"] = 0
            d1_copy["寬度Cm"] = d1_width
            d1_copy["總長度(cm)"] = d1_width * d1_car_count

            main_copy = main_order.copy()
            main_copy["備註配對工單"] = d1_order_id
            d1_copy["備註配對工單"] = main_order_id

            pair_rows_d1.extend([main_copy, d1_copy])

    # 移除已配對的副工單
    paired_d1_order_ids = set(d["工單號碼"] for d in pair_rows_d1 if "工單號碼" in d)
    remaining_df = remaining_df[~remaining_df["工單號碼"].isin(paired_d1_order_ids)].copy()

    # 合併所有配對結果
    paired_df_88 = pd.DataFrame(pair_rows_88)
    paired_df_68 = pd.DataFrame(pair_rows_68)
    paired_df_d1 = pd.DataFrame(pair_rows_d1)
    all_paired_df = pd.concat([df for df in [paired_df_88, paired_df_68, paired_df_d1] if not df.empty], ignore_index=True)

    # 格式化函式
    def format_df(df):
        df = df.rename(columns={
            "開工日期": "預計開工日",
            "預計完工日期": "預計完工日",
            "工單號碼": "工單編號",
            "料號": "品號",
            "開工數量": "預估良品數",
            "寬度Cm": "公分",
            "客戶需求日期": "客戶需求日"
        })

        # 統一欄位型別與補空值
        df["公分"] = pd.to_numeric(df["公分"], errors="coerce").round(1)
        df["預估良品數"] = pd.to_numeric(df["預估良品數"], errors="coerce").fillna(0).astype(int)
        df["車數"] = pd.to_numeric(df["車數"], errors="coerce").fillna(0).astype(int)
        df["刀次"] = pd.to_numeric(df["刀次"], errors="coerce").fillna(0).round(2)
        df["人員"] = ""
        df["餘量"] = df["預估良品數"]
        df["生產註記"] = ""
        
        # 補足缺少的欄位
        for col in ["預計開工日", "人員", "工單編號", "品號", "餘量",
                    "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]:
            if col not in df.columns:
                df[col] = ""

        return df[["預計開工日", "人員", "工單編號", "品號", "餘量",
                    "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]]


    remaining_df = format_df(remaining_df)
    no_pair_df = format_df(no_pair_df)
    if not all_paired_df.empty:
        all_paired_df = format_df(all_paired_df)

    return {
            "remaining_df": remaining_df,
            "no_pair_df": no_pair_df,
            "paired_df": all_paired_df
        }


# 無須配對工單組
# 若工單數量大且不用換料無需配對 一天最多60刀次
def split_no_pair_rows(df_no_pair, max_knife=55):
    rows = []

    for _, row in df_no_pair.iterrows():
        knife = int(row["刀次"])
        cars = float(row["車數"])
        while knife > 0:
            split_knife = min(knife, max_knife)
            knife -= split_knife

            new_row = row.copy()
            new_row["刀次"] = split_knife
            new_row["預估良品數"] = int(cars * split_knife)
            new_row["餘量"] = new_row["預估良品數"]

            rows.append(new_row)

    return pd.DataFrame(rows)


# 配對完成工單
# 先根據配對工單數量先決定好主工單刀次 剩下的再去配對其他工單 目前先他剩下的特別拿出
def split_paired_rows(df_paired):
    df_paired = df_paired.copy().reset_index(drop=True)
    paired_rows = []

    i = 0
    while i < len(df_paired):
        current_row = df_paired.iloc[i]
        current_key = current_row["工單編號"]
        current_car = int(current_row["車數"])
        #current_cut = int(current_row["刀次"])
        current_cut = math.ceil(current_row["刀次"]) # 刀次無條件進位 多切的部分
        total_qty = current_car * current_cut if current_cut > 0 else int(current_row["預估良品數"])

        if current_cut > 0:
            # 主工單處理
            main_row = current_row.copy()
            main_car = current_car
            main_cut = current_cut
            main_qty = main_car * main_cut
            main_remain_qty = main_qty

            j = i + 1
            sub_rows = []

            while j < len(df_paired) and int(df_paired.iloc[j]["刀次"]) == 0:
                sub_rows.append(df_paired.iloc[j].copy())
                j += 1

            for sub_row in sub_rows:
                sub_car = int(sub_row["車數"])
                sub_total_qty = int(sub_row["預估良品數"])

                if sub_car == 0 or sub_total_qty <= 0:
                    continue

                # 子工單可配刀次數
                max_possible_cut = sub_total_qty // sub_car
                usable_cut = min(max_possible_cut, main_remain_qty // main_car)

                if usable_cut <= 0:
                    continue

                used_main_qty = usable_cut * main_car
                used_sub_qty = usable_cut * sub_car

                # 加入主刀
                paired_main = main_row.copy()
                paired_main["刀次"] = usable_cut
                paired_main["預估良品數"] = used_main_qty
                paired_main["餘量"] = used_main_qty
                paired_rows.append(paired_main)

                # 加入子刀
                paired_sub = sub_row.copy()
                paired_sub["刀次"] = 0
                paired_sub["預估良品數"] = used_sub_qty
                paired_sub["餘量"] = used_sub_qty
                paired_rows.append(paired_sub)

                main_remain_qty -= used_main_qty
                if main_remain_qty <= 0:
                    break

            i = j  # 跳過處理過的子刀
        else:
            i += 1

    # ✅ 統計實際被使用掉的數量（不用 usage_map）
    used_df = pd.DataFrame(paired_rows)
    used_summary = used_df.groupby("工單編號")["預估良品數"].sum().to_dict()

    remaining_rows = []
    for _, row in df_paired.iterrows():
        key = row["工單編號"]
        total_qty = int(row["預估良品數"])
        used_qty = used_summary.get(key, 0)
        remain_qty = total_qty - used_qty

        if remain_qty > 0:
            remain_row = row.copy()
            remain_row["預估良品數"] = remain_qty
            remain_row["餘量"] = remain_qty
            #remain_row["刀次"] = 0 
            remaining_rows.append(remain_row)


    return pd.DataFrame(paired_rows), pd.DataFrame(remaining_rows)


# 將剩餘可不用匹配就可以獨立的工單找出 根據基本資料 搭1料號為0或空 , 剩餘數量小於刀次就要強制多切
def process_schedule_data_second(order_df):
    
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得設定json檔案 -> 基本資料路徑
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
  
    # 基本資料路徑
    #base_path = r"E:\\ribbon_schedule\\data\\基本資料-20250617-All.xlsx"
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return

    # 讀取資料
    base_df = pd.read_excel(base_path, dtype=str)

    # 數值處理
    base_df["寬度Cm"] = base_df["寬度Cm"].astype(float)
    base_df["車數"] = base_df["車數"].astype(float)
    base_df["搭1產出車數"] = base_df["搭1產出車數"].fillna(0).astype(float)
    base_df = base_df.rename(columns={"料號": "品號"})

    order_df["品號"] = order_df["品號"].str.strip()
    order_df["預估良品數"] = order_df["預估良品數"].astype(int)
    order_df = order_df.rename(columns={"車數": "車數_error"}) #因為車數不準確

    # 合併資料
    merged = pd.merge(order_df, base_df, how="left", on="品號")
    #print(merged)
    merged["刀次"] = (merged["預估良品數"] / merged["車數"]).round(2)
    merged["總長度(cm)"] = (merged["寬度Cm"] * merged["車數"])

    # 無須配對條件
    '''
    mask_no_pair = (
        merged["工單編號"].str.startswith(("81A", "81B")) &
        (merged["刀次"] % 1 == 0) &
        ((merged["總長度(cm)"] == 88) | (merged["總長度(cm)"] == 89))
    )
    '''
    mask_no_pair = (
        (merged["工單編號"].str.startswith(("81A", "81B")) &
        (merged["搭1產出車數"] == 0) &
        (merged["刀次"] % 1 == 0)
        )
        | 
        ((merged["刀次"] % 1 == 0) & 
         (merged["預估良品數"] == (merged["車數"] * merged["刀次"]).round()) &
         ((merged["總長度(cm)"] == 88) | (merged["總長度(cm)"] == 89) | (merged["總長度(cm)"] == 68))
        )
    )
    no_pair_df = merged[mask_no_pair].copy()
    remaining_df = merged[~mask_no_pair].copy()

    # ⬇️ 拆出整刀邏輯（非整數刀次但可先出整刀）
    split_rows = []
    remaining_updates = []

    for idx, row in remaining_df.iterrows():
        try:
            car = float(row["車數"])
            cut = float(row["刀次"])
            est_qty = int(row["預估良品數"])
            width = float(row["寬度Cm"])
            output_car = float(row["搭1產出車數"])

            if cut % 1 != 0 and output_car == 0:
                #full_cut = int(cut)
                full_cut = math.ceil(cut)
                full_qty = int(car * full_cut)
                remain_qty = est_qty - full_qty

                if full_cut > 0:
                    full_row = row.copy()
                    full_row["刀次"] = full_cut
                    full_row["預估良品數"] = full_qty
                    full_row["餘量"] = full_qty
                    split_rows.append(full_row)

                    remain_row = row.copy()
                    remain_row["刀次"] = round(cut - full_cut, 2)
                    remain_row["預估良品數"] = remain_qty
                    remain_row["餘量"] = remain_qty
                    remaining_updates.append(remain_row)

                    remaining_df.drop(index=idx, inplace=True)
        except:
            continue

    if split_rows:
        no_pair_df = pd.concat([no_pair_df, pd.DataFrame(split_rows)], ignore_index=True)

    if remaining_updates:
        remaining_df = pd.concat([remaining_df, pd.DataFrame(remaining_updates)], ignore_index=True)

    # 格式化欄位
    def format_df(df):
        for col in ["預計開工日", "人員", "工單編號", "品號", "餘量",
                    "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]:
            if col not in df.columns:
                df[col] = ""
        return df[["預計開工日", "人員", "工單編號", "品號", "餘量",
                   "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]]

    remaining_df = format_df(remaining_df)
    remaining_df = remaining_df[remaining_df["預估良品數"].astype(float) >= 0].copy()

    no_pair_df = format_df(no_pair_df)

    # ⬇️ 刀次 > 55 時再拆分
    expanded_rows = []

    for _, row in no_pair_df.iterrows():
        try:
            cut = float(row["刀次"])
            car = float(row["車數"])
            qty = int(row["預估良品數"])

            if cut > 55:
                full_sets = int(cut // 55)
                remaining_cut = round(cut % 55, 2)

                for _ in range(full_sets):
                    new_row = row.copy()
                    new_row["刀次"] = 55
                    new_row["預估良品數"] = int(car * 55)
                    new_row["餘量"] = new_row["預估良品數"]
                    expanded_rows.append(new_row)

                if remaining_cut > 0:
                    new_row = row.copy()
                    new_row["刀次"] = remaining_cut
                    new_row["預估良品數"] = int(car * remaining_cut)
                    new_row["餘量"] = new_row["預估良品數"]
                    expanded_rows.append(new_row)
            else:
                expanded_rows.append(row)
        except:
            expanded_rows.append(row)

    no_pair_df = pd.DataFrame(expanded_rows)

    return {
        "remaining_df": remaining_df,
        "no_pair_df": no_pair_df
    }


def prepare_final_schedule(df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining):
    # 建立主工單欄位
    df_no_pair_split["主工單編號"] = df_no_pair_split["工單編號"]
    df_no_pair_second["主工單編號"] = df_no_pair_second["工單編號"]

    df_paired_split = df_paired_split.copy()
    df_paired_split["主工單編號"] = None

    final_remaining = final_remaining.copy()
    final_remaining["主工單編號"] = None

    # 判斷主子工單
    for idx in df_paired_split.index:
        row = df_paired_split.loc[idx]
        if row["刀次"] != 0:
            df_paired_split.at[idx, "主工單編號"] = row["工單編號"]
        else:
            prev_df = df_paired_split.loc[:idx - 1]
            prev_main = prev_df[prev_df["刀次"] != 0].iloc[-1]
            df_paired_split.at[idx, "主工單編號"] = prev_main["工單編號"]

    final_remaining["刀次"] = pd.to_numeric(final_remaining["刀次"], errors="coerce").fillna(0).astype(int)

    for pos in range(len(final_remaining)):
        row_r = final_remaining.iloc[pos]
        cut_count = row_r["刀次"]

        if cut_count != 0:
            final_remaining.at[final_remaining.index[pos], "主工單編號"] = row_r["工單編號"]
        else:
            prev_df = final_remaining.iloc[:pos]
            prev_valid = prev_df[prev_df["刀次"] != 0]

            if prev_valid.empty:
                final_remaining.at[final_remaining.index[pos], "主工單編號"] = ""
            else:
                final_remaining.at[final_remaining.index[pos], "主工單編號"] = prev_valid.iloc[-1]["工單編號"]


    # 合併資料
    combined_df = pd.concat([df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining], ignore_index=True)

    # 指派固定人員：81B/81C 由 A159 處理
    condition_main = combined_df["主工單編號"].str.startswith(("81B", "81C"))
    combined_df.loc[condition_main, "人員"] = "A159"

    # 過濾需分配的人
    unassigned_df = combined_df[combined_df["人員"].isna() | (combined_df["人員"] == "")]
    main_quantities = (
        unassigned_df
        .groupby("主工單編號")["預估良品數"]
        .sum()
        .reset_index()
        .sort_values(by="預估良品數", ascending=False)
        .reset_index(drop=True)
    )

    # 平均分配給 A830 和 B201（輪流加總，誰少就給誰）
    assign_A830, assign_B201 = [], []
    total_A830, total_B201 = 0, 0

    for _, row in main_quantities.iterrows():
        main_id = row["主工單編號"]
        qty = row["預估良品數"]
        if total_A830 <= total_B201:
            assign_A830.append(main_id)
            total_A830 += qty
        else:
            assign_B201.append(main_id)
            total_B201 += qty

    # 指派人員
    combined_df.loc[combined_df["主工單編號"].isin(assign_A830), "人員"] = "A830"
    combined_df.loc[combined_df["主工單編號"].isin(assign_B201), "人員"] = "B201"

    # 排序：主工單開工日、主工單編號、刀次
    main_orders = combined_df[combined_df["工單編號"] == combined_df["主工單編號"]]
    main_dates = main_orders.groupby("工單編號")["預計開工日"].min()
    combined_df["主工單開工日"] = combined_df["主工單編號"].map(main_dates)

    combined_df = combined_df.sort_values(by=["主工單開工日", "主工單編號", "刀次"], ascending=[True, True, False])

    return combined_df


# 先篩選完再來排
def choose_date(df: pd.DataFrame) -> pd.DataFrame:
    global start_date, end_date

    df = df.copy()
    df["日期篩選用"] = pd.to_datetime(df["開工日期"], errors="coerce")

    root = ctk.CTk()
    root.withdraw()
    dialog = DateFilterDialog(root)
    root.wait_window(dialog)

    start_date = dialog.start_date_choose
    end_date = dialog.end_date_choose

    # ✅ 若使用者未選擇日期就關閉，直接中止
    if not start_date and not end_date:
        raise ValueError("❌ 請選擇日期區間後再繼續")

    # ✅ 執行篩選
    if start_date and end_date:
        df = df[(df["日期篩選用"] >= start_date) & (df["日期篩選用"] <= end_date)]
    elif start_date:
        df = df[df["日期篩選用"] >= start_date]
    elif end_date:
        df = df[df["日期篩選用"] <= end_date]

    df = df.drop(columns=["日期篩選用"])

    return df


# 全部排完才刪選時間 或是 刪除多餘欄位
def remaining_cut_clean(df_remaining_second):
    
    """
    根據主工單完工日篩選 final_output（含主+子），
    根據每列完工日篩選 df_remaining_second。
    """
    '''
    # === 主工單日期篩選用資料 ===
    main_orders = final_output[final_output["工單編號"] == final_output["主工單編號"]].copy()
    main_orders["日期篩選用"] = pd.to_datetime(main_orders["預計完工日"], errors="coerce")

    # === 剩餘工單的日期欄位轉換 ===
    df_remaining_second = df_remaining_second.copy()
    df_remaining_second["日期篩選用"] = pd.to_datetime(df_remaining_second["預計完工日"], errors="coerce")

    # === 彈出選擇視窗 ===
    root = ctk.CTk()
    root.withdraw()
    dialog = DateFilterDialog(root)
    root.wait_window(dialog)

    start_date = dialog.start_date
    end_date = dialog.end_date

    # === 主工單：根據主單的完工日篩選整組 ===
    if start_date or end_date:
        if start_date and end_date:
            valid_main_ids = main_orders[
                (main_orders["日期篩選用"] >= start_date) &
                (main_orders["日期篩選用"] <= end_date)
            ]["工單編號"]
        elif start_date:
            valid_main_ids = main_orders[main_orders["日期篩選用"] >= start_date]["工單編號"]
        elif end_date:
            valid_main_ids = main_orders[main_orders["日期篩選用"] <= end_date]["工單編號"]

        final_output = final_output[final_output["主工單編號"].isin(valid_main_ids)]

    # === 剩餘工單：直接用自己的完工日來篩選 ===
    if start_date:
        df_remaining_second = df_remaining_second[df_remaining_second["日期篩選用"] >= start_date]
    if end_date:
        df_remaining_second = df_remaining_second[df_remaining_second["日期篩選用"] <= end_date]

    # 清掉篩選欄
    final_output = final_output.drop(columns=["日期篩選用"], errors="ignore")
    df_remaining_second = df_remaining_second.drop(columns=["日期篩選用"], errors="ignore")
    '''

    # 移除輔助欄位並回傳
    #final_output = final_output.drop(columns=["主工單開工日", "主工單編號"])

    # 將剩餘sheet的刀次清空
    df_remaining_second["刀次"] = ""

    return df_remaining_second

def do_people(final_df):
    # 確保主工單開工日是 datetime
    final_df["主工單開工日"] = pd.to_datetime(final_df["主工單開工日"], errors='coerce')

    # 根據主工單開工日與主工單編號排序（主工單與子工單會排在一起）
    sorted_df = final_df.sort_values(by=["主工單開工日", "主工單編號", "刀次"], ascending=[True, True, False])

    # 移除輔助欄位並回傳
    #sorted_df = sorted_df.drop(columns=["主工單開工日", "主工單編號"])
    

    # 拆分三位人員
    A159_part = sorted_df[sorted_df["人員"] == "A159"].copy()
    A830_part = sorted_df[sorted_df["人員"] == "A830"].copy()
    B201_part = sorted_df[sorted_df["人員"] == "B201"].copy()

    return A159_part, A830_part, B201_part

def open_file(filepath):
    if platform.system() == 'Windows':
        os.startfile(filepath)
    elif platform.system() == 'Darwin':  # macOS
        subprocess.call(['open', filepath])
    else:  # Linux
        subprocess.call(['xdg-open', filepath])


#--------------------日期排程----------------------------        
def final_schedule_list(A159_part, A830_part, B201_part):
    
    global start_date, end_date

    # 彈出假期選擇視窗
    root = ctk.CTk()
    root.withdraw()
    dialog = MultiPersonLeaveDialog(root, start_date, end_date)  # date 格式
    root.wait_window(dialog)

    if dialog.result is None:
        raise Exception("❌ 使用者取消操作，請重新選擇休假日")

    break_dates = {
        "A159": set(pd.to_datetime(date).date() for date in dialog.result.get("A159", [])),
        "A830": set(pd.to_datetime(date).date() for date in dialog.result.get("A830", [])),
        "B201": set(pd.to_datetime(date).date() for date in dialog.result.get("B201", [])),
    }

    
    schedule_A159 = generate_schedule_for_person(A159_part, break_dates["A159"])
    schedule_A830 = generate_schedule_for_person(A830_part, break_dates["A830"])
    schedule_B201 = generate_schedule_for_person(B201_part, break_dates["B201"])

    start_date = start_date.date()

    
    # 容合
    if not schedule_A830.empty and schedule_A830["刀次"].notna().any() and (schedule_A830["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_A830["刀次"] > 0
        schedule_A830.loc[mask, "預估良品數"] = schedule_A830.loc[mask, "刀次"] * schedule_A830.loc[mask, "車數"]

        # 統一日期格式
        schedule_A830["實際排程日期"] = pd.to_datetime(schedule_A830["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_A830["主工單識別碼"] = schedule_A830["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_A830["排程依據日"] = schedule_A830.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list 用）
        schedule_A830["預計開工日"] = schedule_A830["排程依據日"]
        schedule_A830["預計完工日"] = schedule_A830["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_A830["原始順序"] = schedule_A830.index

        # 按照排程依據日 + 原始順序，確保主子工單一起移動且相對順序不變
        schedule_A830 = schedule_A830.sort_values(
            by=["排程依據日", "原始順序"],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_A830 = final_cal_list(schedule_A830, start_d=start_date, break_dates=break_dates["A830"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_A830["預計開工日"] = pd.to_datetime(schedule_A830["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_A830["預計完工日"] = pd.to_datetime(schedule_A830["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_A830.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單編號", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("A830刀次欄位沒有資料或全部為空，跳過後續處理")

    
    # 家偉
    if not schedule_A159.empty and schedule_A159["刀次"].notna().any() and (schedule_A159["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_A159["刀次"] > 0
        schedule_A159.loc[mask, "預估良品數"] = schedule_A159.loc[mask, "刀次"] * schedule_A159.loc[mask, "車數"]

        # 統一日期格式
        schedule_A159["實際排程日期"] = pd.to_datetime(schedule_A159["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_A159["主工單識別碼"] = schedule_A159["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_A159["排程依據日"] = schedule_A159.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list 用）
        schedule_A159["預計開工日"] = schedule_A159["排程依據日"]
        schedule_A159["預計完工日"] = schedule_A159["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_A159["原始順序"] = schedule_A159.index

        # 按照排程依據日 + 原始順序，確保主子工單一起移動且相對順序不變
        schedule_A159 = schedule_A159.sort_values(
            by=["排程依據日", "原始順序"],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_A159 = final_cal_list(schedule_A159, start_d=start_date, break_dates=break_dates["A159"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_A159["預計開工日"] = pd.to_datetime(schedule_A159["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_A159["預計完工日"] = pd.to_datetime(schedule_A159["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_A159.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單編號", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("A159刀次欄位沒有資料或全部為空，跳過後續處理")
    

    # 旺斌
    if not schedule_B201.empty and schedule_B201["刀次"].notna().any() and (schedule_B201["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_B201["刀次"] > 0
        schedule_B201.loc[mask, "預估良品數"] = schedule_B201.loc[mask, "刀次"] * schedule_B201.loc[mask, "車數"]

        # 統一日期格式
        schedule_B201["實際排程日期"] = pd.to_datetime(schedule_B201["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_B201["主工單識別碼"] = schedule_B201["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_B201["排程依據日"] = schedule_B201.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list 用）
        schedule_B201["預計開工日"] = schedule_B201["排程依據日"]
        schedule_B201["預計完工日"] = schedule_B201["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_B201["原始順序"] = schedule_B201.index

        # 按照排程依據日 + 原始順序，確保主子工單一起移動且相對順序不變
        schedule_B201 = schedule_B201.sort_values(
            by=["排程依據日", "原始順序"],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_B201 = final_cal_list(schedule_B201, start_d=start_date, break_dates=break_dates["B201"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_B201["預計開工日"] = pd.to_datetime(schedule_B201["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_B201["預計完工日"] = pd.to_datetime(schedule_B201["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_B201.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單編號", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("B201刀次欄位沒有資料或全部為空，跳過後續處理")

    

    return schedule_A830, schedule_B201, schedule_A159


def generate_schedule_for_person(df: pd.DataFrame, break_dates: set, max_lookback_days: int = 60) -> pd.DataFrame:

    df = df.copy()
    df["預計開工日"] = pd.to_datetime(df["預計開工日"])
    df["預計完工日"] = pd.to_datetime(df["預計完工日"])
    df["實際排程日期"] = None

    # 每日產能限制（週一到週五）
    daily_limits = {0: 65, 1: 65, 2: 55, 3: 65, 4: 55}
    #daily_limits = {0: 55, 1: 55, 2: 55, 3: 55, 4: 55}

    # 建立主工單索引（主工單：刀次 > 0）
    last_main_idx = None
    df["主工單索引"] = None
    for idx, row in df.iterrows():
        if row["刀次"] > 0:
            last_main_idx = idx
        df.at[idx, "主工單索引"] = last_main_idx

    # 每組用「最早完工日」決定排程範圍
    df["組內最早完工日"] = df.groupby("主工單索引")["預計完工日"].transform("min")

    groups = df.groupby("主工單索引")
    sorted_groups = sorted(groups, key=lambda g: df.loc[g[0], "組內最早完工日"], reverse=True)

    # 建立排程容量表
    latest_end = df["預計完工日"].max()

    earliest_start = latest_end - timedelta(days=max_lookback_days)

    schedule_capacity = {}
    date = latest_end
    while date >= earliest_start:
        if date.weekday() < 5 and date not in break_dates:
            schedule_capacity[date] = daily_limits.get(date.weekday(), 55)
        date -= timedelta(days=1)

    for main_idx, group in sorted_groups:
        group_df = df.loc[group.index]
        earliest_end = group_df["組內最早完工日"].min()
        start_date = earliest_end - timedelta(days=max_lookback_days)

        total_doz = group_df["刀次"].sum()
        remaining = total_doz
        assigned_list = []
        current_date = earliest_end

        while remaining > 0 and current_date >= start_date:
            if current_date in schedule_capacity and schedule_capacity[current_date] > 0:
                available = schedule_capacity[current_date]
                assign = min(available, remaining)
                schedule_capacity[current_date] -= assign
                assigned_list.append((current_date.strftime("%Y-%m-%d"), assign))
                remaining -= assign
            current_date -= timedelta(days=1)

        if remaining > 0:
            raise Exception(f"排程無法完成，工單 {df.loc[main_idx, '工單編號']} 剩餘刀次 {remaining}")

        # 寫入每筆工單的排程日期
        for idx in group.index:
            if idx == main_idx:
                formatted = [f"{d}({c})" for d, c in sorted(assigned_list)]
            else:
                formatted = [d for d, _ in sorted(assigned_list)]
            df.at[idx, "實際排程日期"] = "\n".join(formatted)

    df.drop(columns=["組內最早完工日", "主工單索引"], inplace=True)

    df_new = split_schedule_dates(df)

    return df_new

def split_schedule_dates(df: pd.DataFrame) -> pd.DataFrame:
    new_rows = []
    for _, row in df.iterrows():
        sched_str = row["實際排程日期"]
        total_doz = row["刀次"]
        total_qty = row["需求數量"] if "需求數量" in row else None

        # 若不是字串或沒有 ()，表示沒有分配數量，直接保留
        if not isinstance(sched_str, str) or "(" not in sched_str or total_doz <= 60:
            # 若刀次 ≤ 60 或沒有分配數量資訊，不拆分，只改日期格式
            new_row = row.copy()
            try:
                # 防止錯誤字串，例如含括號
                date_part = str(sched_str).split("(")[0].strip()
                new_row["實際排程日期"] = pd.to_datetime(date_part).strftime("%Y/%m/%d")
            except Exception:
                new_row["實際排程日期"] = None  # 或你可以選擇保留原值
            new_rows.append(new_row)
            continue

        # 拆分多個日期與對應刀次
        lines = sched_str.strip().split("\n")

        for line in lines:
            new_row = row.copy()
            try:
                if "(" in line and ")" in line:
                    # 解析格式如 "2025/07/29(11.0)"
                    date_part = line.split("(")[0].strip()
                    doz_part = float(line.split("(")[1].replace(")", "").strip())

                    ratio = doz_part / total_doz if total_doz else 0
                    qty_part = round(total_qty * ratio) if total_qty is not None else None

                    new_row["刀次"] = doz_part
                    if qty_part is not None:
                        new_row["需求數量"] = qty_part

                    new_row["實際排程日期"] = pd.to_datetime(date_part).strftime("%Y/%m/%d")
                else:
                    # 格式只有日期無括號
                    new_row["實際排程日期"] = pd.to_datetime(line.strip()).strftime("%Y/%m/%d")
            except Exception:
                new_row["實際排程日期"] = None  # 若轉換失敗，設為 None
            new_rows.append(new_row)

    return pd.DataFrame(new_rows)

def final_cal_list(df: pd.DataFrame, start_d: datetime.date, break_dates: set) -> pd.DataFrame:
    df = df.copy()

    # 確保 start_d 是 datetime.date
    if isinstance(start_d, pd.Timestamp):
        start_d = start_d.date()
    elif isinstance(start_d, str):
        start_d = pd.to_datetime(start_d).date()

    # 每日產能限制（週一=0）
    daily_limits = {0: 65, 1: 65, 2: 55, 3: 65, 4: 55}
    capacity_used = {}

    doses = pd.to_numeric(df["刀次"], errors="coerce").fillna(0).tolist()
    new_dates = []
    current_date = start_d

    for dose in doses:
        if dose == 0:
            # 子工單：繼承上一筆日期
            if new_dates:
                new_dates.append(new_dates[-1])
            else:
                new_dates.append(current_date)
            continue

        while True:
            if current_date.weekday() < 5 and current_date not in break_dates:
                used = capacity_used.get(current_date, 0)
                limit = daily_limits.get(current_date.weekday(), 55)
                if used + dose <= limit:
                    capacity_used[current_date] = used + dose
                    new_dates.append(current_date)
                    break
            current_date += timedelta(days=1)

    # 直接用原本的index對應更新日期欄
    df.loc[:, "預計開工日"] = [d.strftime("%Y/%#m/%#d") for d in new_dates]
    df.loc[:, "預計完工日"] = df["預計開工日"]

    return df


class MultiPersonLeaveDialog(ctk.CTkToplevel):
    def __init__(self, parent, start_date, end_date, names=("A159", "A830", "B201")):
        super().__init__(parent)
        self.title("選擇三位人員的休假日")

        width, height = 900, 550
        self.geometry(f"{width}x{height}")
        self.center_window(width, height)

        self.leave_dates = {}
        self.names = names
        self.calendars = {}
        self.date_widgets = {}  # 可以是 label 或 dropdown

        ctk.CTkLabel(self, text="請為三位人員分別選擇休假日", font=("微軟正黑體", 20)).pack(pady=10)

        container = ctk.CTkFrame(self)
        container.pack(pady=10, padx=10, expand=True, fill="both")

        calendar_style = {
            "background": "#004080",
            "foreground": "white",
            "font": ("微軟正黑體", 14),
            "headersbackground": "#004080",
            "headersforeground": "white",
            "normalbackground": "white",
            "weekendbackground": "white",
            "disabledbackground": "#cccccc",
            "bordercolor": "#004080",
            "othermonthforeground": "#999999",
            "selectmode": "day",
            "date_pattern": "yyyy-mm-dd",
        }

        for i, name in enumerate(names):
            frame = ctk.CTkFrame(container)
            frame.grid(row=0, column=i, padx=10, pady=10)

            ctk.CTkLabel(frame, text=name, font=("微軟正黑體", 16)).pack(pady=5)

            cal = Calendar(frame, mindate=start_date, maxdate=end_date, **calendar_style)
            cal.pack(pady=5)
            self.calendars[name] = cal

            ctk.CTkButton(frame, text="加入休假日", font=("微軟正黑體", 14),
                          command=lambda n=name: self.add_date(n)).pack(pady=5)

            label = ctk.CTkLabel(frame, text="已選：無", wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            self.date_widgets[name] = label

            self.leave_dates[name] = set()

        self.error_label = ctk.CTkLabel(self, text="", text_color="red", font=("微軟正黑體", 12))
        self.error_label.pack(pady=5)

        ctk.CTkLabel(self, text="📝 若三人皆無休假，請直接按「✅ 確定」。", font=("微軟正黑體", 14), text_color="gray").pack(pady=(0, 10))

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=15)

        ctk.CTkButton(btn_frame, text="✅ 確定", font=("微軟正黑體", 14), command=self.confirm).pack(side="left", padx=20)
        ctk.CTkButton(btn_frame, text="❌ 取消", font=("微軟正黑體", 14), command=self.cancel).pack(side="left", padx=20)

        self.result = None
        self.grab_set()

    def center_window(self, width, height):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"+{x}+{y}")

    def add_date(self, name):
        selected = self.calendars[name].get_date()
        self.leave_dates[name].add(selected)
        self.update_widget(name)

    def remove_date(self, name, date):
        self.leave_dates[name].discard(date)
        self.update_widget(name)

    def show_delete_menu(self, event, name):
        if not self.leave_dates[name]:
            return

        font = tkfont.Font(family="微軟正黑體", size=18, weight="bold")
        menu = tk.Menu(self, tearoff=0, font=font)

        for date in sorted(self.leave_dates[name]):
            menu.add_command(
                label=f"🗑️　刪除日期：　{date}　",
                command=lambda d=date: self.remove_date(name, d)
            )

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()


    def update_widget(self, name):
        dates = sorted(list(self.leave_dates[name]))

        # 移除原本元件
        widget = self.date_widgets[name]
        widget.pack_forget()

        if len(dates) == 0:
            label = ctk.CTkLabel(widget.master, text="已選：無", wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            self.date_widgets[name] = label

        elif len(dates) <= 2:
            text = "\n".join(dates)
            label = ctk.CTkLabel(widget.master, text=text, wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            label.bind("<Button-1>", lambda e, n=name: self.show_delete_menu(e, n))
            self.date_widgets[name] = label

        else:
            option_menu = ctk.CTkOptionMenu(widget.master, values=dates,
                                            command=lambda d: self.remove_date(name, d),
                                            font=("微軟正黑體", 14),
                                            width=180)
            option_menu.set("休假日列表(點選可刪除)")
            option_menu.pack(pady=5)
            self.date_widgets[name] = option_menu

    def confirm(self):
        self.result = self.leave_dates
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()
        
# 最後分配: 
# 1. 把剩餘工單內料號拿去已經配好的工單篩選若有與"基本資料符合"且"刀次不為0"的資料，則將取其一部份來搭配這原本剩餘工單
# 2. 排序要依照料號 -> 相同料號盡量擺在同一天生產
def final_doAssignAndSort(df_paired_split, df_no_pair_split, df_no_pair_second, remaining):
    # 先找到remaining之中剩餘工單的料號
    # 先去"基本資料"中找到其"搭一料號" 再去三人資料中搜尋是否有符合的
    # 若有符合的就拉其工單來做搭配 如無符合則將工單號留空白補上料號跟數量
    # 將完成的資料一樣排去分給B201, A830
    # 依照料號來重新排列工單 7個工作天相同料號盡量排在同一天生產

    # 基本資料:
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得腳本所在目錄
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    # 基本資料路徑
    #base_path = r"E:\ribbon_schedule\\data\\基本資料-20250617-All.xlsx"
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return

    # 讀取基本資料
    base_df = pd.read_excel(base_path, dtype=str)
    base_df["搭1產出車數"] = pd.to_numeric(base_df["搭1產出車數"], errors="coerce")

    remaining_items = remaining["品號"].unique()  # 取唯一品號清單

    # 過濾基本資料中有在 remaining 的品號
    base_subset = base_df[base_df["料號"].isin(remaining_items)]

    # 建立對照表：{品號: 搭1料號}
    item_to_d1 = dict(zip(base_subset["料號"], base_subset["搭1料號"]))
    #print(item_to_d1)

    # 尋找可以搭配的料號
    all_staff_df = pd.concat([df_paired_split, df_no_pair_split, df_no_pair_second], ignore_index=True)

    # 建立一個空的 DataFrame 來放最終結果
    final_rows = []

    for idx, row in remaining.iterrows():
        original_item = row["品號"]
        original_qty = row["預估良品數"]
        original_order = row.copy()

        d1_code = item_to_d1.get(original_item)
        if not d1_code:
            # 沒有搭1料號，直接加入主工單
            final_rows.append(original_order)
            continue

        candidates = all_staff_df[
            (all_staff_df["品號"] == d1_code) &
            (pd.to_numeric(all_staff_df["刀次"], errors="coerce") > 0)
        ]

        if not candidates.empty:
            paired_row = candidates.iloc[0].copy()

            # 子工單的車數 = 基本資料的搭1產出車數
            # 直接用品號比對取出對應的搭1產出車數
            row_match = base_df[base_df["料號"] == original_order["品號"]]

            if not row_match.empty:
                car_count = pd.to_numeric(row_match.iloc[0]["搭1產出車數"], errors="coerce")
            else:
                car_count = 0  # 找不到就給 0（或視情況 raise 錯誤）

            # 主工單刀次 = 預估良品數 / 車數
            main_car_count = pd.to_numeric(original_order.get("車數", 1), errors="coerce")
            if main_car_count == 0 or pd.isna(main_car_count):
                main_car_count = 1  # 避免除以0
            main_cut_count = math.ceil(pd.to_numeric(original_qty, errors="coerce") / main_car_count)

            # 主工單欄位更新：預計開工日、預計完工日、人員，改成跟子工單一樣
            original_order["預計開工日"] = paired_row.get("預計開工日", original_order.get("預計開工日"))
            original_order["預計完工日"] = paired_row.get("預計完工日", original_order.get("預計完工日"))
            original_order["人員"] = paired_row.get("人員", original_order.get("人員"))

            # 主工單刀次
            original_order["刀次"] = main_cut_count

            # 子工單調整
            paired_row["刀次"] = 0
            paired_row["車數"] = car_count
            paired_row["預估良品數"] = main_cut_count * car_count
            paired_row["餘量"] = paired_row["預估良品數"]

            # 先放主工單，再放搭配的子工單（依序排列）
            final_rows.append(original_order)
            final_rows.append(paired_row)

            # 移除搭配過的工單
            #all_staff_df = all_staff_df[all_staff_df["工單編號"] != paired_row["工單編號"]]

            # 搭配出去的數量（給主工單）
            matched_qty = main_cut_count * car_count  # 主工單實際拿走的良品數

            order_id = paired_row["工單編號"]
            idx_list = all_staff_df.index[all_staff_df["工單編號"] == order_id].tolist()

            if len(idx_list) == 1:
                idx = idx_list[0]
                # 取得原始子工單的車數（不變）
                original_car_count = pd.to_numeric(all_staff_df.at[idx, "車數"], errors="coerce")
                # 取得原始良品數
                original_qty = pd.to_numeric(all_staff_df.at[idx, "預估良品數"], errors="coerce")

                # 計算剩餘良品數
                left_qty = original_qty - matched_qty
                if left_qty < 0:
                    left_qty = 0

                # 計算新的刀次
                new_cut_count = math.ceil(left_qty / original_car_count) if original_car_count > 0 else 0
                new_qty = new_cut_count * original_car_count

                # 更新 all_staff_df
                all_staff_df.at[idx, "預估良品數"] = new_qty
                all_staff_df.at[idx, "餘量"] = new_qty
                all_staff_df.at[idx, "刀次"] = new_cut_count
            else:
                print(f"警告：工單編號 {order_id} 找不到唯一索引，跳過更新")

        else:
            # 沒找到搭配的，直接放主工單
            final_rows.append(original_order)

            # 查找 base_df 對應的搭1料號與搭1產出車數
            row_match = base_df[base_df["料號"] == original_order["品號"]]
            if not row_match.empty:
                d1_code = row_match.iloc[0]["搭1料號"] 
                car_count = pd.to_numeric(row_match.iloc[0]["搭1產出車數"], errors="coerce")
                new_match = base_df[base_df["料號"] == d1_code]
                if not new_match.empty:
                    raw_cm = str(new_match.iloc[0]["寬度Cm"])
                    cm_clean = raw_cm.lower().replace("cm", "").strip()

                    # 轉成 float 再根據是否為整數決定顯示格式
                    cm_value = float(cm_clean)
                    if cm_value.is_integer():
                        new_cm = int(cm_value)   # 顯示為整數
                    else:
                        new_cm = cm_value        # 顯示為小數
                else:
                    new_cm = None
                    raise ValueError(f"在 base_df 中找不到搭1料號 {d1_code} 的公分欄位")

                # 主工單刀次 = 預估良品數 / 主工單車數（向上取整）
                main_car_count = pd.to_numeric(original_order.get("車數", 1), errors="coerce")
                if main_car_count == 0 or pd.isna(main_car_count):
                    main_car_count = 1
                main_cut_count = math.ceil(pd.to_numeric(original_qty, errors="coerce") / main_car_count)

                original_order["刀次"] = main_cut_count
                original_order["預估良品數"] = main_cut_count * main_car_count
                original_order["餘量"] = original_order["預估良品數"]
 
                # 建立子工單（工單號空白）
                new_row = original_order.copy()
                new_row["工單編號"] = ""  # 空白工單
                new_row["品號"] = d1_code
                new_row["公分"] = new_cm
                new_row["車數"] = car_count
                new_row["刀次"] = 0
                new_row["預估良品數"] = main_cut_count * car_count
                new_row["餘量"] = new_row["預估良品數"]
                new_row["客戶需求日"] = ""

                final_rows.append(new_row)

    # 把結果轉成 DataFrame
    final_df = pd.DataFrame(final_rows)

    # 排序，依品號、日期
    # final_df = final_df.sort_values([...])

    # 最後回傳處理後的 DataFrame
    return {
        "all": all_staff_df,
        "final_remaining": final_df
    }


def main():

    # 初階段工單分類 (輸入開始結束日期，根據config找到基本資料)
    result = process_schedule_data()  

    df_remaining = result["remaining_df"]
    df_no_pair = result["no_pair_df"]
    df_paired = result["paired_df"]

    # 無須配對的工單進行分組尚未排程
    df_no_pair_split = split_no_pair_rows(df_no_pair)

    # 配對的工單進行分組尚未排程
    df_paired_split, extra_remaining = split_paired_rows(df_paired)

    # 合併 extra_remaining 並去除重複的工單
    df_remaining = pd.concat([df_remaining, extra_remaining], ignore_index=True)
    df_remaining = df_remaining.drop_duplicates(subset="工單編號", keep="last")

    # 將剩餘可不用匹配就可以獨立的工單找出 根據基本資料 搭1料號為0或空 , 剩餘數量小於刀次就要強制多切
    result_second = process_schedule_data_second(df_remaining)
    df_reamining_second = result_second["remaining_df"]
    df_no_pair_second = result_second["no_pair_df"]
    
    # 沒啥作用 -> 把剩餘工單的刀次改為""
    remaining = remaining_cut_clean(df_reamining_second) 

    # 剩餘工單排序 分配
    result_final = final_doAssignAndSort(df_paired_split, df_no_pair_split, df_no_pair_second, remaining)
    final_remaining = result_final["final_remaining"]
    #all = result_final["all"]

    # 排版,依照預計開工時間排序,尚未加上預計完工時間,初步人員分類
    #final_output = prepare_final_schedule(df_paired_split, df_no_pair_split, df_no_pair_second)
    final_output = prepare_final_schedule(df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining)

    # 工單分配
    A159_part, A830_part, B201_part = do_people(final_output)

    # 日期排程 (一天標準刀次55車，加班可以到65車，滿足出貨不需要安排加班)
    # 輸入休假 彈出視窗 選擇人員以及休假日期
    A830_part_end, B201_part_end, A159_part_end = final_schedule_list(A159_part, A830_part, B201_part)

    #result_final = final_doAssignAndSort(A830_part_end, B201_part_end, A159_part_end, remaining)
    #final_remaining = result_final["final_remaining"]
    #all = result_final["all"]

    # === 寫入 Excel ===
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得目前.py檔的資料夾
    output_name = os.path.join(script_dir, "排程結果_0729.xlsx")   # 存到同一層資料夾
    if os.path.exists(output_name):
        try:
            os.remove(output_name)
        except PermissionError:
            raise Exception(f"Excel 檔案 {output_name} 已開啟，請先關閉後再執行。")
    with pd.ExcelWriter(output_name, engine="openpyxl") as writer:
        #df_paired_split.to_excel(writer, sheet_name="第一次處理後配對", index=False)
        #df_no_pair_split.to_excel(writer, sheet_name="第一次處理無須配對", index=False)
        #df_no_pair_second.to_excel(writer, sheet_name="第二次處理無須配對", index=False)
        A159_part.to_excel(writer, sheet_name="家偉", index=False)
        B201_part.to_excel(writer, sheet_name="旺斌", index=False)
        A830_part.to_excel(writer, sheet_name="容合", index=False)
        A159_part_end.to_excel(writer, sheet_name="家偉_end", index=False)
        B201_part_end.to_excel(writer, sheet_name="旺斌_end", index=False)
        A830_part_end.to_excel(writer, sheet_name="容合_end", index=False)
        #remaining.to_excel(writer, sheet_name="剩餘工單", index=False)
        #final_remaining.to_excel(writer, sheet_name="最後剩餘工單", index=False)

    # Excel 欄寬調整
    def get_display_width(text):
        width = 0
        for ch in str(text):
            width += 4 if unicodedata.east_asian_width(ch) in ('F', 'W', 'A') else 2
        return width

    wb = load_workbook(output_name)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, name="Calibri")

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.font = Font(name="Calibri")

        for col in ws.columns:
            max_len = max((get_display_width(cell.value) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    wb.save(output_name)

    print("End running")

    return output_name

'''
if __name__ == "__main__":
    main()  # 主程式
'''

# 執行 GUI
if __name__ == "__main__":
    app = ScheduleApp()
    app.mainloop()
