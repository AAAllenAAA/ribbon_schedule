'''
Allen
# 每月預排還是只排兩位人員
# 依照換刀數 改變每日每位人員刀次上限
# EX: 整日不用換刀 -> 65 , 換1次刀 -> 60
# MIS電腦可以用新的嗎

初版 55 - 60刀次  
2025/10/07 加入81B, 81C

#加入歷史資料概念 在電子報匯入後同時也匯入歷史資料  
避免出現重複排的問題

#加入global map儲存開工數量
避免出現數量超過的問題
#修正工單是否遺漏，重複，數量太少or超過的問題

#選擇資料區間
#選擇排程開始日(適用每日排程)
#選擇休假日

#加入換線次數與刀次影響table
0:65, 1:55, 2:51, 3:47, 4:43, 5:39

# 每月10號以前 A工單預排
# 每天程式預排 vs 麗如

# 1.B工單來隔天就要先上
# 2.品號相似當天給一人 同公分同車生產
'''
import os
import platform
import subprocess
import sys
import unicodedata
import json
import math
import re
import traceback
import numpy as np
import pandas as pd
import customtkinter as ctk
import tkinter as tk
import tkinter.font as tkfont
from typing import List, Dict, Any, Tuple
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta, date
from tkcalendar import DateEntry, Calendar
from collections import defaultdict
import warnings
import mysql.connector
warnings.filterwarnings("ignore", category=UserWarning)

# 初始化 GUI 風格
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

global start_date, end_date, user_schedule_date


GLOBAL_ORDER_QTY_MAP = {}  # {工單號碼: 剩餘開工數量}

PAIRED_QTY_MAP = {} # key = 工單編號, value = 已配對良品數

def update_global_order_qty_map(df):
    """
    將目前 df 的工單號碼與開工數量儲存為全域 map
    用於後續比對或檢查剩餘可排量
    """
    global GLOBAL_ORDER_QTY_MAP
    GLOBAL_ORDER_QTY_MAP = (
        df.groupby("工單號碼")["開工數量"]
        .sum()
        .to_dict()
    )
    print("✅ GLOBAL_ORDER_QTY_MAP 已更新：")
    for k, v in GLOBAL_ORDER_QTY_MAP.items():
        print(f"  工單 {k} → 剩餘開工 {v}")
    print("=" * 60)
    return GLOBAL_ORDER_QTY_MAP


# 歷史資料 避免排程重複或是數量超過
def load_schedule_history(config: Dict[str, Any]) -> pd.DataFrame:
    """
    讀取上次程式優化後的排程歷史紀錄。
    """
    df_history = pd.DataFrame()
    
    # 讀取 JSON 配置的歷史檔案路徑
    SCHEDULE_HISTORY_FILE = config.get("schedule_history_path")
    if not SCHEDULE_HISTORY_FILE:
        print("警告: config_ribbon.json 缺少 'schedule_history_path' 配置。")
        return df_history
        
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, SCHEDULE_HISTORY_FILE)
    
    if os.path.exists(file_path):
        encodings_to_try = ['utf-8-sig', 'utf-8', 'big5']
        
        for encoding in encodings_to_try:
            try:
                # 嘗試讀取
                df_history = pd.read_csv(file_path, encoding=encoding)
                df_history['工單編號'] = df_history['工單編號'].astype(str).str.strip()
                print(f"✅ 成功載入 {len(df_history)} 筆歷史排程紀錄，使用編碼: {encoding}。")
                return df_history # 載入成功即返回
            except UnicodeDecodeError:
                # 換下一個編碼
                continue
            except Exception as e:
                # 其他非編碼錯誤 (如檔案格式錯誤)
                print(f"警告: 載入歷史檔案失敗 (非編碼錯誤): {e}")
                return pd.DataFrame() # 載入失敗則返回空 DataFrame
        
        # 如果所有編碼都嘗試失敗
        print("警告: 載入歷史檔案失敗: 所有常見編碼 (utf-8-sig, utf-8, big5) 皆無法解析檔案。")
        # --- END: 修正編碼錯誤區塊 ---
        
    return df_history


def save_schedule_history(df: pd.DataFrame, config: Dict[str, Any]):
    """
    將最終優化後的排程結果儲存為歷史紀錄，供下次運行時校準。
    """
    SCHEDULE_HISTORY_FILE = config.get("schedule_history_path")
    if not SCHEDULE_HISTORY_FILE:
        return
        
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, SCHEDULE_HISTORY_FILE)
    
    # 確保只儲存關鍵欄位 (這些欄位是程式上次做的決策)
    cols_to_save = [
        '工單編號', '品號', '餘量', '公分', '車數', '刀次', 
        '預估良品數', '預計開工日', '預計完工日', '人員' 
    ]
    
    df_to_save = df.copy()
    
    # 僅保留有分配到人員的主工單（因為子工單刀次為 0，且子工單的餘量在電子報中會更新）
    # 我們只繼承主工單的「人員」分配，確保換刀優勢。
    df_to_save = df_to_save[df_to_save['人員'].isin(['A159', 'B201', 'A830'])].copy()
    
    # 過濾只保留需要的欄位
    df_to_save = df_to_save[[col for col in cols_to_save if col in df_to_save.columns]]
    
    try:
        # 使用 utf-8-sig 確保 Excel 開啟時中文不亂碼，並使用 mode='w' 覆蓋舊檔案
        df_to_save.to_csv(file_path, index=False, encoding='utf-8-sig')
        print(f"排程結果已儲存到 {SCHEDULE_HISTORY_FILE}，作為下一次的校準基準。")
    except Exception as e:
        print(f"警告: 歷史檔案儲存失敗: {e}")


class ScheduleApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("工單排程工具")
        self.geometry("500x300")

        self.label = ctk.CTkLabel(self, text="碳帶工單排程", font=("Arial", 20))
        self.label.pack(pady=20)

        self.run_button = ctk.CTkButton(self, text="執行排程", command=self.run_schedule)
        self.run_button.pack(pady=10)

        self.output_label = ctk.CTkLabel(self, text="")
        self.output_label.pack(pady=10)

        # 綁定關閉事件
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def run_schedule(self):
        self.run_button.configure(state="disabled")
        self.output_label.configure(text="⏳ 處理中，請稍候...")
        self.update_idletasks()  # 強制更新畫面
        try:
            output_path = main() # 排程主程式
            self.output_label.configure(text="✅ 排程完成！結果已輸出至 Excel")
            self.after(500, lambda: open_file(output_path))

        except Exception as e:
            messagebox.showerror("錯誤", f"執行失敗，請確認資料格式或 Excel 是否已開啟。\n詳細錯誤：{str(e)[:200]}")
            print("🔴 發生錯誤：", e)
            traceback.print_exc()  # 🔍 印出完整錯誤堆疊
            self.output_label.configure(text="❌ 排程失敗")

        finally:
            self.run_button.configure(state="normal")

    def on_closing(self):
        # 這邊可以放你要結束前要做的事情
        self.destroy()   # 關閉視窗
        sys.exit()       # 強制結束程式

# 選擇日期區間彈出視窗
class DateFilterDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("選擇電子報資料日期區間")
        self.geometry("500x360")  # 放大視窗

        self.start_date_choose = None
        self.end_date_choose = None

        # 標題文字
        ctk.CTkLabel(self, text="📅 請選擇日期區間", font=("微軟正黑體", 20, "bold")).pack(pady=(15, 10))

        # 設定共用樣式
        label_font = ("微軟正黑體", 14)
        dateentry_style = {
            "font": ("微軟正黑體", 14),
            "width": 20,
            "background": "#004080",
            "foreground": "white",
            "borderwidth": 2,
            "date_pattern": "yyyy-mm-dd"
        }

        # 起始日期
        ctk.CTkLabel(self, text="起始日期", font=label_font).pack()
        #self.start_entry = DateEntry(self, **dateentry_style)
        self.start_entry = DateEntry(self, year=2025, month=10, day=10, **dateentry_style) # test use
        self.start_entry.pack(pady=(5, 15))

        # 結束日期
        ctk.CTkLabel(self, text="結束日期", font=label_font).pack()
        #self.end_entry = DateEntry(self, **dateentry_style)
        self.end_entry = DateEntry(self, year=2025, month=11, day=30, **dateentry_style) # test use
        self.end_entry.pack(pady=(5, 15))

        # 錯誤訊息
        self.error_label = ctk.CTkLabel(self, text="", text_color="red", font=("微軟正黑體", 12))
        self.error_label.pack(pady=5)

        # 確定按鈕
        ctk.CTkButton(self, text="✅ 確定", font=("微軟正黑體", 14), command=self.confirm).pack(pady=15)

        self.grab_set()  # 鎖定視窗
        self.focus()

        #self.after(100, self.confirm) # 開發使用

    def confirm(self): 
        try:
            self.start_date_choose = datetime.strptime(self.start_entry.get(), "%Y-%m-%d")
            self.end_date_choose = datetime.strptime(self.end_entry.get(), "%Y-%m-%d")
            self.destroy()
        except ValueError:
            self.error_label.configure(text="⚠️ 請選擇有效的日期")



def process_schedule_data():
    # 讓使用者選擇排程資料 Excel
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="請選擇排程資料 Excel 檔案",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not file_path:
        print("❌ 沒有選擇任何檔案，程式結束。")
        return 

    script_dir = os.path.dirname(os.path.abspath(__file__)) 
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
        
    '''
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return
    
    '''
    db_config = config.get("db_config")
    if not db_config:
        print("❌ config_ribbon.json 缺少 db_config 設定")
        return
    
    try:
        conn = mysql.connector.connect(**db_config)
        sql_base_df = "SELECT * FROM pairingrules"  # 你要撈的資料表
        base_df = pd.read_sql(sql_base_df, conn)
        conn.close()
        base_df = base_df.astype(str)  # 保持原本程式邏輯一致
    except Exception as e:
        print(f"❌ 從資料庫讀取基本資料失敗: {e}")
        return
    

    # 讀取資料
    order_df = pd.read_excel(file_path, dtype=str)
    #base_df = pd.read_excel(base_path, dtype=str)

    # 過濾條件
    order_df = order_df[~order_df["工單號碼"].str.startswith(("81R", "81T"))]
    order_df = order_df[order_df["狀態"].str.lower() == "released"] 
    order_df = choose_date(order_df)  # 假設 choose_date 已定義

    
    # rename base_df
    base_df.rename(columns={
        "main_ProductInfo": "料號",
        "mi_SN": "原料材質",
        "mi_Width": "寬度Cm",
        "main_CarNum": "車數",
        "1st_ProductInfo": "搭1料號",
        "1st_CarNum": "搭1產出車數",
        "2nd_ProductInfo": "搭2料號",
        "2nd_CarNum": "搭2產出車數",
        "3th_ProductInfo": "搭3料號",
        "3th_CarNum": "搭3產出車數",
        "4th_ProductInfo": "搭4料號",
        "4th_CarNum": "搭4產出車數",
        "mi_Area": "面積  M2"
    }, inplace=True)
    
    
    # 數值處理
    base_df["寬度Cm"] = pd.to_numeric(base_df["寬度Cm"], errors="coerce").fillna(0)
    base_df["車數"] = pd.to_numeric(base_df["車數"], errors="coerce").fillna(0)
    base_df["搭1產出車數"] = pd.to_numeric(base_df["搭1產出車數"], errors="coerce").fillna(0).astype(float)
    base_df["搭2產出車數"] = pd.to_numeric(base_df["搭2產出車數"], errors="coerce").fillna(0).astype(float)
    order_df["料號"] = order_df["料號"].str.strip()
    order_df["開工數量"] = pd.to_numeric(order_df["開工數量"], errors="coerce").fillna(0).astype(int)

    # 合併資料
    base_df_unique = base_df.drop_duplicates(subset=['料號'], keep='first').copy()
    merged = pd.merge(order_df, base_df_unique, how="left", on="料號")
    merged["總長度(cm)"] = merged["寬度Cm"] * merged["車數"]
    merged["刀次"] = (merged["開工數量"] / merged["車數"]).apply(lambda x: round(x, 2) if x > 0 else 0)

    # ==========================================================
    # 步驟 A: 歷史資料扣除並更新電子報
    # 目的：使用「完工數量」與「昨日已排數量」兩者中的最大值，來扣減「開工數量」，確保不超排。
    # ==========================================================
    df = merged.copy()
    df_history = load_schedule_history(config)

    if not df_history.empty:
        print("歷史資料載入...")

        # 統一欄位名稱
        df_history = df_history.rename(columns={'工單編號': '工單號碼', '品號': '料號'})
        for col in ['工單號碼', '料號']:
            if col in df_history.columns:
                df_history[col] = df_history[col].astype(str).str.strip()
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()

        # 數值欄位
        if '預估良品數' in df_history.columns:
            df_history['預估良品數'] = pd.to_numeric(df_history['預估良品數'], errors='coerce').fillna(0)
        df['開工數量'] = pd.to_numeric(df['開工數量'], errors='coerce').fillna(0)
        df['完工數量'] = pd.to_numeric(df['完工數量'], errors='coerce').fillna(0)

        # 歷史已排數量
        df_history_used_qty = df_history.groupby(['工單號碼', '料號'], as_index=False)['預估良品數'].sum()
        df_history_used_qty = df_history_used_qty.rename(columns={'預估良品數': '昨日已排數量'})
        df = df.merge(df_history_used_qty, on=['工單號碼', '料號'], how='left').fillna(0)

        # ------------------------------------------------------------------
        # ⭐ 修正：使用 np.maximum 確保用最大已處理數量來扣減
        # ------------------------------------------------------------------
        # 1. 找出最大已完成/已排量
        df['最大已處理數量'] = np.maximum(df['完工數量'], df['昨日已排數量'])
        
        # 2. 扣除數量：用原始開工數量減去最大已處理數量，並確保不小於 0
        df['開工數量'] = (df['開工數量'] - df['最大已處理數量']).clip(lower=0)
        
        # 3. 準備列印資訊
        # 將 '最大已處理數量' 暫存為 '昨日已排數量' 以便列印原開工數量
        df['昨日已排數量_PRINT'] = df['最大已處理數量']
        df.drop(columns=['最大已處理數量'], errors='ignore', inplace=True)
        
        # ------------------------------------------------------------------

        # 列印明細
        print("=== 扣除歷史數量後的工單明細 ===")
        for _, row in df.iterrows():
            # 由於我們已用最大值扣減，這裡計算原開工數量時使用 '昨日已排數量_PRINT'
            print(f"工單: {row['工單號碼']}, 料號: {row['料號']}, "
                  f"原開工數量: {row['開工數量'] + row['昨日已排數量_PRINT']}, "
                  f"最大已處理: {row['昨日已排數量_PRINT']}, 剩餘: {row['開工數量']}")
        
        # 移除輔助列印欄位
        df.drop(columns=['昨日已排數量', '昨日已排數量_PRINT'], errors='ignore', inplace=True)


        # 已完成工單
        completed_orders = df[df['開工數量'] <= 0]['工單號碼'].unique()
        if len(completed_orders) > 0:
            print("=== 已完成工單 ===")
            print(completed_orders)

        # 移除完成工單
        df = df[df['開工數量'] > 0].copy()

        # 排除備註引用已完成工單
        if '備註' in df.columns:
            def check_exclude(note):
                if pd.isna(note):
                    return False
                note_prefix = str(note)[:8]
                return note_prefix in completed_orders
            df['exclude_by_note'] = df['備註'].apply(check_exclude)
            excluded_count = df['exclude_by_note'].sum()
            if excluded_count > 0:
                print(f"排除 {excluded_count} 筆備註引用已完成工單的工單")
            df = df[df['exclude_by_note'] == False].copy()
            df.drop(columns=['exclude_by_note'], inplace=True)

    # === 更新全域工單數量 map ===
    update_global_order_qty_map(df)
    
    # 最後整理
    df = df.loc[:, ~df.columns.duplicated()]
    print(f"-> 總計 {len(df)} 筆工單進入排程優化。")
    merged = df.copy()

    # -------------------------------
    # 無須配對條件
    mask_no_pair = (
        merged["工單號碼"].str.startswith(("81A", "81B")) &
        (merged["搭1產出車數"] == 0) &
        (merged["刀次"] % 1 == 0)
    )
    no_pair_df = merged[mask_no_pair].copy().reset_index(drop=True)
    remaining_df = merged[~mask_no_pair].copy().reset_index(drop=True)

    # -------------------------------
    # 88cm 配對
    pair_rows_88 = []
    df_81c_remarks = merged[(merged["工單號碼"].str.startswith("81C")) & (merged["備註"].notna())]
    for _, c_order in df_81c_remarks.iterrows():
        # 🌟 修正點 1: 從備註中提取乾淨的工單 ID 🌟
        raw_remark = str(c_order["備註"]).strip()
        match = re.match(r"(81[A-Z]\d{5})", raw_remark) # 匹配 "81XNNNNN" 格式
        
        if match:
            pair_order_id = match.group(1) # 取得乾淨的 ID，例如 "81A03766"
        else:
            continue # 無效 ID，跳過此 $81C$ 工單
            
        # pair_order_id = c_order["備註"]  <-- 移除或註解掉原本的行
        matched_rows = remaining_df[remaining_df["工單號碼"] == pair_order_id]
        if matched_rows.empty:
            continue
        
        a_order = matched_rows.iloc[0]
        for cars in range(1, 8):
            total_cm = a_order["總長度(cm)"] + c_order["寬度Cm"] * cars
            if abs(total_cm - 88) <= 1:
                a_copy = a_order.copy()
                c_copy = c_order.copy()
                c_copy["車數"] = cars
                c_copy["總長度(cm)"] = c_copy["寬度Cm"] * cars
                c_copy["刀次"] = 0
                pair_rows_88.extend([a_copy, c_copy])
                remaining_df = remaining_df[~remaining_df["工單號碼"].isin([pair_order_id, c_order["工單號碼"]])]
                break

    # -------------------------------
    # 68cm 配對
    pair_rows_68 = []
    df_81c_remain = remaining_df[(remaining_df["工單號碼"].str.startswith("81C")) & (remaining_df["備註"].notna())]
    for _, c_order in df_81c_remain.iterrows():
        
        # 🌟 修正點 2: 從備註中提取乾淨的工單 ID 🌟
        raw_remark = str(c_order["備註"]).strip()
        match = re.match(r"(81[A-Z]\d{5})", raw_remark) # 匹配 "81XNNNNN" 格式
        
        if match:
            pair_order_id = match.group(1) # 取得乾淨的 ID
        else:
            continue # 無效 ID，跳過此 $81C$ 工單
            
        # pair_order_id = c_order["備註"]  <-- 移除或註解掉原本的行
        matched_rows = remaining_df[remaining_df["工單號碼"] == pair_order_id]
        if matched_rows.empty:
            continue
        
        a_order = matched_rows.iloc[0]
        for cars in range(1, 5):
            total_cm = a_order["總長度(cm)"] + c_order["寬度Cm"] * cars
            if abs(total_cm - 68) <= 1:
                a_copy = a_order.copy()
                c_copy = c_order.copy()
                c_copy["車數"] = cars
                c_copy["總長度(cm)"] = c_copy["寬度Cm"] * cars
                c_copy["刀次"] = 0
                pair_rows_68.extend([a_copy, c_copy])
                remaining_df = remaining_df[~remaining_df["工單號碼"].isin([pair_order_id, c_order["工單號碼"]])]
                break

    # -------------------------------
    # 搭1料號配對（改為支援同一料號多筆 base 設定，每筆 candidate 要求所有子料號同時存在）
    pair_rows_d1 = []

    # 這樣在迭代 rem_rows 時，能優先處理較早開工的工單
    remaining_df["開工日期"] = pd.to_datetime(remaining_df["開工日期"], errors='coerce')
    remaining_df = remaining_df.sort_values(by="開工日期", ascending=True).reset_index(drop=True)

    # 準備 remaining_by_item：料號 -> list of rows（此為配對池）
    remaining_by_item = {}
    for _, row in remaining_df.iterrows():
        remaining_by_item.setdefault(row["料號"], []).append(row)

    # 建構 candidates：主料號 -> list of candidate dicts（每筆 base_df 的一列為一個 candidate）
    # candidate structure: { "child_codes": [...], "child_out_cars": [...], "source_row": base_row }
    candidates = {}
    for _, b_row in base_df.iterrows():
        main = str(b_row.get("料號", "")).strip()
        if not main:
            continue
        child_codes = []
        child_outs = []
        # 依原本 base_df 欄位順序把搭1/搭2視為可能子料號（某一列可能只有搭1或只有搭2或同時有）
        if "搭1料號" in b_row and pd.notna(b_row.get("搭1料號")) and str(b_row.get("搭1料號")).strip() != "":
            child_codes.append(str(b_row.get("搭1料號")).strip())
            try:
                child_outs.append(int(float(b_row.get("搭1產出車數", 1))))
            except:
                child_outs.append(1)
        if "搭2料號" in b_row and pd.notna(b_row.get("搭2料號")) and str(b_row.get("搭2料號")).strip() != "":
            child_codes.append(str(b_row.get("搭2料號")).strip())
            try:
                child_outs.append(int(float(b_row.get("搭2產出車數", 1))))
            except:
                child_outs.append(1)
        if not child_codes:
            continue
        cand = {
            "child_codes": child_codes,
            "child_outs": child_outs,
            "source_row": b_row
        }
        candidates.setdefault(main, []).append(cand)

    # 對 remaining_df 中每一個主工單（iteration 使用 copy 以免修改迭代物件），嘗試 candidate（貪婪：第一個可完全滿足 candidate）
    # 我們使用 used_order_ids 來記錄被配走的工單（主 + 子），並從 remaining_by_item 與 remaining_df 移除
    used_order_ids = set()
    # 將 remaining_df 的索引列轉成 list 以固定 iteration 行為
    rem_rows = list(remaining_df.to_dict('records'))

    for main_order in rem_rows:
        main_order_id = main_order["工單號碼"]
        main_item_code = main_order["料號"]

        # 若主工單已被其他配對使用，跳過
        if main_order_id in used_order_ids:
            continue

        # 無 candidate 則跳過
        if main_item_code not in candidates:
            continue

        # 逐 candidate 嘗試
        paired_this_main = False
        for cand in candidates[main_item_code]:
            child_codes = cand["child_codes"]
            child_outs = cand["child_outs"]

            matched_child_rows = []
            can_use = True

            # 對 candidate 的每一個 child_code，嘗試在 remaining_by_item 中找第一筆尚未被使用且不是 main_order 本身的工單
            for idx, ccode in enumerate(child_codes):
                found = None
                if ccode in remaining_by_item:
                    for cand_row in remaining_by_item[ccode]:
                        cid = cand_row["工單號碼"]
                        if cid not in used_order_ids and cid != main_order_id:
                            found = cand_row
                            break
                if found is None:
                    can_use = False
                    break
                matched_child_rows.append((found, child_outs[idx] if idx < len(child_outs) else 1))

            if not can_use:
                continue  # 嘗試下一 candidate

            # 若所有 child_code 都找到對應可用工單，則建立配對（主 + 所有子）
            main_copy = main_order.copy()
            
            # --------------------------
            # 🌟 【重要修正點】校正主工單的車數、刀次和預估良品數 🌟
            # --------------------------
            rule_row = cand["source_row"]
            try:
                # 1. 取得這條規則的主工單車數
                # 確保 new_main_car 是一個浮點數
                new_main_car = float(rule_row.get("車數", main_copy.get("車數", 1)))
                
                # 2. 覆蓋主工單的車數
                main_copy["車數"] = new_main_car
                
                # 3. 重新計算主工單的刀次和預估良品數
                # 使用 '開工數量' 作為計算刀次的基礎量
                main_qty = float(main_copy.get("開工數量", 0)) 
                
                if new_main_car > 0 and main_qty > 0:
                    # 計算刀次：向上取整
                    new_main_cut = math.ceil(main_qty / new_main_car)
                    # 預估良品數：以 (新刀次 * 新車數) 計算
                    new_main_qty = new_main_cut * new_main_car
                else:
                    new_main_cut = 0
                    new_main_qty = 0
                
                main_copy["刀次"] = new_main_cut
                
                # 增加 debug 輸出以確認
                print(f"🔧 校正主工單 {main_order_id} ({main_item_code})：")
                print(f"   車數: {main_order['車數']} → {new_main_car}")
                print(f"   刀次: {main_order['刀次']} → {new_main_cut}")
                print(f"   良品數: {main_order['開工數量']} → {new_main_qty} (新刀次 * 新車數)")
                
            except Exception as e:
                print(f"⚠️ 校正主工單車數/刀次時發生錯誤 ({main_order_id})：{e}，使用原始值。")
                # 確保 '預估良品數' 至少使用原始的 '開工數量'

            # --------------------------


            # 設定備註配對工單為逗號連接的子工單號（保留 traceability）
            main_copy["備註配對工單"] = ",".join([r["工單號碼"] for r, _ in matched_child_rows])
            pair_rows_d1.append(main_copy)
            used_order_ids.add(main_order_id)

            # 加入子工單（並依 candidate 的 out car 設定車數 & 計算總長度）
            for child_row, out_car in matched_child_rows:
                child_copy = child_row.copy()
                # 車數以 candidate 出車數為主（防呆）
                try:
                    child_copy["車數"] = int(out_car) if out_car and int(out_car) > 0 else child_copy.get("車數", 1)
                except:
                    child_copy["車數"] = child_copy.get("車數", 1)
                # 取得子料號的寬度：照之前邏輯從 base_df 找，如果找不到則保留 child_row 的寬度
                child_base_info = base_df[base_df["料號"] == child_copy["料號"]]
                if not child_base_info.empty:
                    try:
                        child_width = float(child_base_info.iloc[0].get("寬度Cm", child_copy.get("寬度Cm", 0)))
                    except:
                        child_width = child_copy.get("寬度Cm", 0)
                else:
                    child_width = child_copy.get("寬度Cm", 0)
                child_copy["寬度Cm"] = child_width
                try:
                    child_copy["總長度(cm)"] = float(child_copy.get("寬度Cm", 0)) * float(child_copy.get("車數", 0))
                except:
                    child_copy["總長度(cm)"] = 0
                child_copy["刀次"] = 0
                child_copy["備註配對工單"] = main_order_id

                pair_rows_d1.append(child_copy)
                used_order_ids.add(child_copy["工單號碼"])

            # 從 remaining_by_item 中移除已被使用的子工單紀錄（避免再次被配到）
            for child_row, _ in matched_child_rows:
                code = child_row["料號"]
                remaining_by_item[code] = [r for r in remaining_by_item.get(code, []) if r["工單號碼"] not in used_order_ids]
            # 同時也移除主工單（若它出現在某個料號的列表中）
            if main_item_code in remaining_by_item:
                remaining_by_item[main_item_code] = [r for r in remaining_by_item.get(main_item_code, []) if r["工單號碼"] not in used_order_ids]

            paired_this_main = True
            break  # 此主工單已配到，跳出 candidate 迴圈（採貪婪第一配對）

        # end for each candidate

    # end for each main_order

    # 依舊照你原本的邏輯，將配對結果轉成 DataFrame 並從 remaining_df 中移除已配走的工單
    paired_d1_order_ids = set(d["工單號碼"] for d in pair_rows_d1 if "工單號碼" in d)
    # 加上被當作子工單而使用的工單也移除（used_order_ids 有包含）
    remaining_df = remaining_df[~remaining_df["工單號碼"].isin(used_order_ids)].copy()

    # -------------------------------
    # 合併所有配對結果
    paired_df_88 = pd.DataFrame(pair_rows_88)
    paired_df_68 = pd.DataFrame(pair_rows_68)
    paired_df_d1 = pd.DataFrame(pair_rows_d1)
    all_paired_df = pd.concat([df for df in [paired_df_88, paired_df_68, paired_df_d1] if not df.empty], ignore_index=True)

    # ==========================================================
    # 搭2配對（根據備註欄位，再去確認搭1或搭2料號）
    # ==========================================================
    pair_rows_d2 = []
    paired_main_ids = set()  # 記錄有被搭2配對的主工單

    for _, row in remaining_df.iterrows():
        remark_order_id = row.get("備註")
        if pd.isna(remark_order_id) or remark_order_id == "":
            continue

        # 去 all_paired_df 找主工單
        main_match = all_paired_df[
            (all_paired_df["工單號碼"] == remark_order_id) & 
            (all_paired_df["刀次"] != 0)
        ]
        if main_match.empty:
            continue

        main_order = main_match.iloc[0]
        main_item_code = main_order["料號"]

        # 取 base_df 看搭1或搭2料號是否有這筆料號
        base_info = base_df[base_df["料號"] == main_item_code]
        if base_info.empty:
            continue

        for col in ["搭1料號", "搭2料號"]:
            paired_code = base_info.iloc[0].get(col)
            if pd.isna(paired_code) or paired_code == "":
                continue
            if paired_code != row["料號"]:
                continue

            # 符合條件，建立搭2工單
            paired_copy = row.copy()
            paired_car_count = base_info.iloc[0].get(f"{col}產出車數", 0)
            if pd.isna(paired_car_count) or paired_car_count == 0:
                paired_car_count = 1  # 防呆
            paired_copy["車數"] = paired_car_count
            paired_copy["總長度(cm)"] = paired_copy["寬度Cm"] * paired_car_count
            paired_copy["刀次"] = 0
            paired_copy["備註配對工單"] = main_order["工單號碼"]

            pair_rows_d2.append(paired_copy)
            paired_main_ids.add(main_order["工單號碼"])

            print(f"✅ 搭2配對：工單 {paired_copy['工單號碼']} 配對到主工單 {main_order['工單號碼']} "
                f"料號 {paired_copy['料號']} 車數 {paired_car_count}")

    # 移除 remaining_df 中已配對的工單
    paired_d2_ids = set(d["工單號碼"] for d in pair_rows_d2 if "工單號碼" in d)
    remaining_df = remaining_df[~remaining_df["工單號碼"].isin(paired_d2_ids)].copy()

    # ==========================================================
    # 將搭2工單插入到對應主工單和子工單之後
    # ==========================================================
    for paired_copy in pair_rows_d2:
        main_id = paired_copy["備註配對工單"]

        # 找出主工單在 all_paired_df 的位置
        main_idx_list = all_paired_df.index[
            (all_paired_df["工單號碼"] == main_id) & (all_paired_df["刀次"] != 0)
        ].tolist()

        if not main_idx_list:
            all_paired_df = pd.concat(
                [all_paired_df, pd.DataFrame([paired_copy])],
                ignore_index=True
            )
            print(f"⚠️ 找不到主工單 {main_id}，直接附加搭2工單 {paired_copy['工單號碼']}")
            continue

        main_idx = main_idx_list[0]

        # 找出這個主工單相關的所有子工單
        related_idx = all_paired_df.index[
            all_paired_df["備註配對工單"] == main_id
        ].tolist()

        # 插入位置：主工單之後，或主工單 + 子工單之後
        insert_pos = max(related_idx + main_idx_list) + 1 if related_idx else main_idx + 1

        # 插入搭2工單
        top = all_paired_df.iloc[:insert_pos]
        bottom = all_paired_df.iloc[insert_pos:]
        all_paired_df = pd.concat(
            [top, pd.DataFrame([paired_copy]), bottom],
            ignore_index=True
        )

        print(f"✅ 插入搭2工單 {paired_copy['工單號碼']} 到主工單 {main_id} 的主工單之後 (位置 {insert_pos})")

    # ==========================================================
    # 針對有被搭2配對的主工單 → 往下兩筆做校正
    # ==========================================================
    for main_id in paired_main_ids:
        main_idx_list = all_paired_df.index[
            (all_paired_df["工單號碼"] == main_id) & (all_paired_df["刀次"] != 0)
        ].tolist()
        if not main_idx_list:
            continue

        main_idx = main_idx_list[0]
        main_row = all_paired_df.loc[main_idx]
        main_item_code = main_row["料號"]
        main_cut = main_row["刀次"]

        base_info = base_df[base_df["料號"] == main_item_code]
        if base_info.empty:
            continue

        d1_code = str(base_info.iloc[0].get("搭1料號", "")).strip()
        d2_code = str(base_info.iloc[0].get("搭2料號", "")).strip()
        d1_car = base_info.iloc[0].get("搭1產出車數", 1)
        d2_car = base_info.iloc[0].get("搭2產出車數", 1)

        # 從主工單往下兩筆
        sub_indices = list(range(main_idx + 1, main_idx + 3))
        for sub_idx in sub_indices:
            if sub_idx >= len(all_paired_df):
                continue

            sub_row = all_paired_df.loc[sub_idx]
            sub_code = str(sub_row["料號"]).strip()

            if sub_code == d1_code:
                car = d1_car
                which = "搭1"
            elif sub_code == d2_code:
                car = d2_car
                which = "搭2"
            else:
                which = "非搭1/搭2"
                continue  # 跳過不匹配的

            all_paired_df.at[sub_idx, "車數"] = car
            all_paired_df.at[sub_idx, "總長度(cm)"] = sub_row["寬度Cm"] * car

            print(f"🔧 校正主工單 {main_id} 的子工單 {sub_row['工單號碼']} → {which}, 車數={car}, 總長度={sub_row['寬度Cm']*car}")


    # -------------------------------
    # 格式化函式
    def format_df(df):
        if not isinstance(df, pd.DataFrame) or df.empty:
            return pd.DataFrame(columns=["預計開工日", "人員", "工單編號", "品號", "餘量",
                                         "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"])
        df = df.copy()
        df.rename(columns={
            "開工日期": "預計開工日",
            "預計完工日期": "預計完工日",
            "工單號碼": "工單編號",
            "料號": "品號",
            "開工數量": "預估良品數",
            "寬度Cm": "公分",
            "客戶需求日期": "客戶需求日"
        }, inplace=True)
        for col in ["公分", "預估良品數", "車數", "刀次"]:
            if col in df.columns and isinstance(df[col], pd.Series):
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            elif col not in df.columns:
                df[col] = 0
        for col in ["人員", "餘量", "生產註記"]:
            if col not in df.columns:
                df[col] = ""
        df["餘量"] = df["預估良品數"]
        return df[["預計開工日", "人員", "工單編號", "品號", "餘量",
                   "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]]

    remaining_df = format_df(remaining_df)
    no_pair_df = format_df(no_pair_df)
    if not all_paired_df.empty:
        all_paired_df = format_df(all_paired_df)

    # 防呆 check paired_df內 是否有重複出現 譬如3722配3723 又有 3723配3722之類的
    # 

    return {
        "paired_df_88": paired_df_88,
        "paired_df_68": paired_df_68,
        "paired_df_d1": paired_df_d1,
        "remaining_df": remaining_df,
        "no_pair_df": no_pair_df,
        "paired_df": all_paired_df,
        "df_history": df_history,
        "merged": merged,
        "base_df": base_df
    }


# 無須配對工單組
# 若工單數量大且不用換料無需配對 一天最多60刀次
def split_no_pair_rows(df_no_pair, max_knife=60):
    rows = []

    for _, row in df_no_pair.iterrows():
        cars = float(row["車數"])
        good_qty = float(row["預估良品數"])
        
        # 計算刀次（無條件進位）
        knife = math.ceil(good_qty / cars)

        remaining_knife = knife
        while remaining_knife > 0:
            split_knife = min(remaining_knife, max_knife)
            remaining_knife -= split_knife

            new_row = row.copy()
            new_row["刀次"] = split_knife
            # 根據拆分刀次計算預估良品數（可能最後一刀不足整數倍）
            new_row["預估良品數"] = int(min(split_knife * cars, good_qty))
            # 更新剩餘數量
            good_qty -= new_row["預估良品數"]
            new_row["餘量"] = new_row["預估良品數"]

            rows.append(new_row)

    return pd.DataFrame(rows)


# 配對完成工單
# 先根據配對工單數量先決定好主工單刀次 剩下的再去配對其他工單 目前先他剩下的特別拿出
def split_paired_rows(df_paired, max_cut=60, debug=False):
    """
    改良版 split_paired_rows
    - 以子工單實際餘量限制主工單可配刀次（依車數換算）
    - 已配成功的主/子段落只放入 paired_rows
    - 最後用 used_df 的實際使用量計算 remaining（original - used）
    - debug=True 可印出配對過程
    """

    df_paired = df_paired.copy().reset_index(drop=True)
    paired_rows = []
    remaining_rows = []

    # 原始量 map (工單 -> 原始預估良品數)
    original_qty_map = df_paired.set_index("工單編號")["預估良品數"].astype(float).to_dict()

    i = 0
    while i < len(df_paired):
        current_row = df_paired.iloc[i]
        current_cut = math.ceil(float(current_row["刀次"]))
        current_car = int(current_row["車數"])
        main_id = current_row["工單編號"]

        if current_cut > 0:
            main_row = current_row.copy()

            # 收集後面子工單（刀次 = 0）
            sub_rows = []
            j = i + 1
            while j < len(df_paired) and int(df_paired.iloc[j]["刀次"]) == 0:
                sub_rows.append(df_paired.iloc[j].copy())
                j += 1

            remaining_cut = current_cut
            has_paired = False  # 標記這支主工單是否至少配到過一次

            # 用子工單的「餘量」欄若存在，優先使用；若沒有，使用原始預估良品數
            for s in sub_rows:
                if "餘量" not in s.index:
                    s["餘量"] = float(s["預估良品數"])

            while remaining_cut > 0:
                # 計算每個子工單能支援的最大主刀次 (floor(餘量 / 子車數))
                max_sub_cut = remaining_cut
                total_sub_capacity_cuts = 0  # sum of each sub's floor(餘量 / sub_car)
                for s in sub_rows:
                    sub_car = int(s["車數"])
                    sub_remain_qty = float(s.get("餘量", s["預估良品數"]))
                    if sub_car <= 0:
                        continue
                    max_cut_for_sub = math.floor(sub_remain_qty / sub_car)
                    # 為了讓主刀次能被所有子支持，取最小值
                    max_sub_cut = min(max_sub_cut, max_cut_for_sub)
                    total_sub_capacity_cuts += max_cut_for_sub

                # split_cut 為本輪實際可拆的刀次（受最大分段與子支援限制）
                split_cut = min(max_cut, max_sub_cut)

                if split_cut <= 0:
                    # 子工單不足以支援一個完整 split_cut（可能部分可配）
                    # 若 total_sub_capacity_cuts > 0，代表可以配部分（以 cuts 為單位）
                    if total_sub_capacity_cuts > 0:
                        # partial_cut 為實際能配的刀次（以 cuts 為單位）
                        partial_cut = min(remaining_cut, total_sub_capacity_cuts)
                        if debug:
                            print(f"⚠️ 子工單不足：主 {main_id} 只能配部分 {partial_cut} / {remaining_cut} 刀次")
                        # 生成主工單配對段
                        main_piece = main_row.copy()
                        main_piece["刀次"] = partial_cut
                        main_piece["預估良品數"] = partial_cut * current_car
                        main_piece["餘量"] = main_piece["預估良品數"]
                        paired_rows.append(main_piece)
                        has_paired = True

                        # 向各子工單分配實際數量（以量為單位）
                        for s in sub_rows:
                            sub_car = int(s["車數"])
                            if sub_car <= 0:
                                continue
                            sub_remain_qty = float(s.get("餘量", s["預估良品數"]))
                            alloc_qty = min(sub_remain_qty, partial_cut * sub_car)
                            if alloc_qty <= 0:
                                continue
                            sub_piece = s.copy()
                            sub_piece["刀次"] = 0
                            sub_piece["預估良品數"] = alloc_qty
                            sub_piece["餘量"] = sub_remain_qty - alloc_qty
                            paired_rows.append(sub_piece)
                            # 更新原始 sub_rows 的餘量（影響後續分配）
                            s["餘量"] = sub_piece["餘量"]
                            if debug:
                                print(f"  ↳ 子 {s['工單編號']} 配 {alloc_qty}，剩 {s['餘量']}")

                        # 更新 remaining_cut：把已配的刀次扣掉
                        remaining_cut -= partial_cut

                        # 剩下仍然可能 >0，但因為子資源已耗盡（或不夠），會在之後由總表計算 remaining
                        if remaining_cut > 0 and debug:
                            print(f"  ➕ 主 {main_id} 尚有 {remaining_cut} 刀次待排（會放入剩餘）")
                    else:
                        # 完全無法配任何量，直接結束這支主工單（剩餘量交給最後統計）
                        if debug:
                            print(f"❌ 子工單完全無法配，主 {main_id} 的所有剩餘刀次都會成為剩餘")
                        # 不在此處把剩餘 append 到 remaining_rows，以避免重複
                    break  # 結束這支主工單的配對迴圈

                # 正常情況下，split_cut > 0，可以一次配這段
                if debug:
                    print(f"✅ 主 {main_id} 配 {split_cut} 刀 (量={split_cut * current_car})")

                main_piece = main_row.copy()
                main_piece["刀次"] = split_cut
                main_piece["預估良品數"] = split_cut * current_car
                main_piece["餘量"] = main_piece["預估良品數"]
                paired_rows.append(main_piece)
                has_paired = True

                # 子工單同步分配
                for s in sub_rows:
                    sub_car = int(s["車數"])
                    if sub_car <= 0:
                        continue
                    sub_remain_qty = float(s.get("餘量", s["預估良品數"]))
                    alloc_qty = min(sub_remain_qty, split_cut * sub_car)
                    if alloc_qty <= 0:
                        continue
                    sub_piece = s.copy()
                    sub_piece["刀次"] = 0
                    sub_piece["預估良品數"] = alloc_qty
                    sub_piece["餘量"] = sub_remain_qty - alloc_qty
                    paired_rows.append(sub_piece)
                    s["餘量"] = sub_piece["餘量"]
                    if debug:
                        print(f"  ↳ 子 {s['工單編號']} 配 {alloc_qty}，剩 {s['餘量']}")

                # 扣掉已配的刀次，繼續處理 remaining_cut
                remaining_cut -= split_cut

            # 完成此主工單（不在此處直接 append remaining；最後統一以 used_df 計算 remaining）
            i = j
        else:
            # 這筆本來就是子工單（沒配對用的主工單），先放到 remaining_rows
            # （會在最後合併 remaining，但這裡放也可，最終會用 original-used 決定）
            remaining_rows.append(current_row.copy())
            i += 1

    # === 第二階段：檢查主工單是否超量（保持原有邏輯） ===
    used_df = pd.DataFrame(paired_rows).reset_index(drop=True)

    if not used_df.empty:
        used_summary = used_df.groupby("工單編號")["預估良品數"].sum().to_dict()
    else:
        used_summary = {}

    for key, total_used in used_summary.items():
        original_qty = float(original_qty_map.get(key, 0))
        if total_used > original_qty:
            over_qty = total_used - original_qty
            mask = (used_df["工單編號"] == key) & (used_df["刀次"].astype(float) > 0)
            main_pieces = used_df.loc[mask].reset_index()
            for idx in main_pieces.index[::-1]:
                piece_qty = float(main_pieces.at[idx, "預估良品數"])
                piece_car = float(main_pieces.at[idx, "車數"])
                if over_qty <= 0:
                    break
                if piece_qty <= over_qty:
                    drop_idx = main_pieces.at[idx, "index"]
                    used_df.drop(drop_idx, inplace=True)
                    over_qty -= piece_qty
                else:
                    new_qty = piece_qty - over_qty
                    new_cut = math.ceil(new_qty / piece_car) if piece_car > 0 else 0
                    update_idx = main_pieces.at[idx, "index"]
                    used_df.at[update_idx, "預估良品數"] = new_qty
                    used_df.at[update_idx, "刀次"] = new_cut
                    used_df.at[update_idx, "餘量"] = new_qty
                    over_qty = 0

    used_df = used_df.reset_index(drop=True)

    # === 生成 remaining：以 original - used 的結果為準（確保不會重複） ===
    # 計算每張工單被 used_df 使用的量
    used_summary_after = {}
    if not used_df.empty:
        used_summary_after = used_df.groupby("工單編號")["預估良品數"].sum().to_dict()

    # 以 original_qty_map 為基準，扣掉 used_summary_after -> 剩餘放入 remaining_rows
    for key, orig_qty in original_qty_map.items():
        used_qty = float(used_summary_after.get(key, 0.0))
        remain_qty = orig_qty - used_qty
        if remain_qty > 0:
            remain_row = df_paired.loc[df_paired["工單編號"] == key].iloc[0].copy()
            remain_row["預估良品數"] = remain_qty
            remain_row["餘量"] = remain_qty
            # 為避免重複，先檢查是否已存在類似的 remaining（可依工單編號合併）
            remaining_rows.append(remain_row)

    # === 移除 used_df 中「孤立子工單被誤放走」的情況：保留主+其後所有子（不做 prev_cut 的錯誤判斷） ===
    # 之前的 prev_cut 判斷會誤把部分成功配對的子工單放到 remaining，
    # 現在 paired_rows 只包含實際配對成功的子/主段落，used_df 已正確代表配對結果，
    # 因此 remaining 由 original-used 決定即可，不再進行 prev_cut 的掃描。

    used_df = used_df.reset_index(drop=True)
    remaining_df = pd.DataFrame(remaining_rows).reset_index(drop=True)

    # === 防呆檢查區段（刀次=0 且品號前三碼不同） ===
    if not used_df.empty and "品號" in used_df.columns:
        invalid_rows = []
        rows_to_move = []  # 記錄要移動的 index

        for i in range(1, len(used_df)):
            prev_row = used_df.iloc[i - 1]
            curr_row = used_df.iloc[i]

            # 條件：連續兩筆刀次都為 0 且品號前三碼不同
            if int(prev_row["刀次"]) == 0 and int(curr_row["刀次"]) == 0:
                prev_part = str(prev_row["品號"])
                curr_part = str(curr_row["品號"])

                if prev_part[:3] != curr_part[:3]:
                    invalid_rows.append({
                        "idx": i,
                        "前一筆工單": prev_row["工單編號"],
                        "前一筆品號": prev_part,
                        "當前工單": curr_row["工單編號"],
                        "當前品號": curr_part
                    })
                    rows_to_move.append(i)

        if invalid_rows:
            print("⚠️ 防呆警告：發現異常資料（連續刀次=0 且品號前三碼不相同），將移至 remaining_df:")
            for r in invalid_rows:
                print(f"  ➜ index={r['idx']} | 前:{r['前一筆工單']}({r['前一筆品號']}) → 現:{r['當前工單']}({r['當前品號']})")

            # 移動異常資料
            rows_to_move = sorted(set(rows_to_move))
            move_rows = used_df.iloc[rows_to_move].copy()

            remaining_df = pd.concat([remaining_df, move_rows], ignore_index=True)
            used_df = used_df.drop(index=rows_to_move).reset_index(drop=True)

    if debug:
        print("\n=== split_paired_rows debug summary ===")
        print(f"paired_rows (used_df) count: {len(used_df)}")
        print(f"remaining_rows count: {len(remaining_df)}\n")

    # === 🔹 校正：同天連續兩筆子工單（刀次=0、工單編號相同） ===
    duplicate_indices = []
    move_rows = []

    i = 1
    while i < len(used_df):
        prev = used_df.iloc[i - 1]
        curr = used_df.iloc[i]

        # 條件：連續兩筆都是子工單（刀次=0），同工單編號、同預計開工日
        if (
            int(prev["刀次"]) == 0
            and int(curr["刀次"]) == 0
            and prev["工單編號"] == curr["工單編號"]
            and str(prev.get("預計開工日", "")) == str(curr.get("預計開工日", ""))
        ):
            # Step 1️⃣ 找主工單：往上找最近一筆刀次>0且同開工日
            main_cut = 0
            for k in range(i - 1, -1, -1):
                if float(used_df.iloc[k]["刀次"]) > 0:
                    main_cut = float(used_df.iloc[k]["刀次"])
                    break

            if main_cut is None:
                main_cut = 0

            # Step 2️⃣ 合併兩筆總量
            total_qty = float(prev["預估良品數"]) + float(curr["預估良品數"])

            # Step 3️⃣ 根據主工單刀次重新計算該保留筆的正確數量
            expected_qty = float(prev["車數"]) * main_cut
            print("float(prev[車數])",float(prev["車數"]))
            print("主工單刀次", main_cut)
            remain_qty = max(0, total_qty - expected_qty)

            # Step 4️⃣ 更新第一筆為正確值
            used_df.at[i - 1, "預估良品數"] = expected_qty
            used_df.at[i - 1, "餘量"] = expected_qty
            print("expected_qty", expected_qty)

            # Step 5️⃣ 將第二筆改成剩餘量丟去 remaining_df
            curr_fixed = curr.copy()
            curr_fixed["預估良品數"] = remain_qty
            curr_fixed["餘量"] = remain_qty

            move_rows.append(curr_fixed)
            duplicate_indices.append(i)

            if debug:
                print(f"⚙️ 修正重複子工單 {curr['工單編號']}（日期 {curr.get('預計開工日', '')}）：")
                print(f"   ▸ 主刀次={main_cut}")
                print(f"   ▸ 合併總量={total_qty} → 保留={expected_qty}, 剩餘={remain_qty}")

            # 跳過這組
            i += 1
        i += 1

    # Step 6️⃣ 移除重複的子工單
    if duplicate_indices:
        used_df = used_df.drop(index=duplicate_indices).reset_index(drop=True)
        remaining_df = pd.concat([remaining_df, pd.DataFrame(move_rows)], ignore_index=True)
    

    return used_df, remaining_df


# 延續split_paired_rows的Step 1 分配結束後再將資料做一次數量校正 
# 不會有remaining部份
def split_paired_rows_step1(df_paired, remaining_rows=None, max_cut=60):
    df_paired = df_paired.copy().reset_index(drop=True)
    paired_rows = []

    # 確保剩餘是 list
    if remaining_rows is None:
        new_remaining_rows = []
    elif isinstance(remaining_rows, pd.DataFrame):
        new_remaining_rows = remaining_rows.to_dict('records')  # 轉成 list of dict
    else:
        new_remaining_rows = remaining_rows.copy()  # list

    i = 0
    while i < len(df_paired):
        current_row = df_paired.iloc[i]
        current_cut = math.ceil(float(current_row["刀次"]))
        current_car = int(current_row["車數"])

        if current_cut > 0:
            main_row = current_row.copy()
            sub_rows = []

            j = i + 1
            while j < len(df_paired) and int(df_paired.iloc[j]["刀次"]) == 0:
                sub_rows.append(df_paired.iloc[j].copy())
                j += 1

            remaining_cut = current_cut
            sub_remaining = {idx: int(row["預估良品數"]) for idx, row in enumerate(sub_rows)}

            while remaining_cut > 0:
                split_cut = min(max_cut, remaining_cut)

                main_piece = main_row.copy()
                main_piece["刀次"] = split_cut
                main_piece["預估良品數"] = split_cut * current_car
                main_piece["餘量"] = main_piece["預估良品數"]
                paired_rows.append(main_piece)

                for idx, sub_row in enumerate(sub_rows):
                    sub_car = int(sub_row["車數"])
                    if sub_car == 0:
                        continue
                    sub_piece = sub_row.copy()
                    sub_piece["刀次"] = 0
                    sub_piece["預估良品數"] = split_cut * sub_car
                    sub_piece["餘量"] = sub_piece["預估良品數"]
                    paired_rows.append(sub_piece)
                    sub_remaining[idx] -= sub_piece["預估良品數"]

                remaining_cut -= split_cut

            # 主工單剩餘量
            if remaining_cut > 0:
                leftover_row = main_row.copy()
                leftover_row["刀次"] = remaining_cut
                leftover_row["預估良品數"] = remaining_cut * current_car
                leftover_row["餘量"] = leftover_row["預估良品數"]
                new_remaining_rows.append(leftover_row.to_dict())

            # 子工單剩餘量
            for idx, remain_qty in sub_remaining.items():
                if remain_qty > 0:
                    leftover_sub = sub_rows[idx].copy()
                    leftover_sub["刀次"] = 0
                    leftover_sub["預估良品數"] = remain_qty
                    leftover_sub["餘量"] = remain_qty
                    new_remaining_rows.append(leftover_sub.to_dict())

            i = j
        else:
            new_remaining_rows.append(current_row.to_dict())
            i += 1

    return pd.DataFrame(paired_rows), pd.DataFrame(new_remaining_rows)




# 將剩餘可不用匹配就可以獨立的工單找出 根據基本資料 搭1料號為0或空 , 剩餘數量小於刀次就要強制多切
def process_schedule_data_second(order_df, base_df):
    
    base_df["寬度Cm"] = base_df["寬度Cm"].astype(float)
    base_df["車數"] = base_df["車數"].astype(int)
    base_df["搭1產出車數"] = base_df["搭1產出車數"].fillna(0).astype(float)
    base_df = base_df.rename(columns={'料號':'品號'})

    # 3️⃣ 前處理 order_df
    order_df["品號"] = order_df["品號"].str.strip()
    order_df["預估良品數"] = order_df["預估良品數"].astype(int)
    if "車數" in order_df.columns:
        order_df = order_df.drop(columns=["車數"])

    # 4️⃣ 對齊車數
    merged = pd.merge(
        order_df,
        base_df[["品號", "車數", "寬度Cm", "搭1產出車數"]],
        how="left",
        on="品號"
    )

    # 5️⃣ 計算刀次、總長度
    merged["刀次"] = (merged["預估良品數"] / merged["車數"]).apply(math.ceil)
    merged["總長度(cm)"] = merged["車數"] * merged["寬度Cm"]

    # 6️⃣ 無須配對條件挑出 no_pair_df
    mask_no_pair = (
        (merged["工單編號"].str.startswith(("81A", "81B")) &
         (merged["搭1產出車數"] == 0) &
         (merged["刀次"] % 1 == 0))
        |
        ((merged["刀次"] % 1 == 0) &
         (merged["預估良品數"] == (merged["車數"] * merged["刀次"]).round()) &
         (merged["總長度(cm)"].isin([68, 88, 89])))
    )

    no_pair_df = merged[mask_no_pair].copy()
    remaining_df = merged[~mask_no_pair].copy()

    # 7️⃣ 拆分刀次>55
    def split_large_cut(df):
        rows = []
        for _, row in df.iterrows():
            cut = int(row["刀次"])
            car = int(row["車數"])
            if cut > 55:
                full_sets = cut // 55
                remain_cut = cut % 55
                for _ in range(full_sets):
                    new_row = row.copy()
                    new_row["刀次"] = 55
                    new_row["預估良品數"] = car * 55
                    new_row["餘量"] = new_row["預估良品數"]
                    rows.append(new_row)
                if remain_cut > 0:
                    new_row = row.copy()
                    new_row["刀次"] = remain_cut
                    new_row["預估良品數"] = car * remain_cut
                    new_row["餘量"] = new_row["預估良品數"]
                    rows.append(new_row)
            else:
                # 重新確保 預估良品數 = 車數 × 刀次
                row["預估良品數"] = car * cut
                row["餘量"] = row["預估良品數"]
                rows.append(row)
        return pd.DataFrame(rows)

    no_pair_df = split_large_cut(no_pair_df)

    # === 8️⃣ remaining_df 也同步確保刀次取整並重算預估良品數 ===
    remaining_df["刀次"] = remaining_df["刀次"].apply(lambda x: int(math.ceil(x)))
    remaining_df["預估良品數"] = remaining_df["車數"] * remaining_df["刀次"]
    remaining_df["餘量"] = remaining_df["預估良品數"]

    # === 9️⃣ 欄位排序 ===
    output_columns = [
        "預計開工日", "人員", "工單編號", "品號", "餘量",
        "公分", "車數", "刀次", "預估良品數",
        "預計完工日", "生產註記", "客戶需求日"
    ]

    def format_df(df):
        for col in output_columns:
            if col not in df.columns:
                df[col] = ""
        return df[output_columns]

    no_pair_df = format_df(no_pair_df)
    remaining_df = format_df(remaining_df)

    # === 🔍 Debug印出確認 ===
    print("🔍 無須配對工單明細（最終計算後）：")
    print(no_pair_df[["工單編號", "品號", "車數", "刀次", "預估良品數"]].to_string(index=False))
    print("✅ 所有刀次已無條件進位，預估良品數已重新計算為 車數×刀次。\n")

    return {
        "remaining_df": remaining_df,
        "no_pair_df": no_pair_df
    }



def get_item_base(item_no):
    """
    從品號中提取基礎品號（前4個字母），作為減少換刀次數的依據。
    """
    if pd.isna(item_no):
        return 'UNKNOWN'
    
    item_str = str(item_no).strip()
    # 嘗試匹配開頭的4個英文字母
    match = re.match(r'([A-Za-z]{4})', item_str)
    
    # 確保只返回前四個字母並轉為大寫
    return match.group(1).upper() if match else 'UNKNOWN'

def pair_final_remaining(remaining, base_df, take_first_scheme_only=True):

    """
    自動生成搭料子工單 (v3 版本)
    ---------------------------------
    ✅ 子工單緊貼主工單 (A → A子 → B → B子)
    ✅ 子工單公分從 base_df 查搭料品號的寬度
    ✅ 預估良品數 & 餘量 = 主工單刀次 × 子工單車數
    ✅ 不修改原始 remaining 結構
    """
    # --- 初始化與前置檢查 ---
    remaining = remaining.copy()
    base_df = base_df.copy()
    base_df = base_df.rename(columns={'料號': '品號'})
    ITEM_COL = '品號'

    # 數字轉換
    for col in ['車數', '刀次']:
        if col in remaining.columns:
            remaining[col] = pd.to_numeric(remaining[col], errors='coerce').fillna(0).astype(int)
    if '車數' in base_df.columns:
        base_df['車數'] = pd.to_numeric(base_df['車數'], errors='coerce').fillna(0).astype(int)

    # --- 1️⃣ 取寬度表 ---
    width_map_df = base_df.drop_duplicates(subset=[ITEM_COL], keep='first')[[ITEM_COL, '寬度Cm']].copy()

    # --- 2️⃣ 選擇方案 ---
    if take_first_scheme_only:
        base_df = base_df.drop_duplicates(subset=[ITEM_COL], keep='first').reset_index(drop=True)

    # --- 3️⃣ 生成搭料對照表 ---
    sub_cols = [c for c in base_df.columns if c.startswith('搭') and ('料號' in c or '品號' in c)]
    all_sub_mappings = []

    for _, row in base_df.iterrows():
        main_item = row[ITEM_COL]
        car_count = row['車數']
        scheme_id = f"{main_item}_方案{row.name+1}"
        for i, sub_col in enumerate(sub_cols):
            qty_col = sub_col.replace('料號', '產出車數').replace('品號', '產出車數')
            if qty_col not in row:
                continue
            sub_item = row[sub_col]
            sub_qty = row[qty_col] if pd.notna(row[qty_col]) else 0
            sub_qty = pd.to_numeric(row[qty_col], errors='coerce')
            if pd.notna(sub_item) and sub_qty > 0:
                all_sub_mappings.append({
                    '品號': main_item,
                    '主工單車數': car_count,
                    '搭料品號': sub_item,
                    '搭料車數': int(sub_qty),
                    '方案代號': scheme_id
                })

    if not all_sub_mappings:
        print("💡 無搭料資料，跳過生成。")
        return remaining

    base_mapping = pd.DataFrame(all_sub_mappings)
    print(f"✅ 搭料方案載入 {len(base_mapping)} 筆（每主工單僅選一方案）。")

    # --- 4️⃣ 主工單清單 ---
    mains = remaining[remaining['刀次'] != 0][['工單編號', ITEM_COL, '車數', '刀次']].drop_duplicates()

    # --- 5️⃣ 比對主工單與搭料方案 ---
    merged = mains.merge(
        base_mapping,
        left_on=[ITEM_COL, '車數'],
        right_on=['品號', '主工單車數'],
        how='inner'
    )
    if merged.empty:
        print("💡 無符合搭料。")
        return remaining

    # --- 6️⃣ 生成子工單 ---
    subs = []
    for _, p in merged.iterrows():
        main_row = remaining.loc[remaining['工單編號'] == p['工單編號']].iloc[0].copy()
        sub = main_row.copy()
        sub['工單編號'] = ''
        sub[ITEM_COL] = p['搭料品號']
        sub['車數'] = int(p['搭料車數'])
        sub['刀次'] = 0
        sub['來源類型'] = f"{p['方案代號']}"
        sub['主工單編號'] = p['工單編號']

        # 公分
        matched_width = width_map_df.loc[width_map_df[ITEM_COL] == p['搭料品號'], '寬度Cm']
        sub['公分'] = matched_width.iloc[0] if not matched_width.empty else None

        # 預估良品 / 餘量
        sub['預估良品數'] = main_row['刀次'] * sub['車數']
        sub['餘量'] = main_row['刀次'] * sub['車數']

        subs.append((p['工單編號'], sub))

    # --- 7️⃣ 插入子工單於主工單後 ---
    new_rows = []
    for _, row in remaining.iterrows():
        new_rows.append(row)
        for main_sn, sub_row in subs:
            if row['工單編號'] == main_sn:
                new_rows.append(sub_row)

    remaining = pd.DataFrame(new_rows)[remaining.columns].reset_index(drop=True)
    print(f"🎯 已生成 {len(subs)} 筆子工單（每主工單僅選第一方案）。")

    return remaining


def prepare_final_schedule(all, final_remaining, df_history, base_df):
    
    # 建立主工單欄位
    all = all.copy()
    all["主工單編號"] = None

    final_remaining = final_remaining.copy()
    final_remaining["主工單編號"] = None

    # ----------------------------------------------------------
    # ⭐ 新增前置作業1：在合併前過濾 final_remaining 中的超額工單
    # ----------------------------------------------------------
    if "GLOBAL_ORDER_QTY_MAP" in globals():
        print("🔄 合併前置作業：檢查 final_remaining 數量是否已超額...")

        # 1. 計算 all_df (已處理工單池) 中每筆工單的已分配數量總和
        # 這裡需要計算所有類型的工單總和，因為 all_df 已經包含了 NEWLY_PAIRED/ORIGINAL_PAIRED/NO_PAIR_LEFT
        all["預估良品數"] = pd.to_numeric(all["預估良品數"], errors='coerce').fillna(0)
        assigned_qty_in_all = all.groupby('工單編號')['預估良品數'].sum()

        sns_to_drop_from_remaining = []

        # 2. 迭代檢查 final_remaining 中的工單
        for sn in final_remaining['工單編號'].unique():
            original_qty = GLOBAL_ORDER_QTY_MAP.get(sn)
            if original_qty is None:
                continue

            # 該工單在 all_df 中已經累積的數量
            accumulated_qty = assigned_qty_in_all.get(sn, 0)

            # 該工單在 final_remaining 中的數量 (因為 final_remaining 應該是原始行，假設每行代表該工單全部或一個批次)
            # 由於 final_remaining 在 cleanup_all 階段後應只包含未配對的原始行，我們用它的原始預估良品數來判斷。
            remaining_rows = final_remaining[final_remaining['工單編號'] == sn]
            
            # 由於 final_remaining 裡的行**理論上**都是上一步驟中未被 NEWLY_PAIRED 的**原始行**，
            # 每個工單編號只會有一筆資料，其 '預估良品數' 欄位應為其總需求量。
            # 如果有多筆資料，我們檢查累積數量是否已經超過總需求。

            # 假設 final_remaining 裡是 cleanup_all 步驟 2 留下的、尚未被消耗的原始行
            for idx, row in remaining_rows.iterrows():
                row_qty_series = pd.Series(row['預估良品數'])
                row_qty = pd.to_numeric(row_qty_series, errors='coerce').fillna(0).iloc[0]
                
                # 如果 all_df 中的累積數量 >= 該工單的總需求量
                if accumulated_qty >= original_qty:
                    # 該工單在 all_df 中已被滿足，final_remaining 中的任何對應行都應被刪除
                    sns_to_drop_from_remaining.append(idx)
                    
                # 這裡不需要處理 accumulated_qty + row_qty > original_qty 的情況，
                # 因為 all_df 的清理（cleanup_all）階段已經處理了剩餘批次超額的問題。
                # 這裡主要目的是防止將已被 all_df 完全滿足的**原始行**重複帶入。

        # 3. 執行過濾
        if sns_to_drop_from_remaining:
            removed_count = len(sns_to_drop_from_remaining)
            final_remaining = final_remaining.drop(sns_to_drop_from_remaining).reset_index(drop=True)
            print(f"✅ 合併前置作業完成: 從 final_remaining 移除 {removed_count} 筆在 all_df 中已滿足需求的工單。")
        else:
            print("✅ 合併前置作業完成: 無需移除 final_remaining 中的工單。")
    else:
        print("⚠️ 找不到 GLOBAL_ORDER_QTY_MAP，略過 final_remaining 的數量檢查。")
 
    # ----------------------------------------------------------
    # ⭐ 新增前置作業2：final_remaining 中的工單先配對空工單
    # ----------------------------------------------------------
    final_remaining = pair_final_remaining(final_remaining, base_df, take_first_scheme_only=True)

    # 判斷主子工單
    for pos in range(len(all)):
        row = all.iloc[pos]
        if row["刀次"] != 0:
            all.at[all.index[pos], "主工單編號"] = row["工單編號"]
        else:
            prev_df = all.iloc[:pos]
            prev_main = prev_df[prev_df["刀次"] != 0]
            if prev_main.empty:
                all.at[all.index[pos], "主工單編號"] = None
            else:
                all.at[all.index[pos], "主工單編號"] = prev_main.iloc[-1]["工單編號"]

    final_remaining["刀次"] = pd.to_numeric(final_remaining["刀次"], errors='coerce').fillna(0).astype(int)

    for pos in range(len(final_remaining)):
        row_r = final_remaining.iloc[pos]
        cut_count = row_r["刀次"]
        if cut_count != 0:
            final_remaining.at[final_remaining.index[pos], "主工單編號"] = row_r["工單編號"]
        else:
            prev_df = final_remaining.iloc[:pos]
            prev_valid = prev_df[prev_df["刀次"] != 0]
            if prev_valid.empty:
                final_remaining.at[final_remaining.index[pos], "主工單編號"] = ""
            else:
                final_remaining.at[final_remaining.index[pos], "主工單編號"] = prev_valid.iloc[-1]["工單編號"]

    
    # 合併資料
    combined_df = pd.concat([all, final_remaining], ignore_index=True)

    # ==========================================================
    # 🎯 重新計算/校準：主工單編號 + 唯一主工單ID (修正版)
    # ==========================================================
    combined_df["主工單編號"] = None
    # 創建一個新的、唯一的群組ID
    combined_df["唯一主工單ID"] = None
    last_main_id = None # 用來追蹤最近的主工單行索引

    for pos in range(len(combined_df)):
        row = combined_df.iloc[pos]
        
        # 修正：安全地取出刀次值，避免 'int' object has no attribute 'fillna' 錯誤
        刀次值 = pd.to_numeric(pd.Series(row["刀次"], dtype='object'), errors='coerce').fillna(0).iloc[0]
        
        if 刀次值 != 0:
            # 情況一：這是主工單
            unique_id = f"{row['工單編號']}_{combined_df.index[pos]}" # 使用工單編號 + 行索引作為唯一ID
            
            combined_df.at[combined_df.index[pos], "主工單編號"] = row["工單編號"]
            combined_df.at[combined_df.index[pos], "唯一主工單ID"] = unique_id
            last_main_id = unique_id
            
        else:
            # 情況二：這是子工單/搭料
            if last_main_id is not None:
                # 子工單綁定到最近一次出現的主工單
                main_id = last_main_id.split('_')[0] # 提取工單編號部分
                
                combined_df.at[combined_df.index[pos], "主工單編號"] = main_id 
                combined_df.at[combined_df.index[pos], "唯一主工單ID"] = last_main_id
            # else: 保持 None

    # ==========================================================
    # 🎯 數量校準邏輯 (移除多餘的映射，直接使用 '唯一主工單ID' 映射)
    # ==========================================================

    # 確保數值欄位是數字
    combined_df["刀次"] = pd.to_numeric(combined_df["刀次"], errors="coerce").fillna(0)
    combined_df["車數"] = pd.to_numeric(combined_df["車數"], errors="coerce").fillna(1)


    # 1. 建立唯一映射表 (使用 '唯一主工單ID' 建立映射表) <--- 沿用第一個邏輯
    main_info = combined_df[combined_df["刀次"] != 0].set_index("唯一主工單ID")[["刀次", "車數"]]
    main_info.columns = ["主工單刀次", "主工單車數"]

    # 2. 映射主工單的刀次和車數到所有行 (使用 '唯一主工單ID' 進行映射) <--- 沿用第一個邏輯
    combined_df["主工單刀次"] = combined_df["唯一主工單ID"].map(main_info["主工單刀次"])
    combined_df["主工單車數"] = combined_df["唯一主工單ID"].map(main_info["主工單車數"])

    # 3. 計算並校準數量

    # 創建一個計算結果欄位
    combined_df["計算數量"] = 0

    # a. 主工單行 (刀次 != 0)
    main_mask = combined_df["刀次"] != 0
    # 主工單數量 = 自己的刀次 * 自己的車數
    combined_df.loc[main_mask, "計算數量"] = combined_df.loc[main_mask, "刀次"] * combined_df.loc[main_mask, "車數"]

    # b. 子工單行 (刀次 == 0)
    sub_mask = combined_df["刀次"] == 0
    # 子工單數量 = 主工單刀次 * 子工單車數 (這裡的 '主工單刀次' 是來自 '唯一主工單ID' 的映射，所以是正確的)
    combined_df.loc[sub_mask, "計算數量"] = combined_df.loc[sub_mask, "主工單刀次"] * combined_df.loc[sub_mask, "車數"]

    # 4. 寫回目標欄位
    combined_df["預估良品數"] = combined_df["計算數量"].fillna(0)
    combined_df["餘量"] = combined_df["計算數量"].fillna(0)

    # 5. 清理輔助欄位
    combined_df.drop(columns=["計算數量", "主工單刀次", "主工單車數"], inplace=True) 


    # ==========================================================
    # 防呆檢查：確認 GLOBAL_ORDER_QTY_MAP 與實際加總是否一致，必要時自動補足
    # ==========================================================
    try:
        if "GLOBAL_ORDER_QTY_MAP" in globals():
            print("\n🔍 開始檢查 GLOBAL_ORDER_QTY_MAP 與實際數據差異...\n")

            # 確保數值欄位是數字
            combined_df["預估良品數"] = pd.to_numeric(combined_df["預估良品數"], errors="coerce").fillna(0)
            combined_df["刀次"] = pd.to_numeric(combined_df["刀次"], errors="coerce").fillna(0)
            combined_df["車數"] = pd.to_numeric(combined_df["車數"], errors="coerce").fillna(1)
            combined_df["餘量"] = pd.to_numeric(combined_df.get("餘量", 0), errors="coerce").fillna(0)

            for order_id, planned_qty in GLOBAL_ORDER_QTY_MAP.items():
                planned_qty = float(planned_qty)

                # 計算 combined_df 的總餘量
                total_actual = combined_df.loc[combined_df["工單編號"] == order_id, "餘量"].sum()

                if total_actual < planned_qty:
                    diff = planned_qty - total_actual
                    print(f"⚠️ 工單 {order_id} 開工數量 {planned_qty:.0f}，實際加總僅 {total_actual:.0f}（差 {diff:.0f}）")

                    # 找出該工單在 combined_df 的資料
                    df_target = combined_df[combined_df["工單編號"] == order_id]

                    # 過濾刀次不為0的主工單
                    candidate_df = df_target[df_target["刀次"] > 0]

                    if candidate_df.empty:
                        print(f"   → 找不到刀次不為0的工單 {order_id}，無法自動補足。")
                        continue

                    # 取第一筆作為補充主工單
                    candidate_idx = candidate_df.index[0]
                    car_count = combined_df.loc[candidate_idx, "車數"]
                    add_cut = math.ceil(diff / car_count)
                    add_qty = add_cut * car_count

                    # 更新欄位
                    combined_df.loc[candidate_idx, "刀次"] += add_cut
                    combined_df.loc[candidate_idx, "預估良品數"] += add_qty
                    combined_df.loc[candidate_idx, "餘量"] += add_qty

                    print(f"   ✅ 已自動補足工單 {order_id}：增加 {add_cut} 刀（約 {add_qty} 件）")

                elif total_actual > planned_qty:
                    diff = total_actual - planned_qty
                    print(f"⚠️ 工單 {order_id} 開工數量 {planned_qty:.0f}，實際加總超過 {total_actual:.0f}（超過 {diff:.0f}）")

                    # 顯示超量的每筆工單明細
                    df_target = combined_df[combined_df["工單編號"] == order_id]
                    for idx, row in df_target.iterrows():
                        print(f"   → 工單 {row['工單編號']}，刀次 {row['刀次']}, 車數 {row['車數']}, 預估良品數 {row['預估良品數']}")

        else:
            print("⚠️ 找不到 GLOBAL_ORDER_QTY_MAP，略過防呆檢查。")

    except Exception as e:
        print(f"❌ 防呆檢查時發生錯誤：{e}")


    # ==========================================================
    # 先依歷史資料綁定人員
    # ==========================================================
    if df_history is not None and not df_history.empty:
        if '工單號碼' not in df_history.columns or '人員' not in df_history.columns or '刀次' not in df_history.columns:
            raise ValueError("歷史資料缺少 '工單號碼'、'人員' 或 '刀次' 欄位，無法進行精確判斷。")
        
        # 1. 篩選歷史資料：只保留歷史上刀次 != 0 的工單 (即歷史主工單)
        df_history["刀次_num"] = pd.to_numeric(df_history["刀次"], errors='coerce').fillna(0)
        df_main_history = df_history[df_history["刀次_num"] != 0].copy()
        
        # 2. 建立歷史人員映射表
        # 注意：使用 .drop_duplicates 確保同一個工單編號只對應一個歷史人員（如果歷史資料有重複）
        history_mapping_df = df_main_history.drop_duplicates(subset=['工單號碼'], keep='first')
        history_mapping = history_mapping_df.set_index('工單號碼')['人員'].to_dict()

        # 3. 確定應用映射的範圍：當前工單編號在『歷史主工單』映射表中
        mask_history_apply = combined_df['工單編號'].isin(history_mapping.keys())

        # 4. 應用映射
        # 這裡會將歷史主工單的人員指派給 combined_df 中對應的所有行（無論當前是主工單還是子工單）
        combined_df.loc[mask_history_apply, '人員'] = combined_df.loc[mask_history_apply, '工單編號'].map(history_mapping)

    # ==========================================================
    # 對多主工單引用的子工單解除限定人員
    # ==========================================================
    suborder_counts = combined_df.groupby("工單編號")["主工單編號"].nunique()
    conflicted_suborders = suborder_counts[suborder_counts > 1].index.tolist()
    combined_df.loc[combined_df["工單編號"].isin(conflicted_suborders), "人員"] = None

    # 1. 定義 main_orders (從 combined_df 中過濾出主工單行)
    main_orders = combined_df[combined_df["工單編號"] == combined_df["主工單編號"]].copy() 
    # 加上 .copy() 是一個好習慣，確保操作是在副本上，避免 SettingWithCopyWarning。

    # 2. 強制轉換日期欄位 (解決 TypeError 錯誤)
    main_orders["預計開工日"] = pd.to_datetime(main_orders["預計開工日"], errors='coerce')
        
    # 3. 群組聚合並取最早日期
    main_dates = main_orders.groupby("工單編號")["預計開工日"].min()

    # 4. 映射回 combined_df
    combined_df["主工單開工日"] = combined_df["主工單編號"].map(main_dates)

    '''
    # ==========================================================
    # 均分未指派主工單
    # 加入客戶需求日進去指派
    # ==========================================================
    mask_81A = combined_df["主工單編號"].str.startswith("81A", na=False)
    mask_81B = combined_df["主工單編號"].str.startswith("81B", na=False)
    mask_81C = combined_df["主工單編號"].str.startswith("81C", na=False)
    mask_81 = mask_81A | mask_81B | mask_81C

    assigned_main_ids = combined_df[
        (combined_df["工單編號"] == combined_df["主工單編號"]) &
        (combined_df["人員"].isin(["A159", "B201", "A830"]))
    ]["工單編號"].unique()

    unassigned_df = combined_df[
        mask_81 &
        ~combined_df["主工單編號"].isin(assigned_main_ids)
    ].copy()

    main_quantities = (
        unassigned_df
        .groupby("主工單編號")["刀次"]
        .sum()
        .reset_index()
        .sort_values(by="刀次", ascending=False)
        .reset_index(drop=True)
    )

    assign_A159, assign_B201, assign_A830 = [], [], []
    sorted_people = [
        {"name": "A159", "total": 0, "list": assign_A159},
        {"name": "B201", "total": 0, "list": assign_B201},
        {"name": "A830", "total": 0, "list": assign_A830}
    ]

    for _, row in main_quantities.iterrows():
        main_id = row["主工單編號"]
        qty = row["刀次"]
        sorted_people.sort(key=lambda x: x["total"])
        person = sorted_people[0]
        person["list"].append(main_id)
        person["total"] += qty

    combined_df.loc[combined_df["主工單編號"].isin(assign_A159), "人員"] = "A159"
    combined_df.loc[combined_df["主工單編號"].isin(assign_B201), "人員"] = "B201"
    combined_df.loc[combined_df["主工單編號"].isin(assign_A830), "人員"] = "A830"
    
    
    # 排序
    main_orders = combined_df[combined_df["工單編號"] == combined_df["主工單編號"]]
    main_dates = main_orders.groupby("工單編號")["預計開工日"].min()
    combined_df["主工單開工日"] = combined_df["主工單編號"].map(main_dates)
    combined_df = combined_df.sort_values(by=["主工單開工日", "主工單編號", "刀次"], ascending=[True, True, False])

    
    # check "人員": 如果有空的 就先看他是主還是子工單 如果是子工單就拉主工單的"人員"使用 如果是主工單就報錯
    empty_person_mask = combined_df['人員'].isna()
    if empty_person_mask.any():
        for idx in combined_df[empty_person_mask].index:
            row = combined_df.loc[idx]
            main_id = row["主工單編號"]
            if row["工單編號"] != main_id:
                # 子工單，繼承主工單人員
                main_person = combined_df.loc[combined_df["工單編號"] == main_id, "人員"].iloc[0]
                combined_df.at[idx, "人員"] = main_person
            else:
                # 主工單空人員，防呆報錯
                raise ValueError(f"主工單 {row['工單編號']} 尚未分配人員，請檢查排程邏輯。")
    '''

    return combined_df


# 先篩選完再來排
def choose_date(df: pd.DataFrame) -> pd.DataFrame:
    global start_date, end_date

    df = df.copy()
    df["日期篩選用"] = pd.to_datetime(df["開工日期"], errors="coerce")

    root = ctk.CTk()
    root.withdraw()
    dialog = DateFilterDialog(root)
    root.wait_window(dialog)

    start_date = dialog.start_date_choose
    end_date = dialog.end_date_choose

    # ✅ 若使用者未選擇日期就關閉，直接中止
    if not start_date and not end_date:
        raise ValueError("❌ 請選擇日期區間後再繼續")

    # ✅ 執行篩選
    if start_date and end_date:
        df = df[(df["日期篩選用"] >= start_date) & (df["日期篩選用"] <= end_date)]
    elif start_date:
        df = df[df["日期篩選用"] >= start_date]
    elif end_date:
        df = df[df["日期篩選用"] <= end_date]

    df = df.drop(columns=["日期篩選用"])

    return df


# 全部排完才刪選時間 或是 刪除多餘欄位
# 找到可以搭配的
'''
找品號 前面非數字的(英文部份)
1. 非B110A開頭 -> 88cm, 反之 -> 68cm
2. 搭配完剛好88cm or 68cm
3. 搭配完主工單數量要剛好的優先
4. 剩下沒搭配的刀次改為空
'''
def extract_prefix(item_code):
    m = re.match(r"^[^\d]+", item_code)
    return m.group(0) if m else ""

def find_88cm_combination(row_i, row_j, max_car=88):
    for ci in range(1, max_car+1):
        for cj in range(1, max_car+1):
            total = ci * row_i["公分"] + cj * row_j["公分"]
            if abs(total - 88) < 1e-6:
                if (row_i["預估良品數"] / ci).is_integer():
                    main_cut = int(row_i["預估良品數"] / ci)
                    child_qty = main_cut * cj
                    # 新增條件：子工單用量不得超過餘量
                    if child_qty <= row_j["預估良品數"]:
                        return ci, cj
    return None

def find_68cm_combination(row_i, row_j, max_car=68):
    for ci in range(1, max_car+1):
        for cj in range(1, max_car+1):
            total = ci * row_i["公分"] + cj * row_j["公分"]
            if abs(total - 68) < 1e-6:
                if (row_i["預估良品數"] / ci).is_integer():
                    main_cut = int(row_i["預估良品數"] / ci)
                    child_qty = main_cut * cj
                    # 新增條件：子工單用量不得超過餘量
                    if child_qty <= row_j["預估良品數"]:
                        return ci, cj
    return None

# 將可以配對的進行配對 根據 mode 去找要88cm or 68cm的品號
def remaining_cut_clean_and_repair(df_paired_split, df, mode="88cm", cut_limit=55):
    """
    mode: "88cm" → 處理非 B110A
          "68cm" → 處理 B110A
    """
    df = df.copy()
    df["公分"] = pd.to_numeric(df["公分"], errors="coerce")
    df["預估良品數"] = pd.to_numeric(df["預估良品數"], errors="coerce")
    df["品號前綴"] = df["品號"].apply(extract_prefix)
    df["品號結尾"] = df["品號"].str[-2:]
    df["is_B110A"] = df["品號"].str.startswith("B110A")

    merged_indices = set()
    leftover_rows = []

    # 判斷處理範圍
    if mode == "88cm":
        target_df = df[~df["is_B110A"]]
        find_func = find_88cm_combination
    elif mode == "68cm":
        target_df = df[df["is_B110A"]]
        find_func = find_68cm_combination
    else:
        raise ValueError("mode 必須是 '88cm' 或 '68cm'")

    # --- 核心配對流程 ---
    for prefix, group in target_df.groupby("品號前綴"):
        group = group.reset_index()
        n = len(group)

        for i in range(n):
            idx_i = group.at[i, "index"]
            if idx_i in merged_indices:
                continue

            row_i = df.loc[idx_i]

            for j in range(i + 1, n):
                idx_j = group.at[j, "index"]
                if idx_j in merged_indices:
                    continue
                row_j = df.loc[idx_j]

                if row_i["品號結尾"] != row_j["品號結尾"]:
                    continue

                res = find_func(row_i, row_j)
                if res:
                    ci, cj = res

                    main_car = ci
                    main_cut = int(row_i["預估良品數"] / ci)

                    child_car = cj
                    child_cut = 0
                    child_qty = main_cut * child_car

                    main_qty = row_i["預估良品數"] - child_qty
                    if main_qty < 0:
                        main_qty = 0

                    main_order = row_i.copy()
                    main_order["車數"] = main_car
                    main_order["刀次"] = main_cut
                    main_order["預估良品數"] = main_qty

                    child_order = row_j.copy()
                    child_order["車數"] = child_car
                    child_order["刀次"] = child_cut
                    child_order["預估良品數"] = child_qty

                    # === 檢查主工單的剩餘可配量 ===
                    main_order_id = str(row_i["工單編號"])

                    # 從 global map 撈取目前可用數量（經過歷史扣除後的最新數量）
                    original_qty = int(GLOBAL_ORDER_QTY_MAP.get(main_order_id, 0))

                    if "工單編號" in df_paired_split.columns:
                        used_qty = df_paired_split.loc[
                            df_paired_split["工單編號"] == main_order_id, "預估良品數"
                        ].sum()
                    else:
                        used_qty = 0

                    available_qty = original_qty - used_qty
                    print(f"要配對工單{main_order_id} 開工數量 {original_qty} 已使用數量 {used_qty}")

                    # 如果主工單已經用完，就跳過
                    if available_qty <= 0:
                        print(f"⚠️ 工單 {main_order_id} 已使用完（已用 {used_qty}/{original_qty}），跳過。")
                        continue

                    df_pair_split = split_pair_orders(main_order, child_order, cut_limit=cut_limit)
                    df_paired_split = pd.concat([df_paired_split, df_pair_split], ignore_index=True)

                    merged_indices.update([idx_i, idx_j])

                    leftover_qty = row_j["預估良品數"] - child_qty
                    if leftover_qty > 0:
                        leftover_order = row_j.copy()
                        leftover_order["預估良品數"] = leftover_qty
                        leftover_order["刀次"] = math.ceil(leftover_qty / row_j["車數"])
                        leftover_order["車數"] = row_j["車數"]
                        leftover_rows.append(leftover_order)

                    break

    # --- 未配對 ---
    leftover_indices = set(df.index) - merged_indices
    remaining_rows = df.loc[list(leftover_indices)].copy()
    remaining_rows["刀次"] = remaining_rows["刀次"].fillna("")

    if leftover_rows:
        remaining_rows = pd.concat([remaining_rows, pd.DataFrame(leftover_rows)], ignore_index=True)

    # 1. 確保 '預計開工日' 是日期類型
    remaining_rows["預計開工日"] = pd.to_datetime(remaining_rows["預計開工日"], errors='coerce')
    
    # 2. 進行排序，並確保索引連續
    remaining_rows = remaining_rows.sort_values(
        by="預計開工日", 
        ascending=True
    ).copy()

    for col in ["品號前綴", "品號結尾", "is_B110A"]:
        if col in remaining_rows.columns:
            remaining_rows = remaining_rows.drop(columns=[col])
        if col in df_paired_split.columns:
            df_paired_split = df_paired_split.drop(columns=[col])

    return df_paired_split.reset_index(drop=True), remaining_rows.reset_index(drop=True)


# 一天最多55刀次 主子工單要拆分的話要一起拆分 
def split_pair_orders(main_order, child_order, cut_limit=55):
    """
    拆分主工單與子工單，並保持一一對應交錯輸出
    - main_order: 刀次 > 0
    - child_order: 刀次 = 0
    """
    results = []

    total_cut = int(main_order["刀次"])
    if total_cut <= 0:
        return [main_order.to_frame().T, child_order.to_frame().T]

    # 每刀對應的數量
    base_child_qty_per_cut = child_order["預估良品數"] / total_cut
    base_main_qty_per_cut = main_order["預估良品數"] / total_cut

    while total_cut > cut_limit:
        # 主工單
        new_main = main_order.copy()
        new_main["刀次"] = cut_limit
        new_main["預估良品數"] = int(base_main_qty_per_cut * cut_limit)

        # 子工單
        new_child = child_order.copy()
        new_child["刀次"] = 0
        new_child["預估良品數"] = int(base_child_qty_per_cut * cut_limit)

        # 成對加入
        results.append(new_main.to_frame().T)
        results.append(new_child.to_frame().T)

        total_cut -= cut_limit

    if total_cut > 0:
        new_main = main_order.copy()
        new_main["刀次"] = total_cut
        new_main["預估良品數"] = int(base_main_qty_per_cut * total_cut)

        new_child = child_order.copy()
        new_child["刀次"] = 0
        new_child["預估良品數"] = int(base_child_qty_per_cut * total_cut)

        results.append(new_main.to_frame().T)
        results.append(new_child.to_frame().T)

    # 串成 DataFrame
    return pd.concat(results, ignore_index=True)



def do_check(final_df):
    
    # 🚨 檢查：確保 '唯一主工單ID' 存在
    if "唯一主工單ID" not in final_df.columns:
        raise ValueError("缺少 '唯一主工單ID' 欄位，無法進行正確分批排序。請檢查 prepare_final_schedule 函數。")

    # 確保主工單開工日是 datetime
    final_df["主工單開工日"] = pd.to_datetime(final_df["主工單開工日"], errors='coerce')

    # 核心修正：使用 '唯一主工單ID' 進行分批排序，以維持拆分後的相對順序
    # 1. 主工單開工日 (時間優先)
    # 2. 唯一主工單ID (維持每個分拆批次的內在順序)
    # 3. 刀次 (確保在同一批次內，主工單排在子工單之前)
    sorted_df = final_df.sort_values(
        by=["主工單開工日", "唯一主工單ID", "刀次"], 
        ascending=[True, True, False]
    )

    return sorted_df


def open_file(filepath):
    if platform.system() == 'Windows':
        os.startfile(filepath)
    elif platform.system() == 'Darwin':  # macOS
        subprocess.call(['open', filepath])
    else:  # Linux
        subprocess.call(['xdg-open', filepath])


#--------------------日期排程----------------------------        
def final_schedule_list(sort_df, df_history):
    
    global start_date, end_date, user_schedule_date  

    # 彈出假期選擇視窗
    root = ctk.CTk()
    root.withdraw()
    dialog = MultiPersonLeaveDialog(root, start_date, end_date)  # date 格式
    root.wait_window(dialog)

    if dialog.result is None:
        raise Exception("❌ 使用者取消操作，請重新選擇休假日")
    print(dialog.result)

    break_dates = {
        "A159": set(pd.to_datetime(date).date() for date in dialog.result.get("leaves").get("A159", [])),
        "A830": set(pd.to_datetime(date).date() for date in dialog.result.get("leaves").get("A830", [])),
        "B201": set(pd.to_datetime(date).date() for date in dialog.result.get("leaves").get("B201", [])),
    }
    print(break_dates)

    # 排程開始日期
    schedule_start = dialog.result.get("start_date") 
    user_schedule_date = schedule_start

    start_date = start_date.date()

    return sort_df, break_dates



def generate_schedule(df: pd.DataFrame, break_dates: dict, max_lookback_days: int = 165) -> pd.DataFrame:
    
    df = df.copy()
    df["預計開工日"] = pd.to_datetime(df["預計開工日"])
    df["預計完工日"] = pd.to_datetime(df["預計完工日"])
    df["實際排程日期"] = None
    
    # 🎯 產能參數定義
    BASE_DAILY_CAPACITY = 165 
    CAPACITY_DEDUCTION_PER_ABSENCE = 55 
    
    # ----------------------------------------------------
    # ⭐ 修正區 1: 使用穩定的 '唯一主工單ID' 進行分組，取代索引遞迴 ⭐
    # ----------------------------------------------------
    GROUP_ID_COLUMN = "唯一主工單ID" 
    
    # 假設 '唯一主工單ID' 欄位名稱是準確的，且主子工單相同。
    df["組別ID"] = df[GROUP_ID_COLUMN].astype(str)
    
    # 確保所有行都有一個組別ID（如果子工單的該欄位為空，則會繼承上一個工單的值）
    df["組別ID"].fillna(method='ffill', inplace=True)
    
    # 如果列表開頭仍然是 NaN（例如第一行就是子工單），使用 工單編號 暫時填補
    df["組別ID"].fillna(df["工單編號"], inplace=True) 
    
    # ----------------------------------------------------
    
    # 每組用「最早完工日」決定排程範圍
    # 分組邏輯現在使用 "組別ID"
    df["組內最早完工日"] = df.groupby("組別ID")["預計完工日"].transform("min")

    groups = df.groupby("組別ID")
    
    # 排序：根據組內最早完工日逆向排序
    # 這裡的 g[0] 現在是 '組別ID'，g[1].index[0] 是該群組第一行的 DataFrame 索引
    sorted_groups = sorted(groups, key=lambda g: df.loc[g[1].index[0], "組內最早完工日"], reverse=True)

    # 建立排程容量表
    latest_end = df["預計完工日"].max()

    earliest_start = latest_end - timedelta(days=max_lookback_days)

    schedule_capacity = {}
    date = latest_end
    
    # 步驟 1: 建立每日請假人數統計表
    absence_counts = {}
    for person_id, dates_set in break_dates.items():
        for leave_date in dates_set:
            if isinstance(leave_date, pd.Timestamp):
                 leave_date = leave_date.date()
            absence_counts[leave_date] = absence_counts.get(leave_date, 0) + 1
            
    # 步驟 2: 根據請假人數計算每日總產能
    while date >= earliest_start:
        day = date.to_pydatetime().date()
        
        if date.weekday() < 5: # 週一到週五
            absent_count = absence_counts.get(day, 0) 
            current_capacity = BASE_DAILY_CAPACITY - (absent_count * CAPACITY_DEDUCTION_PER_ABSENCE)
            
            if current_capacity > 0:
                schedule_capacity[day] = current_capacity
                
        date -= timedelta(days=1)


    # 執行逆向排程
    for group_id, group in sorted_groups:
        group_df = df.loc[group.index]
        earliest_end = group_df["組內最早完工日"].min().date()
        
        start_date = earliest_end - timedelta(days=max_lookback_days)

        total_doz = group_df["刀次"].astype(float).sum()
        remaining = total_doz
        assigned_list = []
        current_date = earliest_end

        # ⭐ 找出主工單的 DataFrame 索引 (用於寫入邏輯的判斷和錯誤報告)
        # 尋找群組中刀次 > 0 的第一筆工單
        main_order_candidates = group_df[group_df["刀次"].astype(float) > 0]
        
        # 假設每個群組至少有一筆主工單
        if not main_order_candidates.empty:
            main_idx = main_order_candidates.index[0]
        else:
            # 如果群組中沒有刀次 > 0 的工單 (理論上不該發生)，則跳過或使用第一行的索引
            main_idx = group_df.index[0] 

        while remaining > 0 and current_date >= start_date:
            current_date_key = current_date
            
            if current_date_key in schedule_capacity and schedule_capacity[current_date_key] > 0:
                available = schedule_capacity[current_date_key]
                assign = min(available, remaining)
                schedule_capacity[current_date_key] -= assign
                
                assigned_list.append((current_date_key.strftime("%Y-%m-%d"), assign))
                remaining -= assign
                
            current_date -= timedelta(days=1)

        if remaining > 0:
            # 使用我們找到的主工單編號進行錯誤報告
            raise Exception(f"排程無法完成，工單 {df.loc[main_idx, '工單編號']} 剩餘刀次 {remaining}")

        # 寫入每筆工單的排程日期
        for idx in group.index:
            # 判斷當前行是否為我們找到的主工單 (刀次 > 0 的行)
            if idx == main_idx: 
                # 主工單：寫入日期(刀次)
                formatted = [f"{d}({c})" for d, c in sorted(assigned_list)]
            else:
                # 子工單：只寫入日期 (這確保子工單獲得拆分後的所有日期)
                formatted = [d for d, _ in sorted(assigned_list)]
            df.at[idx, "實際排程日期"] = "\n".join(formatted)

    # ----------------------------------------------------
    # ⭐ 修正區 2: 刪除新的群組欄位 ⭐
    # ----------------------------------------------------
    df.drop(columns=["組內最早完工日", "組別ID"], inplace=True, errors='ignore')

    df_new = split_schedule_dates(df)

    return df_new, df

# 正則表達式：用於解析日期與括號內的數字，確保解析的健壯性
pattern = re.compile(r"(\d{4}[-/]\d{2}[-/]\d{2})(?:\(([\d.]+)\))?")

def split_schedule_dates(df: pd.DataFrame) -> pd.DataFrame:
    new_rows = []
    processed_indices = set() # 用來記錄哪些子工單已經被處理過，避免重複計算
    
    # 必須使用 index 迭代，才能向前查看下一行
    for i in df.index:
        
        # 如果這個 index 已經被前一個 Master 當作 Child 處理過，則跳過
        if i in processed_indices:
            continue
            
        # 1. 識別當前 Master Row
        master_idx = i
        master_row = df.loc[master_idx]
        current_master_id = str(master_row.get("主工單編號", "")).strip()
        
        # 2. 建立 Group：找出所有跟隨在 Master 後面的 Child Row
        group_to_split = [(master_idx, master_row)]
        
        # 向前查看 (Look-ahead)
        j = i + 1
        while j in df.index:
            child_row = df.loc[j]
            child_master_id = str(child_row.get("主工單編號", "")).strip()

            # 判斷是否為子工單的邏輯：
            # A. 子工單的主工單編號是空的 (經典子單)
            # B. 子工單的主工單編號跟 Master 相同 (同一組的另一個項目)
            is_child_of_group = (
                child_master_id == "" or 
                (child_master_id == current_master_id and child_master_id != "")
            )

            if is_child_of_group:
                group_to_split.append((j, child_row))
                processed_indices.add(j) # 標記為已處理
                j += 1
            else:
                break # 遇到不屬於這組的新單，停止分組

        # =================================================================
        # 3. 處理這個 Group
        # =================================================================
        
        # 抓取 Master Row 的排程資訊 (必須用 Master Row 裡的數據來驅動)
        sched_str = str(master_row.get("實際排程日期", "")).strip()
        total_doz = master_row.get("刀次", 0)
        total_qty = master_row.get("需求數量")
        total_good_qty = master_row.get("預估良品數")
        
        is_splitting_master = isinstance(sched_str, str) and "(" in sched_str and total_doz > 0
        
        # 使用你原來的刀次限制
        if total_doz <= 60:
            is_splitting_master = False
            
        matches = []
        if is_splitting_master:
            matches = pattern.findall(sched_str)


        # 情況 A: 需要拆分 (Group 內有 Master 且有複雜排程)
        if matches:
            
            # 依照日期/比例進行雙迴圈展開 (確保輸出順序正確：日期1主子 -> 日期2主子)
            for date_str, doz_str in matches:
                
                # 計算比例
                ratio = 1.0
                current_doz = 0.0
                if doz_str and total_doz > 0:
                    current_doz = float(doz_str)
                    ratio = current_doz / total_doz

                # 遍歷 Group 內的所有行 (Master + Children)
                for index, original_row in group_to_split:
                    
                    new_row = original_row.copy()
                    
                    # 1. 修改日期
                    new_row["實際排程日期"] = pd.to_datetime(date_str).strftime("%Y/%m/%d")

                    # 2. 更新數量 (使用各自的原始數量進行比例計算)
                    row_qty = original_row.get("需求數量")
                    row_good_qty = original_row.get("預估良品數")

                    if row_qty is not None and pd.notna(row_qty):
                        new_row["需求數量"] = round(row_qty * ratio)
                    
                    if row_good_qty is not None and pd.notna(row_good_qty):
                        new_row["預估良品數"] = round(row_good_qty * ratio)

                    # 3. 更新刀次 (只更新 Master Row 的刀次，Child Row 的刀次通常為 0 或維持原值)
                    # 判斷是否為 Master Row (即 Group 的第一行)
                    if index == master_idx and doz_str:
                        new_row["刀次"] = current_doz
                    # 否則保持原值 (通常 Child Row 的刀次是 0)
                    
                    new_rows.append(new_row)
        
        # 情況 B: 不需要拆分 (Master Row 沒有複雜排程)
        else:
            # 只嘗試修正日期格式，然後原樣保留整個 Group
            for index, original_row in group_to_split:
                new_row = original_row.copy()
                
                try:
                    date_part = str(original_row.get("實際排程日期", "")).split("(")[0].strip()
                    if date_part:
                        new_row["實際排程日期"] = pd.to_datetime(date_part).strftime("%Y/%m/%d")
                    else:
                        new_row["實際排程日期"] = None
                except Exception:
                    new_row["實際排程日期"] = None
                    
                new_rows.append(new_row)

    return pd.DataFrame(new_rows)



def final_cal_list(df: pd.DataFrame, start_d: datetime.date, break_dates: dict) -> pd.DataFrame:
    df = df.copy()

    if isinstance(start_d, pd.Timestamp):
        start_d = start_d.date()
    elif isinstance(start_d, str):
        start_d = pd.to_datetime(start_d).date()
        
    # 🎯 修正點 1：統一總產能基準和扣除參數
    BASE_DAILY_CAPACITY = 165 
    CAPACITY_DEDUCTION_PER_ABSENCE = 55 
    
    # 🎯 新增邏輯：計算每日請假人數統計表 (與上一個函式邏輯相同)
    absence_counts = {}
    for person_id, dates_set in break_dates.items():
        for leave_date in dates_set:
            if isinstance(leave_date, pd.Timestamp):
                 leave_date = leave_date.date()
            absence_counts[leave_date] = absence_counts.get(leave_date, 0) + 1
            
    capacity_used = {}

    doses = pd.to_numeric(df["刀次"], errors="coerce").fillna(0).tolist()
    sku_list = df.get("品號", pd.Series([None]*len(doses))).tolist()
    new_dates = []
    current_date = start_d

    for dose, sku in zip(doses, sku_list):
        if dose == 0:
            # 子工單：繼承上一筆日期
            new_dates.append(new_dates[-1] if new_dates else current_date)
            continue

        while True:
            wd = current_date.weekday()
            
            # 🎯 修正點 2：統一週末和請假判斷
            if wd < 5: # 週一到週五
                
                # 取得當天請假人數
                absent_count = absence_counts.get(current_date, 0) 
                
                # 計算動態產能上限
                limit = BASE_DAILY_CAPACITY - (absent_count * CAPACITY_DEDUCTION_PER_ABSENCE)
                
                # 如果計算出的產能為零或負數，跳過這一天
                if limit <= 0:
                    current_date += timedelta(days=1)
                    continue

                # 🎯 修正點 3：移除原有的特殊品號/特殊日期的判斷邏輯
                # 因為現在是總產能排程，不再需要根據品號或週三/週五來調整產能
                
                used_info = capacity_used.get(current_date, {"count": 0}) 

                if used_info["count"] + dose <= limit:
                    # 滿足條件，排程
                    used_info["count"] += dose
                    capacity_used[current_date] = used_info
                    new_dates.append(current_date)
                    break # 跳出 while True 迴圈，進行下一筆工單

            current_date += timedelta(days=1) # 跳到下一天繼續檢查

    df.loc[:, "預計開工日"] = [d.strftime("%Y/%m/%d") for d in new_dates]
    df.loc[:, "預計完工日"] = df["預計開工日"]

    return df


class MultiPersonLeaveDialog(ctk.CTkToplevel):
    def __init__(self, parent, start_date, end_date, names=("A159", "A830", "B201")):
        super().__init__(parent)
        self.title("選擇三位人員的休假日")
        width, height = 900, 600
        self.geometry(f"{width}x{height}")
        self.center_window(width, height)

        # 狀態資料
        self.leave_dates = {}
        self.names = names
        self.calendars = {}
        self.date_widgets = {}
        self.start_date_selected = None
        self.result = None
        self.start_date = start_date
        self.end_date = end_date

        # Calendar 統一大小
        self.calendar_font_size = 14

        # 主要容器
        self.step_frame = ctk.CTkFrame(self)
        self.step_frame.pack(expand=True, fill="both", padx=20, pady=20)

        # 顯示第一步
        self.show_step_1()
        self.grab_set()

    # --- 步驟1：選排程起始日 ---
    def show_step_1(self):
        for widget in self.step_frame.winfo_children():
            widget.destroy()

        ctk.CTkLabel(self.step_frame, text="📅 請選擇排程起始日",
                     font=("微軟正黑體", 20)).pack(pady=(15, 5))

        start_frame = ctk.CTkFrame(self.step_frame)
        start_frame.pack(pady=5)

        self.start_calendar = Calendar(
            start_frame, mindate=None, maxdate=None, date_pattern="yyyy-mm-dd",
            background="#006699", foreground="white",
            headersbackground="#006699", headersforeground="white",
            normalbackground="white", weekendbackground="white",
            disabledbackground="#cccccc", bordercolor="#004080",
            othermonthforeground="#999999", selectmode="day",
            font=("微軟正黑體", self.calendar_font_size)
        )
        self.start_calendar.pack(padx=10, pady=5)

        self.start_label = ctk.CTkLabel(self.step_frame, text="目前未選擇",
                                        font=("微軟正黑體", 16), text_color="gray")
        self.start_label.pack(pady=(0, 10))

        btn_frame = ctk.CTkFrame(self.step_frame)
        btn_frame.pack(pady=15)

        ctk.CTkButton(btn_frame, text="✅ 設定為排程開始日",
                      font=("微軟正黑體", 14),
                      command=self.set_start_date).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="➡️ 下一步",
                      font=("微軟正黑體", 14),
                      command=self.show_step_2).pack(side="left", padx=10)

    # --- 步驟2：選休假 ---
    def show_step_2(self):
        if not self.start_date_selected:
            return self.error("請先選擇排程起始日！")

        for widget in self.step_frame.winfo_children():
            widget.destroy()

        ctk.CTkLabel(self.step_frame, text="請為三位人員分別選擇休假日",
                    font=("微軟正黑體", 20)).pack(pady=10)

        # Container 統一放在 step_frame
        container = ctk.CTkFrame(self.step_frame)
        container.pack(pady=10, padx=10, expand=True, fill="both")

        calendar_style = {
            "background": "#004080",
            "foreground": "white",
            "font": ("微軟正黑體", 14),
            "headersbackground": "#004080",
            "headersforeground": "white",
            "normalbackground": "white",
            "weekendbackground": "white",
            "disabledbackground": "#cccccc",
            "bordercolor": "#004080",
            "othermonthforeground": "#999999",
            "selectmode": "day",
            "date_pattern": "yyyy-mm-dd",
        }

        for i, name in enumerate(self.names):
            frame = ctk.CTkFrame(container)
            frame.grid(row=0, column=i, padx=10, pady=10, sticky="n")

            ctk.CTkLabel(frame, text=name, font=("微軟正黑體", 16)).pack(pady=5)

            cal = Calendar(frame, mindate=self.start_date, maxdate=self.end_date, **calendar_style)
            cal.pack(pady=5)
            self.calendars[name] = cal
            self.leave_dates[name] = set()

            ctk.CTkButton(frame, text="加入休假日", font=("微軟正黑體", 14),
                        command=lambda n=name: self.add_date(n)).pack(pady=5)

            label = ctk.CTkLabel(frame, text="已選：無", wraplength=200,
                                justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            self.date_widgets[name] = label

        self.error_label = ctk.CTkLabel(self.step_frame, text="", text_color="red", font=("微軟正黑體", 12))
        self.error_label.pack(pady=5)

        ctk.CTkLabel(self.step_frame, text="📝 若三人皆無休假，請直接按「✅ 確定」。",
                    font=("微軟正黑體", 14), text_color="gray").pack(pady=(0, 10))

        # 按鈕也放 step_frame
        btn_frame = ctk.CTkFrame(self.step_frame)
        btn_frame.pack(pady=15)

        ctk.CTkButton(btn_frame, text="⬅️ 上一步", font=("微軟正黑體", 14),
                    command=self.show_step_1).pack(side="left", padx=20)
        ctk.CTkButton(btn_frame, text="✅ 確定", font=("微軟正黑體", 14),
                    command=self.confirm).pack(side="left", padx=20)
    
    # --- 工具函式 ---
    def center_window(self, width, height):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"+{x}+{y}")

    def set_start_date(self):
        selected = self.start_calendar.get_date()
        self.start_date_selected = selected
        self.start_label.configure(text=f"排程開始日：{selected}", text_color="#00CC99")
        if hasattr(self, "error_label"):
            self.error_label.configure(text="")

    def add_date(self, name):
        selected = self.calendars[name].get_date()
        self.leave_dates[name].add(selected)
        self.update_widget(name)

    def remove_date(self, name, date):
        self.leave_dates[name].discard(date)
        self.update_widget(name)

    def show_delete_menu(self, event, name):
        if not self.leave_dates[name]:
            return
        font = tkfont.Font(family="微軟正黑體", size=18, weight="bold")
        menu = tk.Menu(self, tearoff=0, font=font)
        for date in sorted(self.leave_dates[name]):
            menu.add_command(label=f"🗑️ 刪除日期：{date}",
                             command=lambda d=date: self.remove_date(name, d))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def update_widget(self, name):
        dates = sorted(list(self.leave_dates[name]))
        widget = self.date_widgets[name]
        widget.pack_forget()

        if len(dates) == 0:
            label = ctk.CTkLabel(widget.master, text="已選：無", wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            self.date_widgets[name] = label
        elif len(dates) <= 2:
            text = "\n".join(dates)
            label = ctk.CTkLabel(widget.master, text=text, wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            label.bind("<Button-1>", lambda e, n=name: self.show_delete_menu(e, n))
            self.date_widgets[name] = label
        else:
            option_menu = ctk.CTkOptionMenu(widget.master, values=dates,
                                            command=lambda d: self.remove_date(name, d),
                                            font=("微軟正黑體", 14),
                                            width=180)
            option_menu.set("休假日列表(點選可刪除)")
            option_menu.pack(pady=5)
            self.date_widgets[name] = option_menu

    def error(self, msg):
        if hasattr(self, "error_label"):
            self.error_label.configure(text=msg)
        else:
            ctk.CTkLabel(self.step_frame, text=msg, font=("微軟正黑體", 14),
                         text_color="red").pack(pady=5)

    def confirm(self):
        if not self.start_date_selected:
            self.error("⚠️ 請先選擇排程起始日！")
            return
        self.result = {
            "start_date": self.start_date_selected,
            "leaves": self.leave_dates
        }
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()


# 最後分配: 
# 1. 把剩餘工單內料號拿去已經配好的工單篩選若有與"基本資料符合"且"刀次不為0"的資料，則將取其一部份來搭配這原本剩餘工單
# 假設 df_paired_split, df_no_pair_split, df_no_pair_second, remaining 
# 已經被定義並傳入
def final_doAssignAndSort_DEBUG(df_paired_split, df_no_pair_split, df_no_pair_second, remaining, base_df):
    
    # 內建輔助函數：還原車數和重算刀次
    def _restore_single_df(df_to_restore, base_df, original_car_map):
        """
        將工單的車數還原為其原始設定車數 (來自 base_df)，
        並根據剩餘的良品數量重新計算刀次和預估良品數。
        """
        if df_to_restore is None or df_to_restore.empty:
            return pd.DataFrame()
        
        df_restored = df_to_restore 
        
        df_restored.loc[:, "remain_qty"] = df_restored["預估良品數"]
        df_restored.loc[:, "original_car"] = df_restored["品號"].map(original_car_map).fillna(1)
        
        df_restored.loc[:, "original_car"] = pd.to_numeric(df_restored["original_car"], errors='coerce').clip(lower=1).fillna(1)

        # 重新計算刀次：新刀次 = 無條件進位(剩餘良品數 / 原始車數)
        df_restored.loc[:, "new_cut"] = df_restored.apply(
            lambda row: math.ceil(row["remain_qty"] / row["original_car"]),
            axis=1
        )
        
        # 更新 DataFrame
        df_restored.loc[:, "車數"] = df_restored["original_car"]
        df_restored.loc[:, "刀次"] = df_restored["new_cut"]
        
        df_restored.loc[:, "預估良品數"] = df_restored["刀次"] * df_restored["車數"]
        df_restored.loc[:, "餘量"] = df_restored["預估良品數"]

        # 清理輔助欄位
        df_restored = df_restored.drop(columns=["remain_qty", "original_car", "new_cut"], errors='ignore')
        
        return df_restored
    
    # 轉型 base_df 數值欄位
    try:
        # === 建立 料號 -> 原始車數 的對應表 (供還原時使用) ===
        # ⭐ 關鍵修正：篩選出搭1料號為空（None/NaN/空字串）的行，這才是該料號的自產基本配置。
        # 由於 '搭1料號' 可能是字串，使用 .fillna('') 或 .isna() 進行判斷
        base_df_for_restore = base_df[
            (base_df['搭1料號'].isna()) | (base_df['搭1料號'] == '') | (base_df['搭1料號'].str.strip().eq(''))
        ].copy()
        
        # 如果有重複的料號（有多行搭1料號為空），只取第一筆或依其他規則（這裡使用 drop_duplicates）
        original_car_map = base_df_for_restore.drop_duplicates(subset=['料號']).set_index("料號")["車數"].to_dict()
        base_df["車數"] = pd.to_numeric(base_df["車數"], errors="coerce")
        base_df["搭1產出車數"] = pd.to_numeric(base_df["搭1產出車數"], errors="coerce")
        base_df["搭2產出車數"] = pd.to_numeric(base_df.get("搭2產出車數"), errors="coerce")
    except Exception as e:
        print(f"❌ 讀取配置或基本資料失敗: {e}")
        return {"all": pd.DataFrame(), "debug_log": pd.DataFrame(), "final_remaining": remaining, 
                "df_no_pair_second_mutable": df_no_pair_second, "df_no_pair_split_mutable": df_no_pair_split}
    
    debug_log = [] 
    newly_paired_rows = [] 

    # === 創建可變的子工單池副本並重設索引 ===
    df_no_pair_split_mutable = df_no_pair_split.copy().reset_index(drop=True) if df_no_pair_split is not None else pd.DataFrame()
    df_no_pair_second_mutable = df_no_pair_second.copy().reset_index(drop=True) if df_no_pair_second is not None else pd.DataFrame()
    # =================================================================

    remaining_sorted = remaining.copy()
    remaining_sorted["預計開工日"] = pd.to_datetime(remaining_sorted["預計開工日"], errors='coerce')
    remaining_sorted = remaining_sorted.sort_values(by="預計開工日", ascending=True).reset_index(drop=True)

    print("\n--- 階段 A：開始處理 remaining (主) 與外部池 (輔) 互配 ---")
    
    for idx, row in remaining_sorted.iterrows(): 
        original_item = row["品號"]
        original_qty_val = pd.to_numeric(row["預估良品數"], errors="coerce") if not pd.isna(row["預估良品數"]) else 0
        sn = row["工單編號"]
        original_order = row.copy()
        
        main_car_count = pd.to_numeric(original_order.get("車數", 1), errors="coerce") or 1
        if main_car_count <= 0: main_car_count = 1
        main_cut_count = math.ceil(original_qty_val / main_car_count)
        original_order["刀次"] = main_cut_count
        
        match_rows = base_df[base_df["料號"] == original_item]
        if match_rows.empty: continue
            
        paired_found = False
        
        for _, m in match_rows.iterrows():
            if paired_found: break 

            main_car_count_rule = pd.to_numeric(m.get("車數"), errors="coerce") or 1
            if pd.isna(main_car_count_rule) or main_car_count_rule <= 0:
                main_car_count_rule = 1 
            
            d1_code = m["搭1料號"]
            d1_car_count = pd.to_numeric(m["搭1產出車數"], errors="coerce") or 0
            d2_code = m.get("搭2料號", "")
            d2_car_count = pd.to_numeric(m.get("搭2產出車數", ""), errors="coerce") or 0
            
            targets = []
            if d1_code: targets.append((d1_code, d1_car_count, main_car_count_rule))
            if d2_code: targets.append((d2_code, d2_car_count, main_car_count_rule))
            
            for target_code, sub_car_count, main_car_count in targets:
                if pd.isna(original_qty_val) or original_qty_val <= 0: continue
                if pd.isna(sub_car_count) or sub_car_count <= 0: continue 
                
                main_cut_count_paired = math.ceil(original_qty_val / main_car_count)
                if main_cut_count_paired <= 0: continue

                for df_src_ref, src_name in [(df_no_pair_split_mutable, "df_no_pair_split"),
                                             (df_no_pair_second_mutable, "df_no_pair_second")]: 
                    if paired_found: break
                    if df_src_ref.empty: continue
                        
                    match_in_df = df_src_ref[df_src_ref["品號"] == target_code].copy()
                    if match_in_df.empty: continue
                        
                    match_in_df["預計開工日"] = pd.to_datetime(match_in_df["預計開工日"], errors='coerce')
                    match_in_df = match_in_df.sort_values(by="預計開工日", ascending=True)

                    paired_row = match_in_df.iloc[0].copy()
                    original_paired_index = paired_row.name 
                        
                    alloc_qty = main_cut_count_paired * sub_car_count 
                    sub_remain_qty = pd.to_numeric(paired_row["預估良品數"], errors='coerce') or 0
                        
                    if alloc_qty > 0 and alloc_qty <= sub_remain_qty:
                        paired_found = True
                        leftover_qty = sub_remain_qty - alloc_qty
                        
                        log_entry = {
                            "主工單_SN": sn,
                            "主工單_品號": original_item,
                            "主工單_刀次": main_cut_count_paired, 
                            "主工單_車數": main_car_count,
                            "子工單_SN": paired_row.get("工單編號"),
                            "子工單_品號": target_code,
                            "來源池": src_name,
                            "子工單_車數": sub_car_count,
                            "子工單_消耗數量": alloc_qty,
                            "剩餘數量": leftover_qty,
                        }
                        debug_log.append(log_entry)

                        main_alloc_qty = main_cut_count_paired * main_car_count
                        
                        # 主工單新增項
                        paired_main_row = original_order.copy()
                        paired_main_row["預估良品數"] = main_alloc_qty
                        paired_main_row["餘量"] = main_alloc_qty
                        paired_main_row["刀次"] = main_cut_count_paired
                        paired_main_row["車數"] = main_car_count
                        # ⭐ 新增來源標註
                        paired_main_row["來源類型"] = "NEWLY_PAIRED_MAIN"
                        
                        # 子工單新增項
                        paired_sub_row = paired_row.copy()
                        paired_sub_row["預估良品數"] = alloc_qty
                        paired_sub_row["餘量"] = alloc_qty
                        paired_sub_row["刀次"] = 0
                        paired_sub_row["車數"] = sub_car_count
                        # ⭐ 新增來源標註
                        paired_sub_row["來源類型"] = "NEWLY_PAIRED_SUB"
                        
                        newly_paired_rows.extend([paired_main_row, paired_sub_row])
                        
                        # === 重新計算剩餘量及刀次（無條件進位） ===
                        if leftover_qty > 0:
                            new_cut = math.ceil(leftover_qty / sub_car_count)
                            df_src_ref.loc[original_paired_index, "刀次"] = new_cut
                            df_src_ref.loc[original_paired_index, "車數"] = sub_car_count
                            df_src_ref.loc[original_paired_index, "餘量"] = leftover_qty
                            df_src_ref.loc[original_paired_index, "預估良品數"] = new_cut * sub_car_count 
                        else:
                            df_src_ref.drop(original_paired_index, inplace=True)
                        # ====================================================================
                        break
                if paired_found: break
            if paired_found: break

    debug_df = pd.DataFrame(debug_log)
    
    # ----------------------------------------------------
    # 階段 B：還原 df_no_pair 系列工單的車數和刀次 
    # ----------------------------------------------------
    print(f"\n--- 階段 B：還原 df_no_pair 系列工單的車數和刀次 ---")
    
    df_no_pair_split_mutable = _restore_single_df(df_no_pair_split_mutable, base_df, original_car_map)
    df_no_pair_second_mutable = _restore_single_df(df_no_pair_second_mutable, base_df, original_car_map)

    # ⭐ 在合併前標註來源
    if not df_no_pair_split_mutable.empty:
        df_no_pair_split_mutable.loc[:, "來源類型"] = "NO_PAIR_SPLIT_LEFT"
    if not df_no_pair_second_mutable.empty:
        df_no_pair_second_mutable.loc[:, "來源類型"] = "NO_PAIR_SECOND_LEFT"
        
    print(f"還原後 df_no_pair_split 剩餘筆數: {len(df_no_pair_split_mutable)}")
    print(f"還原後 df_no_pair_second 剩餘筆數: {len(df_no_pair_second_mutable)}")
    
    # ----------------------------------------------------
    
    # ⭐ 標註原始已配對工單
    if df_paired_split is not None and not df_paired_split.empty:
        df_paired_split.loc[:, "來源類型"] = "ORIGINAL_PAIRED"

    df_paired_split_updated = pd.concat([df_paired_split, pd.DataFrame(newly_paired_rows)], ignore_index=True)
    
    # 合併所有資源池
    remaining_dfs = [df_paired_split_updated, df_no_pair_split_mutable, df_no_pair_second_mutable]
    
    remaining_dfs = [df.reset_index(drop=True) for df in remaining_dfs if (df is not None and not df.empty)]
    updated_unassigned_pool = pd.concat(remaining_dfs, ignore_index=True) if remaining_dfs else pd.DataFrame()
        
    print(f"\n--- 階段 A/B 總結 ---")
    print(f"成功配對筆數: {len(debug_df)} 筆")
    print(f"最終資源池 (all) 總工單數: {len(updated_unassigned_pool)} 筆 (包含已配對結果)")

    return {
        "all": updated_unassigned_pool, 
        "debug_log": debug_df, 
        "final_remaining": remaining, 
        "df_no_pair_second_mutable": df_no_pair_second_mutable,
        "df_no_pair_split_mutable": df_no_pair_split_mutable
    }


def _clean_cm(x):
    s = str(x).lower().replace("cm", "").strip()
    try:
        v = float(s)
        return int(v) if v.is_integer() else v
    except:
        return pd.to_numeric(x, errors="coerce")


def process_pair_order(pair_code, car_count, main_cut_count,
                       all_staff_df, original_order, final_rows, width_map,
                       df_paired_split=None, df_no_pair_split=None, df_no_pair_second=None):
    
    # 1. 安全檢查 (不需查找，因為 final_doAssignAndSort 已經判斷過)
    if not pair_code or str(pair_code).strip() == "" or pd.isna(pair_code) or car_count <= 0:
        # 如果 final_doAssignAndSort 傳遞了一個空的 pair_code，代表沒有找到搭配
        # 應生成空白行
        if not pair_code or str(pair_code).strip() == "":
            # 必須從 original_order 中取出預計要配對的品號和車數來生成空白行
            # ⚠️ 這裡需要知道原本在 final_doAssignAndSort 哪一行要配對，但這裡資訊不足。
            # 由於 final_doAssignAndSort 已經過濾並只傳遞了成功的品號，
            # 我們假設如果 pair_code 為空，就不生成任何行，而是讓 original_order 保持獨立。
            return
        
        # 由於 final_doAssignAndSort 只傳遞了成功找到配對的 pair_code，
        # 如果執行到這裡，代表 logic error。但根據您原來的程式碼，如果找不到，它會生成空白行。
        # 這裡假設如果 pair_code 存在但 car_count <= 0，則返回。
        return
    
    # --- 關鍵：計算要分配的數量 ---
    qty_to_use = main_cut_count * car_count
    width_value = width_map.get(str(pair_code))
    
    # 2. 查找工單 (此查找僅用於獲取工單詳細資訊，不進行數量檢查！)
    # 由於數量已在 final_doAssignAndSort 中扣除，此處查找工單是為了獲取其他元數據 (如日期, 人員等)
    # 我們查找的是 all_staff_df，且不再需要檢查刀次 > 0 (因為數量已在源頭扣除)
    candidates = all_staff_df[
        (all_staff_df["品號"] == pair_code)
        # 💡 注意: 這裡不能限制刀次 > 0，因為 paired_split 裡的工單刀次可能是 0。
    ].copy()

    if candidates.empty:
        # 沒找到工單，但 final_doAssignAndSort 說找到了搭配
        # 這種情況應視為原工單已配完，現在只是要生成搭配行
        # 如果要補空白，則照原邏輯 (但這會生成過多空白行，請確保邏輯一致)
        print(f"   [PairOrder] ⚠️ 找不到 {pair_code} 的詳細資訊，生成空白搭配行。")
        new_row = original_order.copy()
        new_row["工單編號"] = ""
        new_row["品號"] = pair_code
        new_row["車數"] = car_count
        new_row["刀次"] = 0
        new_row["預估良品數"] = qty_to_use
        new_row["餘量"] = new_row["預估良品數"]
        new_row["客戶需求日"] = ""
        if width_value is not None: new_row["公分"] = width_value
        final_rows.append(new_row)
        return

    # 取第一筆候選 (用來複製元數據)
    paired_row = candidates.iloc[0].copy()
    main_order_id = paired_row.get("工單編號", "")

    # 3. 確定搭配行資訊 (簡化，因為數量已在 final_doAssignAndSort 中處理)
    
    # --- 生成搭配行 ---
    paired_row["刀次"] = 0
    paired_row["車數"] = car_count
    paired_row["預估良品數"] = qty_to_use # 這裡直接使用 final_doAssignAndSort 計算出的數量
    paired_row["餘量"] = qty_to_use
    if width_value is not None:
        paired_row["公分"] = width_value
        
    # 確保工單編號和日期是正確的 (因為它從 all_staff_df 來的，可能是舊數據)
    # 我們假設 final_doAssignAndSort 已經將日期同步到 original_order
    # 這裡應該使用找到的工單ID和日期
    
    final_rows.append(paired_row)
    print(f"   [PairOrder] ✅ 為主工單 {original_order['工單編號']} 成功生成子搭配行 {main_order_id} ({pair_code})，數量: {qty_to_use}")

    # 🛑 刪除所有 in_no_pair 的複雜檢查邏輯，避免重複配對和數量錯誤
    # 這些檢查和數量扣除已經在 final_doAssignAndSort 中處理完畢。
    return


def split_large_batches(rows, max_cut=55):
    split_result = []
    i = 0
    while i < len(rows):
        main_row = rows[i]
        # 預設子工單為 None
        sub_row = rows[i + 1] if (i + 1) < len(rows) and rows[i + 1].get("刀次", None) == 0 else None
            
        main_cut_raw = main_row.get("刀次", 0)
        main_cut = pd.to_numeric(main_cut_raw, errors="coerce")

        if pd.isna(main_cut):
            main_cut = 0
        else:
            main_cut = int(main_cut)

        # 若刀次沒超過限制或無子工單，直接加入
        if main_cut <= max_cut or sub_row is None:
            split_result.append(main_row)
            if sub_row is not None:
                split_result.append(sub_row)
            i += 2 if sub_row is not None else 1
            continue

        # 刀次超過 max_cut 的情況拆分
        remaining_cut = main_cut
        main_qty = main_row.get("預估良品數", 0)
        sub_qty = sub_row.get("預估良品數", 0) if sub_row is not None else 0

        while remaining_cut > 0:
            current_cut = min(max_cut, remaining_cut)
            ratio = current_cut / main_cut

            new_main = main_row.copy()
            new_main["刀次"] = current_cut
            new_main["預估良品數"] = round(main_qty * ratio)
            new_main["餘量"] = new_main["預估良品數"]
            split_result.append(new_main)

            if sub_row is not None:
                new_sub = sub_row.copy()
                new_sub["刀次"] = 0
                new_sub["預估良品數"] = round(sub_qty * ratio)
                new_sub["餘量"] = new_sub["預估良品數"]
                split_result.append(new_sub)

            remaining_cut -= current_cut

        i += 2 if sub_row is not None else 1

    return split_result

# 品號排序
# 相似品號"盡量"會擺在同一時期生產
def sort_by_customer_due_date(df: pd.DataFrame, break_dates: set = None) -> pd.DataFrame:
    df = df.copy().reset_index(drop=True)
    df["預計開工日"] = pd.to_datetime(df["預計開工日"], errors='coerce')

    # 清理「客戶需求日」欄位中異常值
    if "客戶需求日" in df.columns:
        df["客戶需求日"] = df["客戶需求日"].replace(["0", 0, "0000/00/00", "NaT"], "")

    if break_dates is None:
        break_dates = set()

    # 可設定的品號前綴規則 因應各種不同品號
    prefix_rules = {
        "B110": 5,
        "TTR": 4,
        "UTMX": 4
    }
    default_prefix_len = 3

    def next_workdays(start_date, n):
        days = []
        d = start_date
        while len(days) <= n:  # 改成包含 start_date + 後續 n 天
            if d.weekday() < 5 and d not in break_dates:
                days.append(d)
            d += timedelta(days=1)
        return days

    result_frames = []
    idx = 0
    while idx < len(df):
        start_date = df.at[idx, "預計開工日"].date()
        workdays = next_workdays(start_date, 7)
        end_date = workdays[-1]

        mask = (df["預計開工日"].dt.date >= start_date) & (df["預計開工日"].dt.date <= end_date)
        chunk = df[mask].copy()

        if chunk.empty:
            # 沒有資料就跳到下一個日期
            next_start_idx = df.index[df["預計開工日"].dt.date > start_date].min()
            if pd.isna(next_start_idx):
                break
            idx = next_start_idx
            continue

        # 主子工單群組 (保持原順序)
        groups = []
        i = 0
        while i < len(chunk):
            row = chunk.iloc[i]
            group = [row]
            if row["刀次"] > 0:
                i += 1
                while i < len(chunk) and chunk.iloc[i]["刀次"] == 0:
                    group.append(chunk.iloc[i])
                    i += 1
            else:
                i += 1
            groups.append(group)

        # 品號分組（保留 others 在原 chunk 順序）
        grouped_by_code = defaultdict(list)
        others = []

        for g in groups:
            pn = str(g[0]["品號"]) if pd.notna(g[0]["品號"]) else None
            if pn:
                match_len = next((length for prefix, length in prefix_rules.items() if pn.startswith(prefix)), default_prefix_len)
                code = pn[:match_len]
                grouped_by_code[code].append(g)
            else:
                others.append(g)

        # 依 code 排序，每個 code 內保留原 chunk 順序
        sorted_groups = []
        for code in sorted(grouped_by_code.keys()):
            sorted_groups.extend(grouped_by_code[code])
        # others 保留在原位置
        sorted_groups.extend(others)

        sorted_chunk = pd.DataFrame([row for group in sorted_groups for row in group])
        result_frames.append(sorted_chunk)

        # 跳到下一個日期區間的第一筆
        next_start_idx = df.index[df["預計開工日"].dt.date > end_date].min()
        if pd.isna(next_start_idx):
            break
        idx = next_start_idx

    final_df = pd.concat(result_frames, ignore_index=True)
    
    # === 以上為品號分類 ===
    # === 合併同工單+品號（主工單刀次不為0）的群組 ===
    groups = []
    i = 0
    while i < len(final_df):
        row = final_df.iloc[i]
        group = [row]
        if row["刀次"] > 0:
            i += 1
            while i < len(final_df) and final_df.iloc[i]["刀次"] == 0:
                group.append(final_df.iloc[i])
                i += 1
        else:
            i += 1
        groups.append(group)

    key_to_groups = defaultdict(list)
    for idx, group in enumerate(groups):
        main = group[0]
        if pd.notna(main["工單編號"]) and pd.notna(main["品號"]) and main["刀次"] != 0:
            key = (main["工單編號"], main["品號"])
            key_to_groups[key].append((idx, group))

    used_indices = set()
    new_groups = []
    i = 0
    while i < len(groups):
        if i in used_indices:
            i += 1
            continue

        group = groups[i]
        main = group[0]
        key = (main["工單編號"], main["品號"])

        if key in key_to_groups and len(key_to_groups[key]) > 1:
            related = key_to_groups[key]
            base_idx, _ = related[0]
            if i == base_idx:
                merged_group = []
                for idx2, g in related:
                    if idx2 not in used_indices:
                        merged_group.extend(g)
                        used_indices.add(idx2)
                new_groups.append(merged_group)
            i += 1
        else:
            new_groups.append(group)
            i += 1

    final_df = pd.DataFrame([row for group in new_groups for row in group]).reset_index(drop=True)

    # === 根據每區最早的「客戶需求日」來整體排序 ===
    block = []
    temp_group = []
    for row in final_df.itertuples(index=False):
        if row.刀次 > 0 and temp_group:
            block.append(temp_group)
            temp_group = []
        temp_group.append(row)
    if temp_group:
        block.append(temp_group)

    def get_earliest_due(group):
        dates = [pd.to_datetime(r.客戶需求日, errors='coerce') for r in group if pd.notna(r.客戶需求日)]
        return min(dates) if dates else pd.Timestamp.max

    block.sort(key=get_earliest_due)

    final_df = pd.DataFrame([r._asdict() for group in block for r in group])

    # 日期欄轉格式
    final_df["預計開工日"] = pd.to_datetime(final_df["預計開工日"], errors='coerce')
    final_df["預計開工日"] = final_df["預計開工日"].dt.strftime("%Y/%#m/%#d")

    # 客戶需求日移到最前面
    cols = final_df.columns.tolist()
    if "客戶需求日" in cols:
        cols.insert(0, cols.pop(cols.index("客戶需求日")))
        final_df = final_df[cols]

    return final_df



def reorder_main_and_subs(df: pd.DataFrame) -> pd.DataFrame:
    """
    將每筆主工單與其後子工單重新排序，
    依序排列成「主工單 → 子工單 → 下一筆主工單 → 子工單」。
    子工單刀次為0，良品數 = 車數 * 對應主工單刀次。
    (已修正：使用 safe_int_conversion 確保 "車數" 和 "刀次" 轉換安全。)
    """
    df = df.copy().reset_index(drop=True)
    new_rows = []
    i = 0
    while i < len(df):
        row = df.iloc[i].copy()
        
        # 修正點 A: 使用 safe_int_conversion 確保主工單的刀次和車數安全
        current_cut = safe_int_conversion(row.get("刀次", 0))
        current_car = safe_int_conversion(row.get("車數", 0))

        if current_cut > 0:
            # 主工單
            main_row = row.copy()
            new_rows.append(main_row)
            
            # 找子工單
            j = i + 1
            while j < len(df):
                sub_row = df.iloc[j].copy()
                
                # 修正點 B: 使用 safe_int_conversion 確保子工單的刀次檢查安全
                sub_cut = safe_int_conversion(sub_row.get("刀次", 0))
                if sub_cut != 0:
                    break # 遇到下一張主工單，跳出子工單迴圈

                # 修正點 C (錯誤發生處): 使用 safe_int_conversion 確保子工單的車數安全
                sub_car = safe_int_conversion(sub_row.get("車數", 0))
                
                # 計算子工單預估良品數
                sub_row["預估良品數"] = sub_car * current_cut
                new_rows.append(sub_row)
                j += 1
                
            i = j # 將主迴圈索引跳到下一張主工單
        else:
            # 刀次=0，但沒有對應主工單 (孤立的子工單或單純刀次為0的工單)
            new_rows.append(row.copy())
            i += 1
            
    # 建立新的 DataFrame
    out_df = pd.DataFrame(new_rows)
    
    # 確保輸出欄位順序一致（可選，但保持習慣）
    if not df.empty and not out_df.empty:
        out_df = out_df[df.columns.intersection(out_df.columns).tolist() + 
                         [c for c in out_df.columns if c not in df.columns]]
    elif df.empty:
        return pd.DataFrame(columns=df.columns)
        
    return out_df.reset_index(drop=True)




def merge_same_day_orders_multi(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    df = df.copy()
    df["預計開工日"] = pd.to_datetime(df["預計開工日"])
    merged_rows = []

    i = 0
    while i < len(df):
        row = df.iloc[i]
        row_date = row["預計開工日"]

        if row["刀次"] > 0:
            main_cut = row["刀次"]
            main_car = row["車數"]
            main_id = row["工單編號"]

            # 找對應子工單
            j = i + 1
            subs = []
            while j < len(df) and df.iloc[j]["刀次"] == 0 and df.iloc[j]["預計開工日"] == row_date:
                sub_id = str(df.iloc[j]["工單編號"]).strip()
                if sub_id == "":
                    sub_id = "EMPTY_SUB"
                subs.append((sub_id, df.iloc[j]))
                j += 1

            merged = False
            if merged_rows:
                # 找前一筆主工單
                prev_idx = -1
                for k in range(len(merged_rows)-1, -1, -1):
                    if merged_rows[k]["刀次"] > 0:
                        prev_idx = k
                        break
                if prev_idx >= 0:
                    prev_main = merged_rows[prev_idx]
                    prev_main_id = prev_main["工單編號"]
                    prev_date = prev_main["預計開工日"]

                    # 找前一組子工單
                    prev_subs = []
                    m = prev_idx + 1
                    while m < len(merged_rows) and merged_rows[m]["刀次"] == 0 and merged_rows[m]["預計開工日"] == prev_date:
                        prev_sub_id = str(merged_rows[m]["工單編號"]).strip()
                        if prev_sub_id == "":
                            prev_sub_id = "EMPTY_SUB"
                        prev_subs.append((prev_sub_id, merged_rows[m]))
                        m += 1

                    # 判斷是否可合併
                    if prev_main_id == main_id and prev_date == row_date:
                        if len(subs) == 0 and len(prev_subs) == 0:
                            # 重新計算主工單
                            new_cut = prev_main["刀次"] + main_cut
                            merged_rows[prev_idx]["刀次"] = new_cut
                            merged_rows[prev_idx]["良品數"] = merged_rows[prev_idx]["車數"] * new_cut
                            merged = True
                        elif [s[0] for s in subs] == [ps[0] for ps in prev_subs]:
                            new_cut = prev_main["刀次"] + main_cut
                            merged_rows[prev_idx]["刀次"] = new_cut
                            merged_rows[prev_idx]["良品數"] = merged_rows[prev_idx]["車數"] * new_cut
                            for k, (_, sub_row) in enumerate(prev_subs):
                                merged_rows[prev_idx + 1 + k]["良品數"] = merged_rows[prev_idx + 1 + k]["車數"] * new_cut
                            merged = True

            if not merged:
                # 新增主工單
                new_main = row.copy()
                new_main["良品數"] = main_car * main_cut
                merged_rows.append(new_main)
                # 新增子工單
                for _, sub in subs:
                    new_sub = sub.copy()
                    new_sub["良品數"] = new_sub["車數"] * main_cut
                    merged_rows.append(new_sub)

            i = j
        else:
            merged_rows.append(row.copy())
            i += 1

    merged_df = pd.DataFrame(merged_rows)
    merged_df["預計開工日"] = merged_df["預計開工日"].dt.strftime("%Y-%m-%d")

    # -----------------------------
    # 🔹 統一重新計算良品數與餘量
    # -----------------------------
    recalculated = []
    i = 0
    while i < len(merged_df):
        row = merged_df.iloc[i].copy()
        if row["刀次"] > 0:  # 主工單
            main_cut = row["刀次"]
            main_car = row["車數"]

            row["預估良品數"] = main_car * main_cut
            row["餘量"] = row["預估良品數"]
            recalculated.append(row)

            # 處理子工單
            j = i + 1
            while j < len(merged_df) and merged_df.iloc[j]["刀次"] == 0 and merged_df.iloc[j]["預計開工日"] == row["預計開工日"]:
                sub = merged_df.iloc[j].copy()
                sub["預估良品數"] = sub["車數"] * main_cut
                sub["餘量"] = sub["預估良品數"]
                recalculated.append(sub)
                j += 1

            i = j
        else:
            # 孤立子工單
            row["預估良品數"] = row["車數"] * row["刀次"]
            row["餘量"] = row["預估良品數"]
            recalculated.append(row)
            i += 1

    return pd.DataFrame(recalculated)


def safe_int_conversion(value: Any, default: int = 0) -> int:
    """
    安全地將值轉換為整數。處理 None, NaN, 空字串等情況，避免 ValueError。
    (已提升至頂層，供所有函式共用。)
    """
    # 檢查是否為缺失值 (NaN, None) 或空字串
    if pd.isna(value) or value is None or str(value).strip() == "":
        return default
    try:
        # 先轉為 float 處理科學記號或奇怪格式，再轉為 intS
        return int(float(value))
    except (ValueError, TypeError):
        return default


def merge_order_cutNum(A_df: pd.DataFrame, B_df: pd.DataFrame, C_df: pd.DataFrame,
                       daily_standard: int = 165) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    將每個人 (A_df, B_df, C_df) 的工單往前遞補滿每日標準刀次。
    修正邏輯：每次補刀時即時計算當日品號種類數與新上限，避免超過限制。
    """

    MAX_KNIFE_MAP = {1:65, 2:55, 3:51, 4:47, 5:43, 6:39}
    prefix_rules = {"B110":5, "TTR":4, "UTMX":4, "UTM":3}
    default_prefix_len = 3

    def get_item_group_key(item_code: str) -> str:
        if not isinstance(item_code, str):
            return "UNKNOWN"
        item_code = item_code.strip().replace('.', '')
        for prefix in sorted(prefix_rules.keys(), key=len, reverse=True):
            length = prefix_rules[prefix]
            if item_code.startswith(prefix):
                return item_code[:length] if len(item_code) >= length else item_code
        return item_code[:default_prefix_len] if len(item_code) >= default_prefix_len else item_code

    def process_one(df: pd.DataFrame, person_name: str) -> pd.DataFrame:
        if df is None or df.empty:
            return df.copy()
       
        df = df.copy().reset_index(drop=True)
        df["預計開工日"] = pd.to_datetime(df["預計開工日"], errors="coerce")
        df["預計完工日"] = pd.to_datetime(df["預計完工日"], errors="coerce")

        # 初始化 orders
        orders = []
        last_main = None
        for i, row in df.iterrows():
            knives = safe_int_conversion(row.get("刀次",0))
            cars = safe_int_conversion(row.get("車數",0))
            od = {"orig_idx": i, "orig_row": row.copy(), "remain": knives, "car": cars, "children": [], "slices": {}}
            if knives > 0:
                od["slices"][row["預計開工日"].normalize()] = knives
                last_main = od
                orders.append(od)
            else:
                if last_main is not None:
                    last_main["children"].append(row.copy())
                else:
                    orders.append(od)
                    last_main = od

        if not orders:
            return df

        all_dates = sorted({d.normalize() for od in orders for d in od["slices"].keys() if pd.notna(d)})

        # 移除原程式中冗餘的 daily_status 初始化區塊

        # === 核心排程迴圈：使用 while 進行動態日期管理 ===
        date_idx = 0
        while date_idx < len(all_dates):
            current_date = all_dates[date_idx]
            
            # 1. 計算當日初始狀態
            daily_ods = [od for od in orders if current_date in od["slices"] and od["slices"][current_date]>0]
            total_knives = sum(od["slices"][current_date] for od in daily_ods)
            existing_keys = {get_item_group_key(od["orig_row"].get("品號")) for od in daily_ods}
            
            # 2. 計算即時上限 (使用 MAX_KNIFE_MAP)
            num_items = len(existing_keys)
            day_limit = MAX_KNIFE_MAP.get(num_items, 65 if num_items==0 else 39)

            # --- PHASE 1: 新增【超額遞延】邏輯 (Push-back Excess) ---
            excess = total_knives - day_limit
            
            if excess > 0:
                # 確定遞延目標：找出當日排程中，刀次最大的工單優先推遲
                ods_to_push = sorted(
                    daily_ods, 
                    key=lambda od: (-od["slices"][current_date], od["orig_idx"]) 
                )
                
                # 找到下一個排程日 (後一天)
                next_date = current_date + pd.Timedelta(days=1)
                
                # 確保 next_date 在 all_dates 列表中，如果沒有則加入並重新排序
                if next_date not in all_dates:
                    all_dates.append(next_date)
                    all_dates.sort()
                    # 註：此處不需要更新 date_idx，因為我們是將刀次往未來推。

                pushed_amount = 0
                for od in ods_to_push:
                    if excess <= 0:
                        break
                        
                    # 決定推遲的刀次：不超過工單當日刀次，也不超過總溢出量
                    push_target = min(od["slices"][current_date], excess)
                    
                    # 執行遞延：將刀次從 current_date 移到 next_date
                    od["slices"][current_date] -= push_target
                    # next_date 的刀次增加
                    od["slices"][next_date] = od["slices"].get(next_date, 0) + push_target 
                    
                    excess -= push_target
                    pushed_amount += push_target
                    
                # 更新狀態 (total_knives 已經減少 pushed_amount)
                total_knives -= pushed_amount

            # --- PHASE 2: 原始【不足遞補】邏輯 (Pull-in Under-Capacity) ---
            
            # 使用經過遞延修正後的 total_knives 重新計算可用容量
            available = max(day_limit - total_knives, 0)
            
            # 找未來工單 (必須從 current_date 之後的所有排程中尋找)
            future_cand = []
            for od in orders:
                for sdate, amt in od["slices"].items():
                    # 條件：日期必須晚於 current_date 且 amt > 0
                    if pd.notna(sdate) and sdate > current_date and amt>0:
                        future_cand.append((sdate, od, amt))
            future_cand.sort(key=lambda x:(x[0], x[1]["orig_idx"]))

            # 補刀邏輯 (原程式碼，保持不變)
            while future_cand:
                pick_date, pick_od, pick_amt = future_cand[0]
                pick_key = get_item_group_key(pick_od["orig_row"].get("品號"))

                # 預測加入這筆工單後的品號集合與新上限
                new_existing_keys = existing_keys | {pick_key}
                new_num_items = len(new_existing_keys)
                new_day_limit = MAX_KNIFE_MAP.get(new_num_items, 39)

                # 計算當天剩餘可補刀次
                remaining_capacity = new_day_limit - total_knives
                if remaining_capacity <= 0:
                    break

                # 計算補刀量，不超過剩餘容量或工單本身刀次
                fill_target = min(pick_amt, remaining_capacity)
                if fill_target <= 0:
                    break

                # 執行遞補
                pick_od["slices"][pick_date] -= fill_target
                if pick_od["slices"][pick_date] <= 0:
                    del pick_od["slices"][pick_date]
                pick_od["slices"][current_date] = pick_od["slices"].get(current_date, 0) + fill_target

                total_knives += fill_target
                existing_keys.add(pick_key)

                # 更新候選
                future_cand = [(d, o, a) for d, o, a in future_cand if o["slices"].get(d, 0) > 0]
            
            # 處理下一天
            date_idx += 1


        # 4. 展開成 DataFrame (保持原樣)
        out_rows = []
        for od in orders:
            for sdate, amt in sorted(od["slices"].items(), key=lambda x:x[0]):
                if amt>0:
                    # 主工單
                    main_row = od["orig_row"].copy()
                    main_row["預計開工日"] = sdate
                    main_row["預計完工日"] = sdate
                    main_row["刀次"] = int(amt)
                    main_car = safe_int_conversion(main_row.get("車數",0))
                    main_row["預估良品數"] = main_car*int(amt)
                    out_rows.append(main_row)
                    
                    # 子工單
                    for ch in od["children"]:
                        ch_row = ch.copy()
                        ch_row["預計開工日"] = sdate
                        ch_row["預計完工日"] = sdate
                        ch_row["刀次"] = 0
                        ch_car = safe_int_conversion(ch_row.get("車數",0))
                        # 良品數仍基於 amt (主工單刀次) 計算
                        ch_row["預估良品數"] = ch_car*int(amt) 
                        if "餘量" in ch_row.index:
                            ch_row["餘量"] = ch_row["預估良品數"]
                        out_rows.append(ch_row)

        out_df = pd.DataFrame(out_rows)
        out_df["預計開工日"] = out_df["預計開工日"].dt.strftime("%Y-%m-%d")
        out_df["預計完工日"] = out_df["預計完工日"].dt.strftime("%Y-%m-%d")
        return out_df.reset_index(drop=True)

    A_out = process_one(A_df, "容合 (A830)")
    B_out = process_one(B_df, "旺斌 (B201)")
    C_out = process_one(C_df, "家偉 (A159)")

    return A_out, B_out, C_out




def Erin_use(A830_df, B201_df, A159_df, base_df):
    """
    此函式主要處理 Erin 提出的兩點需求：
    1. 根據品號，從基本資料檔中查找並補上對應的「米平方」欄位。
    2. 新增「週分組」欄位，定義每週為「週五到隔週四」，並依此排序。
    """
    print("正在執行新增需求：加入米平方與週分組...")

    # --- 步驟 1: 讀取基本資料以取得米平方 ---
    try:   
        # 只選取需要的欄位進行合併，更高效
        m2_df = base_df[['料號', '面積  M2']].copy()
        m2_df = m2_df.drop_duplicates(subset=['料號'], keep='first').copy()
    except Exception as e:
        print(f"讀取基本資料時發生錯誤，無法新增米平方：{e}")
        return A830_df, B201_df, A159_df
    
    # --- 步驟 2 & 3: 處理每個員工的 DataFrame ---
    processed_dfs = []
    staff_dfs = [A830_df, B201_df, A159_df]

    for df in staff_dfs:
        if df is None or df.empty:
            processed_dfs.append(df)
            continue

        # 在處理米平方之前，先檢查並移除 '良品數' 欄位
        if '良品數' in df.columns:
            df = df.drop(columns=['良品數'])

        # --- 2. 加入米平方欄位 ---
        #  修改 #2：使用 left_on 和 right_on 來指定不同的欄位名稱 
        # 左邊 df 用 '品號', 右邊 m2_df 用 '料號'
        merged_df = pd.merge(df, m2_df, left_on='品號', right_on='料號', how='left')

        #  修改 #3：合併後，把多餘的 '料號' 欄位刪除 
        if '料號' in merged_df.columns:
            merged_df = merged_df.drop(columns=['料號'])

        # 1. 將米平方欄位轉為數字，若有錯則變為空值(NaN)
        merged_df['面積  M2'] = pd.to_numeric(merged_df['面積  M2'], errors='coerce')

        # 2. 進行傳統四捨五入，並轉為支援空值的整數型別
        #    (這裡使用一個小技巧：先加0.5再轉成整數，來實現傳統四捨五入)
        merged_df['面積  M2'] = np.floor(merged_df['面積  M2'] + 0.5)

        # 3. 轉成支援空值的整數型別
        merged_df['面積  M2'] = merged_df['面積  M2'].astype('Int64')

        # --- 3. 新增週分組欄位 ---
        merged_df['開工日_dt'] = pd.to_datetime(merged_df['預計開工日'], errors='coerce')

        def get_week_start_friday(date):
            if pd.isna(date):
                return None
            if date.weekday() >= 4:
                return date - pd.Timedelta(days=date.weekday() - 4)
            else:
                return date - pd.Timedelta(days=date.weekday() + 3)

        merged_df['週分組_起始日'] = merged_df['開工日_dt'].apply(get_week_start_friday)
        
        def format_week_range(start_date):
            if pd.isna(start_date):
                return ""
            end_date = start_date + pd.Timedelta(days=6)
            return f"{start_date.strftime('%Y/%m/%d')} ~ {end_date.strftime('%Y/%m/%d')}"
        
        merged_df['週分組'] = merged_df['週分組_起始日'].apply(format_week_range)

        merged_df = merged_df.rename(columns={'面積  M2': '米平方'})

        # --- 4. 重新排序並整理欄位 ---
        sorted_df = merged_df.sort_values(by=['週分組_起始日', '開工日_dt'], ascending=[True, True])
        sorted_df = sorted_df.drop(columns=['開工日_dt', '週分組_起始日'])
        
        cols = sorted_df.columns.tolist()
        week_col = cols.pop(cols.index('週分組'))
        m2_col = cols.pop(cols.index('米平方'))
        cols.insert(0, week_col)
        #cols.append(m2_col)
        cols.insert(6, m2_col)
        final_df = sorted_df[cols]

        processed_dfs.append(final_df)

    print("✅ 新增需求處理完成。")
    return processed_dfs[0], processed_dfs[1], processed_dfs[2]

def remaining_paired_detail(df_paired_split, remaining, base_df):

    df_paired_split = df_paired_split.copy()

    remaining["預計開工日"] = pd.to_datetime(remaining["預計開工日"], errors='coerce')
    remaining = remaining.sort_values(by="預計開工日", ascending=True).copy().reset_index(drop=True)

    # 建立 remaining 工單剩餘量 map
    remaining_qty_map = remaining.set_index("工單編號")["預估良品數"].to_dict()
    
    # 建立寬度 map
    # 這裡假設 _clean_cm 已定義
    width_map = dict(zip(base_df["料號"], base_df["寬度Cm"].apply(_clean_cm)))

    paired_ids = set()

    for idx, main_row in remaining.iterrows():
        main_id = main_row["工單編號"]
        if main_id in paired_ids:
            continue

        main_item = main_row["品號"]
        main_remain_qty = remaining_qty_map.get(main_id, 0)
        if main_remain_qty <= 0:
            continue

        # 查基本資料找到搭1/搭2
        row_match = base_df[base_df["料號"] == main_item]
        if row_match.empty:
            continue

        paired_successfully = False
        
        # 🚨 預設變數，用於儲存成功的搭配車數
        main_output_car = 0
       
        # --- 遍歷所有可能的搭配方案 ---
        for i, match_row in row_match.iterrows():
            if paired_successfully: break # 找到搭配，跳出基本資料的其他方案
                
            d1_code = match_row["搭1料號"]
            d1_car_count = pd.to_numeric(match_row["搭1產出車數"], errors="coerce") or 0
            d2_code = match_row.get("搭2料號", "")
            d2_car_count = pd.to_numeric(match_row.get("搭2產出車數", ""), errors="coerce") or 0
            
            # 尋找 remaining 中可以搭配的子工單
            candidates = remaining[
                (remaining["品號"].isin([d1_code, d2_code])) &
                (~remaining["工單編號"].isin(paired_ids)) &
                (remaining["預估良品數"] > 0) &
                (remaining["工單編號"] != main_id) # ⚠️ 確保不自己搭自己
            ]

            if candidates.empty:
                continue # 嘗試下一種搭配方案

            # 找到第一個子工單（因 remaining 已排序，這將是日期最早的子工單）
            sub_row = candidates.iloc[0].copy()
            sub_id = sub_row["工單編號"]
            sub_item = sub_row["品號"]
            
            # 搭配車數來源於當前 match_row
            sub_car = d1_car_count if sub_item == d1_code else d2_car_count
            
            # 🚨 關鍵修正 1：主工單結果行的車數必須使用搭配方案的車數
            main_output_car = pd.to_numeric(match_row.get("車數", 1), errors="coerce") or 1
            
            # 原始工單的車數 (用於計算最大刀次)
            main_car = int(main_row["車數"])
            
            # 標記成功並跳出外層遍歷
            paired_successfully = True
            break 
        
        if not paired_successfully:
            continue # 如果所有基本資料的搭配方案都找不到子工單，則跳過主工單

        # --- 計算分配刀次 (使用搭配方案的車數) ---
        
        # ⚠️ 修正：這裡的主工單應使用搭配方案的車數 (main_output_car) 來計算最大刀次
        max_main_cut = main_remain_qty // main_output_car if main_output_car > 0 else 0 
        
        max_sub_cut = remaining_qty_map[sub_id] // sub_car if sub_car > 0 else 0
        split_cut = min(max_main_cut, max_sub_cut)
        
        if split_cut <= 0:
            continue

        # 分配量 (使用搭配方案的車數進行分配)
        alloc_main_qty = split_cut * main_output_car
        alloc_sub_qty = split_cut * sub_car

        # 更新剩餘量
        remaining_qty_map[main_id] -= alloc_main_qty
        remaining_qty_map[sub_id] -= alloc_sub_qty

        # --- [ 新增除錯 Print 區塊 ] ---
        print(f"\n>>> 成功配對 [SN: {main_id} 與 {sub_id}]")
        print(f"主工單 | SN: {main_id} | 料號: {main_item} | 車數: {main_output_car} | 刀次: {split_cut} | 分配量: {alloc_main_qty}")
        print(f"子工單 | SN: {sub_id} | 料號: {sub_item} | 車數: {sub_car} | 刀次: 0 | 分配量: {alloc_sub_qty}")
        print(f"資源池 | 主剩餘: {remaining_qty_map[main_id]} | 子剩餘: {remaining_qty_map[sub_id]}")
        # --- [ 區塊結束 ] ---

        # --- 生成結果行 (使用搭配方案的車數) ---
        
        # 主工單 row
        paired_main = main_row.copy()
        paired_main["預估良品數"] = alloc_main_qty
        paired_main["餘量"] = alloc_main_qty
        paired_main["刀次"] = split_cut
        # 🚨 關鍵修正 2：更新 paired_main 的車數
        paired_main["車數"] = main_output_car 
        if width_map.get(main_item):
            paired_main["公分"] = width_map[main_item]

        # 子工單 row
        paired_sub = sub_row.copy()
        paired_sub["預估良品數"] = alloc_sub_qty
        paired_sub["餘量"] = alloc_sub_qty
        paired_sub["刀次"] = 0
        # 這裡原本就正確使用 sub_car，保持不變
        paired_sub["車數"] = sub_car
        if width_map.get(sub_item):
            paired_sub["公分"] = width_map[sub_item]

        # 放入 df_paired_split
        df_paired_split = pd.concat([df_paired_split, pd.DataFrame([paired_main, paired_sub])], ignore_index=True)

        # 如果主子工單都配完，標記
        if remaining_qty_map[main_id] <= 0:
            paired_ids.add(main_id)
        if remaining_qty_map[sub_id] <= 0:
            paired_ids.add(sub_id)

    # ... (更新 remaining 邏輯保持不變) ...
    new_remaining = []
    for idx, row in remaining.iterrows():
        # ⚠️ 這裡需要確保 remaining 內的工單的車數也被更新
        remain_qty = remaining_qty_map.get(row["工單編號"], 0)
        
        # ... (省略：如果這張工單是配對成功後還有剩餘的，它的車數應該維持舊的還是新的？
        #      由於 remaining 只是剩餘池，通常會維持原樣，直到下次作為主工單才計算。
        #      因此，我們只更新數量，不更新 remaining 內的車數。)
        
        if remain_qty > 0:
            new_row = row.copy()
            new_row["預估良品數"] = remain_qty
            new_row["餘量"] = remain_qty
            new_remaining.append(new_row)

    remaining_df = pd.DataFrame(new_remaining).reset_index(drop=True)
    
    return df_paired_split, remaining_df



def check_Qty(A830_part_end, B201_part_end, A159_part_end):

    try:
        if "GLOBAL_ORDER_QTY_MAP" not in globals():
            print("⚠️ 找不到 GLOBAL_ORDER_QTY_MAP，略過防呆檢查。")
            return False

        # 合併三個人員的資料
        all_staff_df = pd.concat([A830_part_end, B201_part_end, A159_part_end], ignore_index=True)

        # 確保數值欄位是數字
        all_staff_df["餘量"] = pd.to_numeric(all_staff_df.get("餘量", 0), errors="coerce").fillna(0)

        # 檢查每個工單是否滿足計畫數量
        for order_id, planned_qty in GLOBAL_ORDER_QTY_MAP.items():
            planned_qty = float(planned_qty)
            total_actual = all_staff_df.loc[all_staff_df["工單編號"] == order_id, "餘量"].sum()

            if total_actual < planned_qty:
                print(f"⚠️ 工單 {order_id} 開工數量 {planned_qty:.0f}，實際加總僅 {total_actual:.0f}（差 {planned_qty - total_actual:.0f}）")
                return False

        # 所有工單都沒缺少
        return True

    except Exception as e:
        print(f"❌ 防呆檢查時發生錯誤：{e}")
        return False



def check_df_paired_split_2(df_paired, extra_remaining, base_df):

    df_paired = df_paired.copy().reset_index(drop=True)
    
    anomalous_rows = []
    indices_to_keep = []
    
    # 【內嵌函數 - 檢查品號前三碼】
    def _check_prefix_match(str1, str2, prefix_length=3):
        if pd.isna(str1) or pd.isna(str2): return False
        str1 = str(str1).strip()
        str2 = str(str2).strip()
        if len(str1) >= prefix_length and len(str2) >= prefix_length:
            return str1[:prefix_length] == str2[:prefix_length]
        return False
        
    i = 0
    while i < len(df_paired):
        # ... (此處的 while 迴圈邏輯與您上一個回覆完全相同，邏輯正確)
        main_index = i
        main_row = df_paired.iloc[main_index]
        
        # --- 類型轉換容錯 ---
        try:
            current_cut = float(main_row.get("刀次", 0))
        except (ValueError, TypeError):
            current_cut = 0.0
        # ---------------------------------------------
        
        
        # ---------------------------------------------
        # 1. 檢查刀次是否為 0
        # ---------------------------------------------
        if current_cut == 0:
            
            # 刀次為 0，需要進行品號前三碼判斷
            is_prefix_match_result = False
            
            if indices_to_keep: # 檢查是否有被保留的索引 (即前面有非零刀次工單)
                # 取得最近一個保留工單的索引 (即最近一個非零刀次的工單)
                prev_non_zero_index = indices_to_keep[-1] 
                prev_non_zero_row = df_paired.iloc[prev_non_zero_index]
                
                current_item_code = main_row["品號"]
                prev_item_code = prev_non_zero_row["品號"]
                
                # 💥 使用內嵌邏輯判斷品號前三碼是否相同
                if _check_prefix_match(current_item_code, prev_item_code):
                    is_prefix_match_result = True
            
            # 根據相似度決定是否為異常
            if is_prefix_match_result:
                # 刀次=0 且品號前三碼與【最近非零刀次工單】相同 -> 保留 (正常情況)
                indices_to_keep.append(main_index)
                
            else:
                # 刀次=0 且品號前三碼與【最近非零刀次工單】不相似 (或前面無非零刀次工單) -> 異常
                anomalous_rows.append(main_row.copy())

        # ---------------------------------------------
        # 2. 刀次不為 0 -> 保留 (正常情況)
        # ---------------------------------------------
        else:
             indices_to_keep.append(main_index)

        i += 1
        
    # ===============================================
    # 執行資料移動與返回
    # ===============================================
    
    # 新的 df_paired (只包含正常組/單行)
    df_paired_new = df_paired.loc[indices_to_keep].reset_index(drop=True)
    
    # 異常工單 (不再與 extra_remaining 合併)
    df_anomalous = pd.DataFrame(anomalous_rows)

    # ===============================================
    # 處理異常工單
    # ===============================================
    if df_anomalous.empty:
        return df_paired_new, extra_remaining 
        
    # --- 步驟 A: 數量校正 (優先執行) ---
    def correct_anomalous_qty(row):
        order_no = row['工單編號']
        
        if 'GLOBAL_ORDER_QTY_MAP' in globals() and order_no in globals()['GLOBAL_ORDER_QTY_MAP']:
            qty_value = globals()['GLOBAL_ORDER_QTY_MAP'][order_no]
            
            qty_series = pd.Series([qty_value])
            new_qty_series = pd.to_numeric(qty_series, errors='coerce')
            new_qty = new_qty_series.fillna(0).iloc[0]
            
            if new_qty > 0:
                # 這裡的餘量和預估良品數都會被更新為校正後的開工數量
                row['餘量'] = new_qty
                row['預估良品數'] = new_qty
            
        return row
    
    df_anomalous = df_anomalous.apply(correct_anomalous_qty, axis=1)
    
    # --- 步驟 B: 重新配對 (加入刀次計算) ---
    
    remaining_df = df_anomalous.copy() 
    pair_rows_d1 = []
    
    # 1. 建立配對查找字典
    remaining_by_item = {}
    for _, row in remaining_df.iterrows():
        remaining_by_item.setdefault(row["品號"], []).append(row.to_dict())

    # 2. 準備 base_df 映射
    base_df_map = base_df.drop_duplicates(subset=['料號'], keep='first').set_index('料號')
    
    used_order_ids = set()
    
    # 3. 遍歷主工單並嘗試配對
    for index, main_order in df_anomalous.iterrows():
        main_order_id = main_order["工單編號"]
        main_item_code = main_order["品號"]
        
        if main_order_id in used_order_ids:
            continue
            
        if main_item_code not in base_df_map.index:
            continue
            
        base_info = base_df_map.loc[main_item_code]
        
        # 獲取主工單的基本屬性 (從 base_df)
        main_car_count_val = pd.to_numeric(base_info.get("車數"), errors='coerce')
        main_car_count_new = np.nan_to_num(main_car_count_val, nan=0.0) # 主工單車數

        if main_car_count_new <= 0:
             continue 

        # 獲取主工單的數量 (從 df_anomalous - 已經被校正過)
        main_qty = main_order.get('預估良品數', 0.0)
        
        if main_qty <= 0:
             continue

        # 💥 修正 1：計算新的主工單刀次
        # 刀次 = 預估良品數 / 車數
        main_cut_new = main_qty / main_car_count_new if main_car_count_new > 0 else 0.0
        
        if main_cut_new <= 0:
            continue
        
        # 獲取子工單的搭1料號和產出車數 (子工單車數)
        d1_code = base_info.get("搭1料號")
        d1_car_count_val = pd.to_numeric(base_info.get("搭1產出車數"), errors='coerce') 
        d1_car_count = np.nan_to_num(d1_car_count_val, nan=0.0) # 子工單車數
        
        if pd.isna(d1_code) or d1_code == "" or d1_car_count <= 0:
            continue
            
        # 查找子工單的寬度
        if d1_code not in base_df_map.index:
            continue
        
        d1_base_info = base_df_map.loc[d1_code]
        d1_width_val = pd.to_numeric(d1_base_info.get("寬度Cm"), errors='coerce')
        d1_width = np.nan_to_num(d1_width_val, nan=0.0)
        
        if d1_width <= 0:
            continue
            
        if d1_code not in remaining_by_item:
            continue
            
        # 尋找可用的子工單進行配對
        found_sub_order = None
        for i, d1_order_dict in enumerate(remaining_by_item[d1_code]):
            d1_order_id = d1_order_dict["工單編號"]
            if d1_order_id not in used_order_ids:
                found_sub_order = pd.Series(d1_order_dict)
                break
                
        if found_sub_order is not None:
            
            # 💥 修正 2：計算新的子工單預估良品數
            # 子工單預估良品數 = 子工單車數 * 主工單刀次
            d1_qty_new = d1_car_count * main_cut_new 
            
            # 找到配對，執行資料填充和標記
            used_order_ids.add(main_order_id)
            used_order_ids.add(found_sub_order["工單編號"])
            
            # --- 處理主工單 (main_copy) ---
            main_copy = main_order.copy()
            main_copy["備註配對工單"] = found_sub_order["工單編號"] 
            
            # 💥 更新：主工單刀次和車數
            main_copy["刀次"] = main_cut_new  # 保持浮點數精度
            main_copy["車數"] = int(main_car_count_new)
            
            # --- 處理子工單 (d1_copy) ---
            d1_copy = found_sub_order.copy()
            d1_copy["備註配對工單"] = main_order_id
            
            # 💥 更新：子工單刀次、車數和良品數
            d1_copy["刀次"] = 0 
            d1_copy["車數"] = int(d1_car_count) 
            d1_copy["預估良品數"] = d1_qty_new
            d1_copy["餘量"] = d1_qty_new # 假設配對後餘量=良品數
            
            d1_copy["寬度Cm"] = d1_width
            d1_copy["總長度(cm)"] = d1_width * d1_car_count 

            pair_rows_d1.extend([main_copy, d1_copy])

    # ... (步驟 E: 合併邏輯，保持不變) ...
    
    # 4. 分離未配對成功的行 (剩餘工單)
    df_repaired = pd.DataFrame(pair_rows_d1)
    
    # 找出所有未被使用的工單
    df_unmatched = remaining_df[~remaining_df["工單編號"].isin(used_order_ids)].copy()
    
    
    # ===============================================
    # 步驟 E: 執行最終合併
    # ===============================================

    # 1. 處理成功修復的工單：將修復後的工單 (df_repaired) 合併到 df_paired_new
    df_paired_new = pd.concat([df_paired_new, df_repaired], ignore_index=True)
    
    # 💥 修正點 1：從 extra_remaining 中移除已修復的工單
    if extra_remaining is not None and not extra_remaining.empty:
        extra_remaining = extra_remaining[~extra_remaining["工單編號"].isin(used_order_ids)].copy()
        

    # 💥 修正點 2：處理未配對成功的剩餘工單：合併到 extra_remaining
    if not df_unmatched.empty:
        if extra_remaining is not None and not extra_remaining.empty:
             df_unmatched_cols = df_unmatched.reindex(columns=extra_remaining.columns, fill_value=np.nan)
        else:
             df_unmatched_cols = df_unmatched
             
        extra_remaining = pd.concat([extra_remaining, df_unmatched_cols], ignore_index=True)


    return df_paired_new, extra_remaining



def post_optimization_clean_up(A159_df, A830_df, B201_df):
    
    print("正在執行最終清理和主子工單群組重新排序...")
    
    # 1. 合併三位人員的資料
    all_final_orders = pd.concat([A159_df, A830_df, B201_df], ignore_index=True)
    
    # 確保 '預計開工日' 是 datetime 格式，便於排序
    all_final_orders['預計開工日'] = pd.to_datetime(all_final_orders['預計開工日'], errors='coerce')
    
    # 儲存原始的行順序，以便在群組內部保持穩定
    all_final_orders['Original_Order'] = all_final_orders.index
    
    # 清空 '人員' 欄位
    if '人員' in all_final_orders.columns:
        all_final_orders['人員'] = ""
        
    # ----------------------------------------------------
    # 2. 確定群組 ID, 群組日期, 和 81B 優先級
    # ----------------------------------------------------
    
    # 找到所有工單所屬的「群組主工單編號」
    # 對於主工單 (刀次>0)，其群組ID是自己的工單編號
    # 對於子工單 (刀次=0)，其群組ID是其 '主工單編號'
    all_final_orders['Group_ID'] = np.where(
        all_final_orders['刀次'] > 0, 
        all_final_orders['工單編號'], 
        all_final_orders['主工單編號']
    )
    
    # 🌟 修正步驟 A：找出每個 Group_ID 對應的 (主工單) 預計開工日
    # 這裡我們只取主工單 (刀次 > 0) 的日期作為群組日期
    group_date_map = all_final_orders[all_final_orders['刀次'] > 0] \
        .set_index('工單編號')['預計開工日'] \
        .to_dict()
        
    # 將群組日期映射回所有工單 (包括子工單)
    all_final_orders['Group_Date'] = all_final_orders['Group_ID'].map(group_date_map)
    
    # 找出所有 Group_ID 中，以 '81B' 開頭的群組，並設定優先級標籤
    priority_group_ids = all_final_orders[all_final_orders['Group_ID'].str[:3] == '81B']['Group_ID'].unique()
    
    all_final_orders['Group_Priority_Rank'] = np.where(
        all_final_orders['Group_ID'].isin(priority_group_ids), 
        0,  # 81B 群組優先 (最高)
        1   # 其他群組次之
    )
    
    # ----------------------------------------------------
    # 3. 最終排序：確保群組連續、81B 優先，並剩下的按日期排序
    # ----------------------------------------------------
    
    # 排序邏輯：
    # 1. Group_Priority_Rank (將所有 81B 群組拉到最前)
    # 2. Group_Date (在 Group_Priority_Rank 內，依照群組日期排序，解決非 81B 群組的日期順序問題)
    # 3. Group_ID (確保同一日期內，同一個群組的工單會緊密排列)
    # 4. Original_Order (保持主子工單的內部順序)
    '''
    final_output_df = all_final_orders.sort_values(
        by=['Group_Priority_Rank', 'Group_Date', 'Group_ID', 'Original_Order'],
        ascending=[True, True, True, True]
    ).drop(columns=['Group_ID', 'Group_Priority_Rank', 'Group_Date', 'Original_Order']).reset_index(drop=True)
    '''
    final_output_df = all_final_orders.sort_values(
        by=['Group_Priority_Rank', 'Group_Date', 'Group_ID', 'Original_Order'],
        ascending=[True, True, True, True]
    ).reset_index(drop=True)
    
    print("✅ 主子工單群組重新排序完成。")
    
    return final_output_df


def assign_81b_priority_groups(all_final_orders):
    
    print("正在執行 81B 群組優先分配 (以群組公分組合為相似性鍵)...")
    
    # ----------------------------------------
    # 1. 數據準備：篩選 81B 群組並定義相似性鍵
    # ----------------------------------------
    
    if 'Group_ID' not in all_final_orders.columns:
         raise KeyError("錯誤：'Group_ID' 欄位遺失。請確保 post_optimization_clean_up 函數已經運行並保留此欄位。")
    
    # 找出所有 81B 群組 ID
    is_81B_group = all_final_orders['Group_ID'].str[:3] == '81B'
    priority_orders = all_final_orders[is_81B_group].copy()
    remaining_orders = all_final_orders[~is_81B_group].copy()

    # --- 關鍵修正：計算群組公分組合鍵 ---
    
    # 1. 彙總每個 Group_ID 內所有的 '公分'
    # .unique() 確保公分不重複，.sort_values() 確保順序一致，.str.cat() 轉為逗號分隔字串
    group_public_key = priority_orders.groupby('Group_ID')['公分'] \
                                     .apply(lambda x: ','.join(map(str, sorted(x.unique())))) \
                                     .to_dict()

    # 2. 將這個 '公分組合鍵' 映射回所有工單
    priority_orders['相似性鍵'] = priority_orders['Group_ID'].map(group_public_key)
    
    # 為了讓排序能找到，我們也需要對主工單進行一樣的映射
    main_81b_orders = priority_orders[priority_orders['刀次'] > 0].copy()
    main_81b_orders['相似性鍵'] = main_81b_orders['Group_ID'].map(group_public_key)

    # ----------------------------------------
    # 2. 排序與人員分配 (公分組合聚集邏輯)
    # ----------------------------------------
    
    # 🎯 排序：先按相似性鍵 (公分組合) 聚集，再按交期
    main_81b_orders = main_81b_orders.sort_values(
        by=['相似性鍵', '預計開工日', '工單編號'], 
        ascending=[True, True, True]
    ).reset_index(drop=True)

    STAFF_POOL = ["A159", "A830", "B201"]
    key_assignment_map = {} 
    next_staff_index = 0
    final_81b_schedule = []
    
    for _, main_row in main_81b_orders.iterrows():
        group_id = main_row['Group_ID']
        current_key = main_row['相似性鍵'] # 這裡的 Key 是 '6,11' 或 '8,9' 這樣的組合
        
        best_staff_id = None
        
        # 1. 優先：檢查這個 Key (公分組合) 是否已經被分配給某人
        if current_key in key_assignment_map:
            best_staff_id = key_assignment_map[current_key]
        
        # 2. 如果 Key 是新的，則循環分配
        if not best_staff_id:
            best_staff_id = STAFF_POOL[next_staff_index]
            key_assignment_map[current_key] = best_staff_id
            next_staff_index = (next_staff_index + 1) % len(STAFF_POOL)

        # ----------------------------------------
        # 3. 執行排程 (主子工單連續排程)
        # ----------------------------------------
        
        # 找出該主工單群組內的所有工單 (priority_orders 現在有 '相似性鍵' 了)
        group_orders_df = priority_orders[priority_orders['Group_ID'] == group_id].copy()
        
        # 確保順序：使用 Original_Order 來維持主子工單的內部順序
        group_orders_df = group_orders_df.sort_values(
            by=['預計開工日', 'Original_Order'],
            ascending=[True, True]
        ).to_dict('records')

        # 寫入排程結果
        for order in group_orders_df:
            order_copy = order.copy()
            order_copy['人員'] = best_staff_id
            final_81b_schedule.append(order_copy)
        
    print("✅ 81B 群組優先分配完成 (公分組合聚集)。")
    
    # ----------------------------------------
    # 4. 整理最終輸出
    # ----------------------------------------
    
    if not final_81b_schedule:
        return pd.DataFrame(), remaining_orders
        
    final_81b_df = pd.DataFrame(final_81b_schedule)
    
    # 最終排序：依 人員, 相似性鍵 (公分組合), 預計開工日
    final_81b_df = final_81b_df.sort_values(
        by=['人員', '相似性鍵', '預計開工日', 'Original_Order'], 
        kind='stable'
    ).reset_index(drop=True)

    # 移除輔助欄位
    cols_to_drop = ['Group_ID', 'Original_Order', '相似性鍵']
    final_81b_df = final_81b_df.drop(columns=[col for col in cols_to_drop if col in final_81b_df.columns])
    
    # 選擇最終輸出的欄位順序
    cols_to_keep = ['預計開工日', '人員', '工單編號', '品號', '餘量', '公分', '刀次', '預估良品數']
    
    # 返回分配好的 81B 群組和未處理的剩餘工單
    return final_81b_df[cols_to_keep], remaining_orders

# 根據歷史資料 處理前一個工作天尚未做完的工單
def schedule_history_download(part_end_list, history):
    """
    根據歷史資料，將前一天尚未完成的工單在此次排程中優先排前面。
    只移動主工單刀次 > 0，且包含其緊接的子工單（刀次=0）。
    傳入的 part_end_list是 DataFrame。
    """
    from pandas.tseries.offsets import BDay

    # 假設 user_schedule_date 為全域變數
    prev_workday = (pd.to_datetime(user_schedule_date) - BDay(1)).strftime("%Y-%m-%d")
    print(f"📅 前一工作日: {prev_workday}")

    def move_previous_day_orders_to_front(today_df):
        # 取前一天歷史資料主工單刀次 > 0
        df_history = history.copy()
        df_history["預計完工日_str"] = pd.to_datetime(df_history["預計完工日"]).dt.strftime("%Y-%m-%d")
        prev_main_orders = set(
            df_history[(df_history["預計完工日_str"] == prev_workday) &
                       (df_history["刀次"] > 0)]["工單號碼"].tolist()
        )

        print(prev_main_orders)

        new_order_rows = []
        i = 0
        while i < len(today_df):
            row = today_df.iloc[i]
            main_order_id = row["工單編號"]
            cut_count = row["刀次"]

            if cut_count > 0 and main_order_id in prev_main_orders:
                # 找到主工單，包含緊接的子工單一起移動
                group_rows = [row]
                i += 1
                while i < len(today_df) and today_df.iloc[i]["刀次"] == 0:
                    group_rows.append(today_df.iloc[i])
                    i += 1

                # 更新主工單與子工單日期，並複製一份避免警告
                updated_group = []
                for r in group_rows:
                    r_copy = r.copy()
                    r_copy["預計開工日"] = pd.to_datetime(user_schedule_date).strftime("%Y/%m/%d")
                    r_copy["預計完工日"] = pd.to_datetime(user_schedule_date).strftime("%Y/%m/%d")
                    updated_group.append(r_copy)

                # 放到最前面
                new_order_rows = updated_group + new_order_rows
            else:
                # 其他維持原順序
                new_order_rows.append(row.copy())
                i += 1

        # 轉回 DataFrame
        new_df = pd.DataFrame(new_order_rows).reset_index(drop=True)
        return new_df

    # 處理排程清單
    part_end_list = move_previous_day_orders_to_front(part_end_list)

    return part_end_list



def assign_personnel_by_similarity(df: pd.DataFrame, personnel_list: List[str] = None) -> Tuple[pd.DataFrame, Dict]:
    """
    【最終穩健修正版 - 引入每週區塊分配】
    1. 強化空值處理以確保回寫成功。
    2. 引入「週數」作為聚類鍵的一部分，將相似性指派的範圍限制在同一週內，實現每週尺寸工單的負荷分散與人員輪替。
    """
    
    # ----------------------------------------------------
    # 初始化與日誌設定
    # ----------------------------------------------------
    if personnel_list is None:
        personnel_list = ["A159", "A830", "B201"]
        
    df_result = df.copy() 
    debug_log: Dict[str, Dict[str, Any]] = {} 
    
    print(f"--- 偵錯日誌開始 ---")
    print(f"可指派人員: {personnel_list}")
    
    # ----------------------------------------------------
    # ⭐ 步驟 1: 欄位清理與準備 (強化人員欄位空值處理) ⭐
    # ----------------------------------------------------
    
    if '人員' not in df_result.columns:
        df_result['人員'] = pd.NA
    
    # 1. 強制轉換為字串，然後用 .strip() 清理兩側空格
    df_result['人員'] = df_result['人員'].astype(str).str.strip()
    
    # 2. 將所有可能的空值表示 (包括 'nan', 'None', 空字串) 統一替換為 pd.NA
    empty_markers = ['', 'nan', 'none', '<NA>']
    df_result['人員'] = df_result['人員'].apply(
        lambda x: pd.NA if str(x).lower() in empty_markers else x
    )

    # 必須存在的 ID 強制清理
    df_result['唯一主工單ID'] = df_result['唯一主工單ID'].astype(str).str.strip() 
    
    # 處理用於相似性判斷的欄位
    df_result['尺寸公分'] = df_result['公分'].fillna('N/A').astype(str)
    df_result['品號群組'] = df_result['品號'].apply(lambda x: str(x)[:4] if pd.notna(x) and str(x).strip() != '' else 'NONE')
    df_result["預計開工日"] = pd.to_datetime(df_result["預計開工日"], errors='coerce')
    
    total_unique_ids_count = df_result['唯一主工單ID'].nunique()
    print(f"\n[階段 1.1] 總工單群組數 (原始): {total_unique_ids_count}")
    
    # ⭐ 核心修復 1: 避免因 NaN 丟棄整個工單行 ⭐
    df_working = df_result[
        df_result['唯一主工單ID'].ne('') & 
        df_result['刀次'].notna()
    ].copy()
    
    df_working_ids_count = df_working['唯一主工單ID'].nunique()
    print(f"[階段 1.2] 排除 ID/刀次 為空後，剩餘工單群組數: {df_working_ids_count}")
    
    # 確保刀次是數值，並將 NaT 的預計開工日排除在分組外
    df_working['刀次'] = df_working['刀次'].fillna(0)
    df_working_valid_date = df_working.dropna(subset=['預計開工日'])

    if df_working_valid_date.empty:
        print("[階段 1.3] 警告: 沒有工單進入指派邏輯。")
        df_result.drop(columns=['尺寸公分', '品號群組'], inplace=True, errors='ignore')
        return df_result.reset_index(drop=True), debug_log 
    
    # ----------------------------------------------------
    # ⭐ 步驟 1.5: 引入週數作為區塊依據 (關鍵修正) ⭐
    # ----------------------------------------------------
    # 使用 isocalendar().week 取得標準週數 (1-52/53)，並轉為字串格式
    df_working_valid_date['排程週數'] = df_working_valid_date['預計開工日'].dt.isocalendar().week.astype(int).apply(lambda x: f"W{x:02d}")
    
    valid_date_ids_count = df_working_valid_date['唯一主工單ID'].nunique()
    print(f"[階段 1.3] 排除 NaT 日期後，用於分組計算的工單群組數: {valid_date_ids_count}")
    
    excluded_ids = set(df_result['唯一主工單ID'].unique()) - set(df_working_valid_date['唯一主工單ID'].unique())
    if excluded_ids:
        print(f"!!! 警告: 有 {len(excluded_ids)} 個群組被排除在指派邏輯之外。")
        
    # ----------------------------------------------------
    # 步驟 A, B: 指紋創建與相似性指派 
    # ----------------------------------------------------
    def get_group_fingerprint(group: pd.DataFrame):
        unique_sizes = sorted(group['尺寸公分'].unique())
        size_fingerprint = '|'.join(unique_sizes)
        assigned_in_group = group['人員'].notna()
        is_assigned = assigned_in_group.any()
        assigned_person_name = group.loc[assigned_in_group, '人員'].iloc[0] if is_assigned else None
        main_order_row = group[group['刀次'] > 0].iloc[0] if (group['刀次'] > 0).any() else None
        main_item_group = main_order_row['品號群組'] if main_order_row is not None else 'NONE'
        
        # ⭐ 提取週數 (群組內應為同一週)
        main_week = group['排程週數'].iloc[0]
        
        return pd.Series({
            '尺寸指紋': size_fingerprint, '品號群組': main_item_group,
            '群組刀次': group['刀次'].sum(), '是否已指派': is_assigned,
            '已指派人員': assigned_person_name,
            '排程週數': main_week # ⭐ 回傳週數
        })

    group_info = df_working_valid_date.groupby(['唯一主工單ID']).apply(get_group_fingerprint).reset_index()
    
    # ⭐ 修正聚類鍵: 尺寸相似性只在同一週內有效
    group_info['群組聚類鍵'] = (
        group_info['排程週數'] + '-' + 
        group_info['尺寸指紋'] + '-' + 
        group_info['品號群組'] 
    )
    
    assigned_groups_map: Dict[str, str] = {} 
    
    group_info_assigned = group_info[group_info['是否已指派'] == True]
    for _, row in group_info_assigned.iterrows():
        assigned_groups_map[row['唯一主工單ID']] = row['已指派人員']
        debug_log[row['唯一主工單ID']] = {'指派類型': '原始指派', '人員': row['已指派人員'], '刀次': row['群組刀次']}

    group_info_to_assign = group_info[group_info['是否已指派'] == False].copy()
    print(f"[階段 2.1] 待指派工單群組數: {len(group_info_to_assign)}")
    
    # 排序：優先處理聚類鍵相同 (同一週同尺寸)，且刀次高的群組
    group_info_to_assign.sort_values(by=['群組聚類鍵', '群組刀次'], ascending=[True, False], inplace=True)
    available_personnel = personnel_list 
    
    # 追蹤人員狀態及初始負荷計算 
    person_status: Dict[str, Dict[str, Any]] = {
        person: {"capacity_used": 0, "current_fingerprint": None, "current_item": None, "current_week": None}
        for person in available_personnel
    }
    initial_assigned_load = df_working[df_working['唯一主工單ID'].isin(group_info_assigned['唯一主工單ID'])].groupby('人員')['刀次'].sum().fillna(0)
    for person, load in initial_assigned_load.items():
        if person in person_status:
             person_status[person]["capacity_used"] = load

    # 執行相似性指派 (A, B, B', C 邏輯)
    for index, group_row in group_info_to_assign.iterrows():
        group_id = group_row['唯一主工單ID']
        dose = group_row['群組刀次']
        current_fingerprint = group_row['尺寸指紋']
        current_item = group_row['品號群組']
        current_week = group_row['排程週數'] # ⭐ 追蹤當前週數
        best_person = None
        assignment_logic = '未指派' 
        
        # A. 優先匹配：尺寸指紋 AND 品號群組 AND 【同一週】
        for p in available_personnel:
            status = person_status[p]
            if (status["current_fingerprint"] == current_fingerprint and 
                status["current_item"] == current_item and
                status["current_week"] == current_week): # ⭐ 必須在同一週
                best_person = p
                assignment_logic = 'A. 雙重匹配 (同週)'
                break
                
        # B. 次級匹配：僅匹配 尺寸指紋 AND 【同一週】
        if not best_person:
            for p in available_personnel:
                status = person_status[p]
                if (status["current_fingerprint"] == current_fingerprint and
                    status["current_week"] == current_week): # ⭐ 必須在同一週
                    best_person = p
                    assignment_logic = 'B. 尺寸匹配 (同週)'
                    break
        
        # B'. 次次級匹配：僅匹配 品號群組 (不強制同週，因為品號專精較不依賴時間)
        if not best_person:
            for p in available_personnel:
                status = person_status[p]
                if (status["current_item"] == current_item):
                    best_person = p
                    assignment_logic = 'B\'. 品號匹配 (跨週可)'
                    break
        
        # C. 最終強制分配給負荷最低的人員 (跨週或新尺寸時，由負荷決定誰接手)
        if not best_person:
            if available_personnel:
                sorted_personnel = sorted(available_personnel, key=lambda p: person_status[p]["capacity_used"])
                best_person = sorted_personnel[0]
                assignment_logic = f'C. 負荷最低 ({person_status[best_person]["capacity_used"]} 刀)'
            else:
                best_person = "ERROR_NO_PERSONNEL"; assignment_logic = "ERROR: 無可用人員"
        
        # 應用指派
        assigned_to = best_person
        
        # 更新狀態 (重點是更新 current_week, 這樣下一週的工單就不會匹配到這個狀態)
        if assigned_to != "ERROR_NO_PERSONNEL":
            status = person_status[assigned_to]
            status["capacity_used"] += dose
            status["current_fingerprint"] = current_fingerprint
            status["current_item"] = current_item
            status["current_week"] = current_week # ⭐ 更新當前週數
        
        assigned_groups_map[group_id] = assigned_to
        debug_log[group_id] = {'指派類型': assignment_logic, '人員': assigned_to, '刀次': dose}


    print(f"\n[階段 3.1] 待指派工單群組指派完成。")
    print(f"所有指派的映射表大小 (assigned_groups_map): {len(assigned_groups_map)}")
    
    # ----------------------------------------------------
    # 步驟 C: 應用映射回原始 DataFrame (回寫邏輯與前一版相同)
    # ----------------------------------------------------
    
    df_result.reset_index(drop=True, inplace=True)
    assignment_series = df_result['唯一主工單ID'].map(assigned_groups_map)
    original_person_is_empty = df_result['人員'].isna()
    has_assignment_result = assignment_series.notna()
    mask_to_update = original_person_is_empty & has_assignment_result
    df_result.loc[mask_to_update, '人員'] = assignment_series.loc[mask_to_update]

    '''
    # ----------------------------------------------------
    # ⭐ NEW STEP: 步驟 D: 執行工單群組的內部排序 (P3 優先於 P2/P4) ⭐
    # ----------------------------------------------------

    # 1. 將 group_info 的聚類鍵信息和品號群組資訊，映射回 df_result 的每一行
    group_cluster_map = group_info.set_index('唯一主工單ID')

    df_result['指派人員'] = df_result['人員'].fillna('未指派') 
    df_result['群組聚類鍵'] = df_result['唯一主工單ID'].map(group_cluster_map['群組聚類鍵'])
    df_result['群組刀次'] = df_result['唯一主工單ID'].map(group_cluster_map['群組刀次'])

    # 映射 P2/P4 的品號群組，作為次要排序依據
    df_result['品號群組_排序'] = df_result['唯一主工單ID'].map(group_cluster_map['品號群組'])

    # 3. 排序執行 (確保 '品號群組_排序' 在 '群組聚類鍵' 之後)
    df_result['__original_index'] = df_result.index # 保存原始索引，用於穩定排序

    df_result.sort_values(by=[
        '指派人員', 
        '群組聚類鍵',       # P3 優先：所有 4cm (W46-4) 聚在一起
        '品號群組_排序',    # P2/P4 次之：在 4cm 群組內，再按品號群組 (UTMX/TTRS) 排序
        '群組刀次', 
        '__original_index'
    ], ascending=[True, True, True, False, True], inplace=True)

    # 清理並重設索引
    df_result.drop(columns=['指派人員', '群組聚類鍵', '群組刀次', '__original_index', '品號群組_排序'], inplace=True, errors='ignore')
    df_result.reset_index(drop=True, inplace=True)
    '''

    # ----------------------------------------------------
    # 步驟 C (原步驟 3.2): 檢查最終仍為 None/空字串的工單
    # ----------------------------------------------------
    final_none_mask = df_result['人員'].isna()
    final_none_ids = df_result.loc[final_none_mask, '唯一主工單ID'].unique()

    print(f"\n[階段 3.2] 回寫後，最終仍為 None/空字串的工單群組數: {len(final_none_ids)}")

    if final_none_ids.size > 0:
        print(f"!!! 最終發現: 這些工單ID ({list(final_none_ids)[:3]}...) 未被指派。")
    else:
        print(f"[階段 4] 最終檢查：所有工單皆已指派完成。")

    # 清理暫時欄位
    df_result.drop(columns=['尺寸公分', '品號群組', '排程週數'], inplace=True, errors='ignore')

    print(f"--- 偵錯日誌結束 ---")

    return df_result, debug_log


def cleanup_remaining_df(remaining_df):
    """
    將 remaining_df 恢復為每個工單編號 (SN) 只有一筆記錄的結構。
    保留每個 SN 第一次出現的記錄，並確保良品數與刀次/車數對齊。
    """
    if remaining_df.empty:
        return remaining_df
    
    # 🚨 步驟 1: 刪除重複行，只保留第一次出現的 SN
    # 由於在之前的合併中，同一 SN 只是被附加了不同的規則，我們選擇第一條記錄。
    remaining_df_cleaned = remaining_df.drop_duplicates(subset=['工單編號'], keep='first').copy()
    
    # 🚨 步驟 2: 重新計算良品數和餘量 (因為前面的合併和計算可能導致數字有微小差異)
    remaining_df_cleaned["車數"] = pd.to_numeric(remaining_df_cleaned["車數"], errors='coerce').fillna(1).astype(int)
    remaining_df_cleaned["刀次"] = pd.to_numeric(remaining_df_cleaned["刀次"], errors='coerce').apply(math.ceil).astype(int)
    
    # 重新對齊良品數為 車數 x 刀次
    remaining_df_cleaned["預估良品數"] = remaining_df_cleaned["車數"] * remaining_df_cleaned["刀次"]
    remaining_df_cleaned["餘量"] = remaining_df_cleaned["預估良品數"]
    
    return remaining_df_cleaned


def cleanup_remaining_df_second(df_remaining):
    """
    對 remaining DataFrame 進行修正和標準化計算：
    1. 根據現有良品數和車數，重新計算刀次 (無條件進位)。
    2. 根據新的刀次和車數，重新計算預估良品數。

    Args:
        df_remaining (pd.DataFrame): 待處理的工單 DataFrame。

    Returns:
        pd.DataFrame: 修正後的 DataFrame。
    """
    
    if df_remaining is None or df_remaining.empty:
        print("ℹ️ 輸入 DataFrame 為空，跳過修正。")
        return pd.DataFrame()

    # 複製 DataFrame 以避免 In-place 修改原始數據
    df_cleaned = df_remaining.copy()

    # 1. 確保關鍵欄位是數值型，並將缺失或無效值視為 0 或 1 處理
    df_cleaned["預估良品數"] = pd.to_numeric(df_cleaned["預估良品數"], errors="coerce").fillna(0)
    # 確保車數至少為 1，避免除以零的錯誤
    df_cleaned["車數"] = pd.to_numeric(df_cleaned.get("車數"), errors="coerce").fillna(1).clip(lower=1)
    
    print(f"🔧 開始修正 remaining 總筆數: {len(df_cleaned)} 筆...")

    # 2. 應用計算邏輯
    # -------------------------------------------------------------------------
    # 刀次 = math.ceil(預估良品數 / 車數)
    # 使用 apply 確保逐行操作，並使用 math.ceil 進行無條件進位
    df_cleaned["刀次"] = df_cleaned.apply(
        lambda row: math.ceil(row["預估良品數"] / row["車數"])
                    if row["車數"] > 0 and row["預估良品數"] > 0 else 0,
        axis=1
    )
    
    # 3. 重新計算 預估良品數 = 刀次 * 車數 (Check)
    df_cleaned["預估良品數"] = df_cleaned["刀次"] * df_cleaned["車數"]
    # -------------------------------------------------------------------------
    
    print("✅ remaining 修正計算完成。")
    return df_cleaned

def cleanup_all(remaining, all):
    """
    執行最終清理：
    1. 檢查已配對工單的總數量是否超過 GLOBAL_ORDER_QTY_MAP，若超過則刪除剩餘批次。
    2. 從 remaining (原始未分配池) 中刪除本次已配對的主工單。

    Args:
        remaining (pd.DataFrame): 原始未分配工單池 (final_remaining)。
        all_df (pd.DataFrame): 經過配對和還原後的最終資源池 ('all' 的值)。

    Returns:
        tup
    """

    # 確保操作的是副本
    remaining_cleaned = remaining.copy()
    all_df_cleaned = all.copy()

    # ------------------------------------------------------------------------
    # 1. 檢查已配對數量是否超額，並刪除 NO_PAIR_..._LEFT 中的重複工單
    # ------------------------------------------------------------------------
    
    paired_types = ['ORIGINAL_PAIRED', 'NEWLY_PAIRED_MAIN', 'NEWLY_PAIRED_SUB']
    leftover_types = ['NO_PAIR_SPLIT_LEFT', 'NO_PAIR_SECOND_LEFT']
    
    # 1.1 篩選出所有已配對或已分配的工單行
    assigned_df = all_df_cleaned[all_df_cleaned['來源類型'].isin(paired_types)]
    
    # 1.2 計算每筆工單在 all_df 中已分配的數量總和
    paired_qty_sum = assigned_df.groupby('工單編號')['預估良品數'].sum()
    
    print("🔄 Cleanup Step 1: 檢查並調整工單總數量...")
    
    # 1.3 準備剩餘批次 DataFrame 和需要修改的索引
    leftover_df = all_df_cleaned[all_df_cleaned['來源類型'].isin(leftover_types)].copy()
    
    # 暫存要刪除的工單SN (用於超額刪除)
    sns_to_remove_leftover = []
    
    # 暫存要更新的工單SN (用於數量調整)
    indices_to_update = []
    
    global GLOBAL_ORDER_QTY_MAP 

    # 1.4 迭代檢查所有工單 SN
    for sn in all_df_cleaned['工單編號'].unique():
        original_qty = GLOBAL_ORDER_QTY_MAP.get(sn)
        if original_qty is None:
            continue
            
        assigned_qty = paired_qty_sum.get(sn, 0) # 已分配的數量 (paired_types)
        remaining_space = original_qty - assigned_qty
        
        # 該工單在剩餘批次中的行
        sn_leftover_rows = leftover_df[leftover_df['工單編號'] == sn]
        
        if sn_leftover_rows.empty:
            continue
            
        # 情況 A: 總需求已滿或超額 (remaining_space <= 0)
        if remaining_space <= 0:
            # 必須完全刪除這些剩餘批次
            sns_to_remove_leftover.append(sn)
            
        # 情況 B: 尚有空間 (remaining_space > 0)，但需要調整剩餘批次的數量
        elif remaining_space > 0:
            
            # 該工單在剩餘批次中的總數量
            leftover_total_qty = sn_leftover_rows['預估良品數'].sum()
            
            # 只有當剩餘批次總量超過剩餘空間時，才需要調整
            if leftover_total_qty > remaining_space:
                
                # 將剩餘空間分配給第一個找到的剩餘批次行，並將其餘的刪除。
                # 這裡假設您只允許一個剩餘批次承載剩餘空間。
                
                # 獲取第一個剩餘批次的原始索引 (用於在 all_df_cleaned 上定位)
                first_leftover_idx = sn_leftover_rows.index[0]
                
                # 1. 將第一個剩餘批次標記為更新
                indices_to_update.append(first_leftover_idx)
                
                # 2. 將第一個剩餘批次之後的行標記為刪除 (超額刪除)
                if len(sn_leftover_rows) > 1:
                    sns_to_remove_leftover.extend(sn_leftover_rows['工單編號'].iloc[1:].unique().tolist())
                
                # 3. 更新第一個剩餘批次的預估良品數
                # 由於我們正在操作 all_df_cleaned 的副本，我們直接在副本上更新
                all_df_cleaned.loc[first_leftover_idx, '預估良品數'] = remaining_space
                
                # 4. 重新計算刀次和餘量
                sub_car_count = all_df_cleaned.loc[first_leftover_idx, '車數'] # 使用該行的車數
                
                # 如果車數 <= 0，設為 1 以避免除以零
                if pd.isna(sub_car_count) or sub_car_count <= 0:
                    sub_car_count = 1 
                
                new_cut = math.ceil(remaining_space / sub_car_count)
                new_qty = new_cut * sub_car_count
                
                all_df_cleaned.loc[first_leftover_idx, '刀次'] = new_cut
                all_df_cleaned.loc[first_leftover_idx, '餘量'] = new_qty
                all_df_cleaned.loc[first_leftover_idx, '預估良品數'] = new_qty
                
                print(f"   -> 調整工單 {sn}：剩餘批次數量從 {leftover_total_qty} 調整為 {new_qty}。")


    # 1.5 處理超額刪除（包含情況A和情況B中被標記要刪除的剩餘批次）
    if sns_to_remove_leftover:
        # 建立一個邏輯遮罩：將屬於超額工單的剩餘批次標記為 False (要被移除)
        mask = ~((all_df_cleaned['工單編號'].isin(sns_to_remove_leftover)) & 
                 (all_df_cleaned['來源類型'].isin(leftover_types)) &
                 (~all_df_cleaned.index.isin(indices_to_update))) # 不刪除剛更新過的行
        
        initial_all_count = len(all_df_cleaned)
        all_df_cleaned = all_df_cleaned[mask].reset_index(drop=True)
        removed_count = initial_all_count - len(all_df_cleaned)
        
        print(f"✅ Cleanup Step 1a/b 完成: 移除 {removed_count} 筆已超額分配工單的剩餘批次。")
    else:
        print("✅ Cleanup Step 1a/b 完成: 未發現超額工單，無需移除剩餘批次。")
        
    # ------------------------------------------------------------------------
    # 2. 收集 NEWLY_PAIRED_MAIN 的工單，並從 remaining 中刪除
    # ------------------------------------------------------------------------
    
    # 2.1 收集所有在本次配對中作為主工單被消耗的工單編號 (SN)
    paired_main_sns = all_df_cleaned[
        all_df_cleaned['來源類型'] == 'NEWLY_PAIRED_MAIN'
    ]['工單編號'].unique()
    
    if paired_main_sns.size > 0:
        # 2.2 刪除 remaining_cleaned 中對應的原始行
        # 由於 remaining 在這次配對中未被修改，其行代表整個工單的原始數據。
        # 只要工單有 NEWLY_PAIRED_MAIN 紀錄，表示它已經被處理，原始行應被移除。
        mask_remaining = ~remaining_cleaned['工單編號'].isin(paired_main_sns)
        
        initial_remaining_count = len(remaining_cleaned)
        remaining_cleaned = remaining_cleaned[mask_remaining].reset_index(drop=True)
        removed_count = initial_remaining_count - len(remaining_cleaned)
        
        print(f"✅ Cleanup Step 2 完成: 從 remaining 中移除 {removed_count} 筆已配對的主工單原始資料。")
    else:
        print("💡 Cleanup Step 2: 本次無新的主工單配對，remaining 保持不變。")

    
    return remaining_cleaned, all_df_cleaned


def do_people(final_df):
    
    # 拆分三位人員
    A159_part = final_df[final_df["人員"] == "A159"].copy()
    A830_part = final_df[final_df["人員"] == "A830"].copy()
    B201_part = final_df[final_df["人員"] == "B201"].copy()

    return A159_part, A830_part, B201_part


def generate_schedule_for_person(df: pd.DataFrame, break_dates: set, max_lookback_days: int = 65) -> pd.DataFrame:

    df = df.copy()
    df["預計開工日"] = pd.to_datetime(df["預計開工日"])
    df["預計完工日"] = pd.to_datetime(df["預計完工日"])
    df["實際排程日期"] = None

    # 每日產能限制（週一到週五）
    daily_limits = {0: 60, 1: 60, 2: 55, 3: 60, 4: 55}
    #daily_limits = {0: 55, 1: 55, 2: 55, 3: 55, 4: 55}

    # 建立主工單索引（主工單：刀次 > 0）
    last_main_idx = None
    df["主工單索引"] = None
    for idx, row in df.iterrows():
        if row["刀次"] > 0:
            last_main_idx = idx
        df.at[idx, "主工單索引"] = last_main_idx

    # 每組用「最早完工日」決定排程範圍
    df["組內最早完工日"] = df.groupby("主工單索引")["預計完工日"].transform("min")

    groups = df.groupby("主工單索引")
    # ❗ 最終修正：定義一個安全提取函數，解決 Series/Timestamp 類型切換問題 ❗
    def safe_get_sort_key(g):
        # g[0] 是該群組的第一個索引
        result = df.loc[g[0], "組內最早完工日"]
        
        # 判斷結果是否為 Series
        if isinstance(result, pd.Series):
            # 如果是 Series (代表索引重複)，我們取第一個元素 (因為 '組內最早完工日' 應該相同)
            # 使用 .iloc[0] 從 Series 中提取單一 Timestamp
            return result.iloc[0] 
        else:
            # 如果是單一的 Timestamp 標量，直接返回
            return result 

    sorted_groups = sorted(
        groups, 
        key=safe_get_sort_key, # 使用新的安全函數作為排序鍵
        reverse=True
    )

    # 建立排程容量表
    latest_end = df["預計完工日"].max()

    earliest_start = latest_end - timedelta(days=max_lookback_days)

    schedule_capacity = {}
    date = latest_end
    while date >= earliest_start:
        if date.weekday() < 5 and date not in break_dates:
            schedule_capacity[date] = daily_limits.get(date.weekday(), 55)
        date -= timedelta(days=1)

    for main_idx, group in sorted_groups:
        group_df = df.loc[group.index]
        earliest_end = group_df["組內最早完工日"].min()
        start_date = earliest_end - timedelta(days=max_lookback_days)

        total_doz = group_df["刀次"].sum()
        remaining = total_doz
        assigned_list = []
        current_date = earliest_end

        while remaining > 0 and current_date >= start_date:
            if current_date in schedule_capacity and schedule_capacity[current_date] > 0:
                available = schedule_capacity[current_date]
                assign = min(available, remaining)
                schedule_capacity[current_date] -= assign
                assigned_list.append((current_date.strftime("%Y-%m-%d"), assign))
                remaining -= assign
            current_date -= timedelta(days=1)

        if remaining > 0:
            raise Exception(f"排程無法完成，工單 {df.loc[main_idx, '工單編號']} 剩餘刀次 {remaining}")

        # 寫入每筆工單的排程日期
        for idx in group.index:
            if idx == main_idx:
                formatted = [f"{d}({c})" for d, c in sorted(assigned_list)]
            else:
                formatted = [d for d, _ in sorted(assigned_list)]
            df.at[idx, "實際排程日期"] = "\n".join(formatted)

    df.drop(columns=["組內最早完工日", "主工單索引"], inplace=True)

    df_new = split_schedule_dates(df)

    return df_new


def final_cal_list_person(df: pd.DataFrame, start_d: datetime.date, break_dates: set) -> pd.DataFrame:
    df = df.copy()

    if isinstance(start_d, pd.Timestamp):
        start_d = start_d.date()
    elif isinstance(start_d, str):
        start_d = pd.to_datetime(start_d).date()

    daily_limits = {0: 60, 1: 60, 2: 55, 3: 60, 4: 55}
    special_limit = 60
    capacity_used = {}

    doses = pd.to_numeric(df["刀次"], errors="coerce").fillna(0).tolist()
    sku_list = df.get("品號", pd.Series([None]*len(doses))).tolist()
    new_dates = []
    current_date = start_d

    for dose, sku in zip(doses, sku_list):
        if dose == 0:
            # 子工單：繼承上一筆日期
            new_dates.append(new_dates[-1] if new_dates else current_date)
            continue

        while True:
            wd = current_date.weekday()
            if wd < 5 and current_date not in break_dates:
                used_info = capacity_used.get(current_date, {"count": 0, "sku_set": set(), "nonzero_sku_list": []})

                # 判斷當天非子工單品號（刀次>0）有哪些
                # 先從已分配當天的非子工單sku列表+目前這筆sku判斷
                current_nonzero_skus = [s for s in used_info["nonzero_sku_list"] if s is not None]
                # 加上當前正在安排的sku（dose != 0）
                if sku not in current_nonzero_skus:
                    current_nonzero_skus.append(sku)

                # 判斷當天是否只有一個品號(非子工單)
                is_single_sku = (len(set(current_nonzero_skus)) == 1)

                # 判斷是否為非加班日(週三2,週五4)，且是非假日
                is_special_day = (wd in [2, 4]) and (current_date not in break_dates)

                if is_single_sku and is_special_day:
                    limit = special_limit
                else:
                    limit = daily_limits.get(wd, 55)

                if used_info["count"] + dose <= limit:
                    used_info["count"] += dose
                    used_info["sku_set"].add(sku)
                    # 更新非子工單品號清單
                    if dose != 0 and sku is not None:
                        if sku not in used_info["nonzero_sku_list"]:
                            used_info["nonzero_sku_list"].append(sku)
                    capacity_used[current_date] = used_info
                    new_dates.append(current_date)
                    break

            current_date += timedelta(days=1)

    df.loc[:, "預計開工日"] = [d.strftime("%Y/%m/%d") for d in new_dates]
    df.loc[:, "預計完工日"] = df["預計開工日"]

    return df



def final_schedule_list_second(A159_part, A830_part, B201_part, break_dates):
    
    global start_date, end_date, user_schedule_date  # 使用全域變數
  
    schedule_A159 = generate_schedule_for_person(A159_part, break_dates["A159"])
    schedule_A830 = generate_schedule_for_person(A830_part, break_dates["A830"])
    schedule_B201 = generate_schedule_for_person(B201_part, break_dates["B201"])

    # 容合
    if not schedule_A830.empty and schedule_A830["刀次"].notna().any() and (schedule_A830["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_A830["刀次"] > 0
        schedule_A830.loc[mask, "預估良品數"] = schedule_A830.loc[mask, "刀次"] * schedule_A830.loc[mask, "車數"]

        # 統一日期格式
        schedule_A830["實際排程日期"] = pd.to_datetime(schedule_A830["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_A830["主工單識別碼"] = schedule_A830["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_A830["排程依據日"] = schedule_A830.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list_person 用）
        schedule_A830["預計開工日"] = schedule_A830["排程依據日"]
        schedule_A830["預計完工日"] = schedule_A830["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_A830["原始順序"] = schedule_A830.index

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_A830 = final_cal_list_person(schedule_A830, start_d=user_schedule_date, break_dates=break_dates["A830"])
            schedule_A830 = sort_by_customer_due_date(schedule_A830)
            schedule_A830 = final_cal_list_person(schedule_A830, start_d=user_schedule_date, break_dates=break_dates["A830"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_A830["預計開工日"] = pd.to_datetime(schedule_A830["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_A830["預計完工日"] = pd.to_datetime(schedule_A830["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_A830.drop(columns=["實際排程日期", "排程依據日", "主工單開工日" , "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("A830刀次欄位沒有資料或全部為空，跳過後續處理")

    
    # 家偉
    if not schedule_A159.empty and schedule_A159["刀次"].notna().any() and (schedule_A159["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_A159["刀次"] > 0
        schedule_A159.loc[mask, "預估良品數"] = schedule_A159.loc[mask, "刀次"] * schedule_A159.loc[mask, "車數"]

        # 統一日期格式
        schedule_A159["實際排程日期"] = pd.to_datetime(schedule_A159["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_A159["主工單識別碼"] = schedule_A159["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_A159["排程依據日"] = schedule_A159.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list_person 用）
        schedule_A159["預計開工日"] = schedule_A159["排程依據日"]
        schedule_A159["預計完工日"] = schedule_A159["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_A159["原始順序"] = schedule_A159.index


        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_A159 = final_cal_list_person(schedule_A159, start_d=user_schedule_date, break_dates=break_dates["A159"])
            schedule_A159 = sort_by_customer_due_date(schedule_A159)
            schedule_A159 = final_cal_list_person(schedule_A159, start_d=user_schedule_date, break_dates=break_dates["A159"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_A159["預計開工日"] = pd.to_datetime(schedule_A159["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_A159["預計完工日"] = pd.to_datetime(schedule_A159["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_A159.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("A159刀次欄位沒有資料或全部為空，跳過後續處理")
    

    # 旺斌
    if not schedule_B201.empty and schedule_B201["刀次"].notna().any() and (schedule_B201["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_B201["刀次"] > 0
        schedule_B201.loc[mask, "預估良品數"] = schedule_B201.loc[mask, "刀次"] * schedule_B201.loc[mask, "車數"]

        # 統一日期格式
        schedule_B201["實際排程日期"] = pd.to_datetime(schedule_B201["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_B201["主工單識別碼"] = schedule_B201["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_B201["排程依據日"] = schedule_B201.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list_person 用）
        schedule_B201["預計開工日"] = schedule_B201["排程依據日"]
        schedule_B201["預計完工日"] = schedule_B201["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_B201["原始順序"] = schedule_B201.index

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_B201_1 = final_cal_list_person(schedule_B201, start_d=user_schedule_date, break_dates=break_dates["B201"])
            schedule_B201_2 = sort_by_customer_due_date(schedule_B201_1)
            schedule_B201_3 = final_cal_list_person(schedule_B201_2, start_d=user_schedule_date, break_dates=break_dates["B201"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_B201_3["預計開工日"] = pd.to_datetime(schedule_B201_3["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_B201_3["預計完工日"] = pd.to_datetime(schedule_B201_3["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_B201_3.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("B201刀次欄位沒有資料或全部為空，跳過後續處理")

    

    return schedule_A830, schedule_A159, schedule_B201_3


def balance_completion_time(df_A159: pd.DataFrame, df_A830: pd.DataFrame, df_B201: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    【最終修正：僅執行工單轉移，並手動模擬更新開工日，解決無限轉移問題】
    
    所有平衡判斷和轉移依據皆基於 '預計開工日'。不執行欄位重命名。
    """
    
    df_list: Dict[str, pd.DataFrame] = {
        'A159': df_A159.copy(),
        'A830': df_A830.copy(),
        'B201': df_B201.copy()
    }
    
    MAX_ITERATIONS = 10 
    DATE_COLUMN = '預計開工日'
    
    print(f"--- 警告: 內部平衡僅執行工單轉移，平衡依據為 '{DATE_COLUMN}'。 ---")
    
    # 初始開工日檢查 (這是唯一一次從 DataFrame 中讀取數據)
    latest_start_date = {}
    for pid, df in df_list.items():
        if not df.empty and DATE_COLUMN in df.columns:
            try:
                # 確保欄位轉換為日期時間類型
                latest_start_date[pid] = pd.to_datetime(df[DATE_COLUMN]).max().date()
            except Exception as e:
                print(f"警告: 人員 {pid} 的 '{DATE_COLUMN}' 轉換失敗 ({e})，使用默認值。")
                latest_start_date[pid] = pd.to_datetime('2200-01-01').date()
        else:
            latest_start_date[pid] = pd.to_datetime('2200-01-01').date() 
    
    # 複製初始日期供後續手動更新
    current_latest_start_date = latest_start_date.copy()

    
    for i in range(MAX_ITERATIONS):
        
        # 1. 檢查是否仍有人超期 (基於原始數據的檢查)
        is_anyone_overdue = False
        # 這裡仍然必須使用原始 df_list 中的數據進行檢查，因為它包含 '客戶需求日'
        for person_id, df in df_list.items(): 
            if not df.empty and '客戶需求日' in df.columns and DATE_COLUMN in df.columns:
                df['客戶需求日'] = pd.to_datetime(df['客戶需求日']).dt.normalize()
                if (pd.to_datetime(df[DATE_COLUMN]) > df['客戶需求日']).any():
                    is_anyone_overdue = True
                    break
        
        # 2. 判斷平衡條件 (使用手動更新的日期)
        slowest_person = max(current_latest_start_date, key=current_latest_start_date.get)
        fastest_person = min(current_latest_start_date, key=current_latest_start_date.get)
        
        # 檢查平衡：開工日最晚者與最早者的差距
        is_balanced_by_time = (current_latest_start_date[slowest_person] - current_latest_start_date[fastest_person] <= timedelta(days=1))
        
        # 停止條件：達到時間平衡 AND 無人超期
        if is_balanced_by_time and not is_anyone_overdue:
            print(f"迭代 {i+1}: 開工日已平衡，且無人超期，停止轉移。")
            break
            
        # 轉移條件：如果仍未平衡 OR 仍有人超期，則進行轉移。
        if not is_balanced_by_time or is_anyone_overdue:
            
            df_source = df_list[slowest_person]
            df_target = df_list[fastest_person]
            
            if df_source.empty or slowest_person == fastest_person:
                break
                
            if DATE_COLUMN not in df_source.columns:
                 print(f"錯誤: 轉移來源缺少 '{DATE_COLUMN}' 欄位。")
                 break
                 
            # 選擇開工日最晚的那一筆工單作為轉移對象
            transfer_order = df_source.sort_values(by=[DATE_COLUMN, '刀次'], ascending=[False, True]).iloc[0]
            
            # 執行轉移 (DataFrame 操作)
            df_list[slowest_person] = df_source.drop(transfer_order.name) 
            df_list[fastest_person] = pd.concat([df_target, transfer_order.to_frame().T])
            df_list[fastest_person]['人員'] = fastest_person 
            
            # ⭐ 關鍵修正：手動更新目標人員的開工日，使其變慢
            # 假設每次轉移會讓目標人員的開工日期推遲 1 天 (這是一個粗略的模擬)
            current_latest_start_date[fastest_person] += timedelta(days=1)
            # 手動將來源人員的開工日設定為其剩餘工單的最大值 (確保它能真正變輕)
            try:
                current_latest_start_date[slowest_person] = pd.to_datetime(df_list[slowest_person][DATE_COLUMN]).max().date()
            except Exception:
                 # 如果來源工單為空，則設為一個較早的日期
                 current_latest_start_date[slowest_person] = pd.to_datetime('1900-01-01').date()
            
            print(f"迭代 {i+1}: 從 {slowest_person} (原開工日: {latest_start_date[slowest_person]} -> 轉移後模擬: {current_latest_start_date[slowest_person]}) 轉移 1 筆工單給 {fastest_person} (原開工日: {latest_start_date[fastest_person]} -> 轉移後模擬: {current_latest_start_date[fastest_person]})")
            
        else:
            break


    # 3. 最終輸出清理：不執行任何 rename 操作
            
    return df_list['A159'], df_list['A830'], df_list['B201']


def merge_final_result(A830_part_end, B201_part_end, A159_part_end):

    df_list = [A830_part_end, B201_part_end, A159_part_end]

    non_empty_dfs = [df for df in df_list if not df.empty]

    if not non_empty_dfs:
        print("警告: 所有傳入的 DataFrame 均為空，返回空的 DataFrame.")
        return pd.DataFrame()
    
    final_df = pd.concat(non_empty_dfs, ignore_index=True)

    if "預計開工日" in final_df.columns:
        final_df['預計開工日'] = pd.to_datetime(final_df['預計開工日'], errors='coerce')
    else:
        print("錯誤: 整合後的 DataFrame 缺少 '預計開工日'欄位 無法排序")
        return
    
    sort_cols = ['預計開工日', '主工單編號']
    available_sort_cols = [col for col in sort_cols if col in final_df.columns]
    final_df = final_df.sort_values(by=available_sort_cols, ascending=True)

    # 整合時間格式
    DATE_COLS_TO_FORMAT = ['客戶需求日', '預計開工日']

    for col in DATE_COLS_TO_FORMAT:
        if col in final_df.columns:
            final_df[col] = pd.to_datetime(final_df[col], errors='coerce')
            final_df[col] = final_df[col].dt.strftime('%Y/%#m/%#d')
                         
    final_df = final_df.reset_index(drop=True)
    print(f"成功整合 {len(non_empty_dfs)} 個 DataFrame,總行數: {len(final_df)}")

    return final_df


def main():

    # 初階段工單分類 (輸入開始結束日期，根據config找到基本資料)
    result = process_schedule_data()  

    df_remaining_first = result["remaining_df"]
    df_no_pair = result["no_pair_df"]
    df_paired = result["paired_df"]
    df_history = result["df_history"]
    df_88 = result["paired_df_88"]
    df_68 = result["paired_df_68"]
    df_d1 = result["paired_df_d1"]
    base_df = result["base_df"]

    # 無須配對的工單進行分組尚未排程
    df_no_pair_split = split_no_pair_rows(df_no_pair)

    # 配對的工單進行分組尚未排程
    df_paired_split_1, extra_remaining_1 = split_paired_rows(df_paired)
    df_paired_split_2, extra_remaining = split_paired_rows_step1(df_paired_split_1, extra_remaining_1)

    # check 依照刀次為0去找上一筆刀次不為0的資料 確認是否為相似品號 如果不是就有異常
    df_paired_split_3, extra_remaining_2 = check_df_paired_split_2(df_paired_split_2, extra_remaining, base_df)

    # 合併 extra_remaining_2 並去除重複的工單
    df_remaining = pd.concat([df_remaining_first, extra_remaining_2], ignore_index=True)
    df_remaining = df_remaining.drop_duplicates(subset="工單編號", keep="last")

    # 將剩餘可不用匹配就可以獨立的工單找出 根據基本資料 搭1料號為0或空 , 剩餘數量小於刀次就要強制多切
    result_second = process_schedule_data_second(df_remaining, base_df)
    df_reamining_second = result_second["remaining_df"]
    df_no_pair_second = result_second["no_pair_df"]
    df_reamining_second_final = cleanup_remaining_df(df_reamining_second)
    
    # 新增功能 找到可以搭配的 88cm and 68cm
    df_paired_split, remaining = remaining_cut_clean_and_repair(df_paired_split_3, df_reamining_second_final, mode = "88cm") 
    df_paired_split, remaining = remaining_cut_clean_and_repair(df_paired_split, remaining, mode = "68cm") 

    # 處理remaining中可以搭配的 加入一種料號 兩種以上搭配方法的邏輯
    df_paired_split_final, remaining_final = remaining_paired_detail(df_paired_split, remaining, base_df) 

    # 剩餘工單排序 分配
    result_final = final_doAssignAndSort_DEBUG(df_paired_split_final, df_no_pair_split, df_no_pair_second, remaining_final, base_df)
    final_remaining = result_final["final_remaining"]
    all = result_final["all"]
    debug = result_final["debug_log"]
    df_no_pair_second_mutable = result_final["df_no_pair_second_mutable"]
    df_no_pair_split_mutable = result_final["df_no_pair_split_mutable"] 
    
    final_remaining_new = cleanup_remaining_df_second(final_remaining)

    final_remaining_new, all_final = cleanup_all(final_remaining_new, all)

    # 根據同公分 同料原則 優先順序 -> 1.滿足交期 2.同公分且同料 3.同公分 4.同料
    # 排版,依照預計開工時間排序
    final_output = prepare_final_schedule(all_final, final_remaining_new, df_history, base_df)

    # 工單整理
    sorted_df = do_check(final_output)
    
    # 輸入休假 彈出視窗 選擇人員以及休假日期
    sorted_df_end_list, break_dates = final_schedule_list(sorted_df,  df_history)

    # 載入歷史資料 觀察是否有前一日尚未做完的工單 要繼續做完
    sorted_df_end_list_his = schedule_history_download(sorted_df_end_list, df_history)

    # 新版分配工單
    sorted_df_end, debug_log = assign_personnel_by_similarity(sorted_df_end_list_his)

    # 人員分類
    A159_part, A830_part, B201_part = do_people(sorted_df_end)
 
    A830_part_end_list, A159_part_end_list, B201_part_end_list = final_schedule_list_second(A159_part, A830_part, B201_part, break_dates)

    # 第二次分配(負荷平衡)
    #A159_final_balanced, A830_final_balanced, B201_final_balanced = balance_completion_time(A159_part_end_list, A830_part_end_list, B201_part_end_list)

    #A830_final_balanced, A159_final_balanced, B201_final_balanced = final_schedule_list_second(A159_final_balanced, A830_final_balanced, B201_final_balanced, break_dates)

    # 將工單往前遞補滿每日的標準刀次 
    A830_part_end_merge, B201_part_end_merge, A159_part_end_merge = merge_order_cutNum(A830_part_end_list, B201_part_end_list, A159_part_end_list) 
 
    # 最後判斷每天的工單 如果有相同的主工單 那就合併~(包含其子工單也會合併)
    A830_part_end_reorder = reorder_main_and_subs(A830_part_end_merge)
    B201_part_end_reorder = reorder_main_and_subs(B201_part_end_merge)
    A159_part_end_reorder = reorder_main_and_subs(A159_part_end_merge)
    A830_part_end_1 = merge_same_day_orders_multi(A830_part_end_reorder) 
    B201_part_end_1 = merge_same_day_orders_multi(B201_part_end_reorder)  
    A159_part_end_1 = merge_same_day_orders_multi(A159_part_end_reorder) 

    # 嘉真需求 補上米平方 & 按週分組
    A830_part_end, B201_part_end, A159_part_end = Erin_use(A830_part_end_1, B201_part_end_1, A159_part_end_1, base_df)
    
    # 最後防呆 確認數量 不可有少 可多
    result_check = check_Qty(A830_part_end, B201_part_end, A159_part_end)
    print(result_check)

    # 將資料合併 並且依照開工日排列 為了for模型
    final_output = merge_final_result(A830_part_end, B201_part_end, A159_part_end)
    
    # ==============================================================================================================
    # === 寫入 Excel ===
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得目前.py檔的資料夾
    output_name = os.path.join(script_dir, "排程結果_1118(1119).xlsx")   # 存到同一層資料夾
    if os.path.exists(output_name):
        try:
            os.remove(output_name)
        except PermissionError:
            raise Exception(f"Excel 檔案 {output_name} 已開啟，請先關閉後再執行。")
    with pd.ExcelWriter(output_name, engine="openpyxl") as writer:
        base_df.to_excel(writer, sheet_name="基本資料", index=False)
        #df_88.to_excel(writer, sheet_name="88", index=False)
        #df_68.to_excel(writer, sheet_name="68", index=False)
        #df_d1.to_excel(writer, sheet_name="d1", index=False)
        #df_paired.to_excel(writer, sheet_name="配對", index=False)
        #df_no_pair.to_excel(writer, sheet_name="不需配對", index=False)
        #df_remaining_first.to_excel(writer, sheet_name="剩餘工單1", index=False)

        #df_paired_split_1.to_excel(writer, sheet_name="配對後結果1", index=False)
        #extra_remaining_1.to_excel(writer, sheet_name="配對後剩餘1", index=False)
        #df_paired_split_2.to_excel(writer, sheet_name="配對後結果2", index=False)
        #extra_remaining.to_excel(writer, sheet_name="配對後剩餘2", index=False)
        #df_paired_split_3.to_excel(writer, sheet_name="配對後結果3", index=False)
        #extra_remaining_2.to_excel(writer, sheet_name="配對後剩餘3", index=False)

        #df_remaining.to_excel(writer, sheet_name="debug_remaining", index=False)
        #df_reamining_second.to_excel(writer, sheet_name="剩餘2", index=False)
        #df_reamining_second_final.to_excel(writer, sheet_name="剩餘2_final", index=False)

        #df_paired_split.to_excel(writer, sheet_name="配對後結果1+2", index=False)
        #remaining.to_excel(writer, sheet_name="配對後剩餘1+2", index=False)

        #df_no_pair_second.to_excel(writer, sheet_name="第二次配對後不需配對", index=False)
        #df_paired_split_final.to_excel(writer, sheet_name="配對後final", index=False)
        #remaining_final.to_excel(writer, sheet_name="配對後剩餘final", index=False)
        #df_no_pair_split.to_excel(writer, sheet_name="不需配對", index=False)

        #debug.to_excel(writer, sheet_name="doassignandsort_debug", index=False)
        #final_remaining.to_excel(writer, sheet_name="最後剩餘", index=False)
        #df_no_pair_second_mutable.to_excel(writer, sheet_name="all中nopair2", index=False)
        #df_no_pair_split_mutable.to_excel(writer, sheet_name="all中nopair1", index=False)
        #all.to_excel(writer, sheet_name="未分配前all", index=False)
        #final_remaining_new.to_excel(writer, sheet_name="最後剩餘修正", index=False)
        #all_final.to_excel(writer, sheet_name="未分配前all_final", index=False)

        #final_output.to_excel(writer, sheet_name="排列初版", index=False)

        #sorted_df.to_excel(writer, sheet_name="排列初版_all", index=False)

        #sorted_df_end_list.to_excel(writer, sheet_name="排列結果", index=False)

        #sorted_df_end_list_his.to_excel(writer, sheet_name="排列結果+his", index=False)

        #sorted_df_end.to_excel(writer, sheet_name="排列結果+人員", index=False)
        '''
        if debug_log: # 確保字典不是空的
            # 1. 將字典轉換為 DataFrame (orient='index' 將 ID 作為索引)
            debug_df = pd.DataFrame.from_dict(debug_log, orient='index')
            
            # 2. 將索引 (唯一主工單ID) 變為欄位
            debug_df.reset_index(inplace=True, names=['唯一主工單ID'])
            
            # 3. 輸出到 Excel
            debug_df.to_excel(writer, sheet_name="debug_log", index=False)
        '''
        
        A159_part.to_excel(writer, sheet_name="家偉_初版", index=False)
        B201_part.to_excel(writer, sheet_name="旺斌_初版", index=False)
        A830_part.to_excel(writer, sheet_name="容合_初版", index=False)
        
        #A159_part_end_list.to_excel(writer, sheet_name="家偉_merge前", index=False)
        #B201_part_end_list.to_excel(writer, sheet_name="旺斌_merge前", index=False)
        #A830_part_end_list.to_excel(writer, sheet_name="容合_merge前", index=False)

        #A159_final_balanced.to_excel(writer, sheet_name="家偉_平衡", index=False)
        #B201_final_balanced.to_excel(writer, sheet_name="旺斌_平衡", index=False)
        #A830_final_balanced.to_excel(writer, sheet_name="容合_平衡", index=False)

        A159_part_end_merge.to_excel(writer, sheet_name="家偉", index=False)
        B201_part_end_merge.to_excel(writer, sheet_name="旺斌", index=False)
        A830_part_end_merge.to_excel(writer, sheet_name="容合", index=False)

        A159_part_end_reorder.to_excel(writer, sheet_name="家偉_reorder", index=False)
        B201_part_end_reorder.to_excel(writer, sheet_name="旺斌_reorder", index=False)
        A830_part_end_reorder.to_excel(writer, sheet_name="容合_reorder", index=False)

        A159_part_end.to_excel(writer, sheet_name="家偉_final", index=False)
        B201_part_end.to_excel(writer, sheet_name="旺斌_final", index=False)
        A830_part_end.to_excel(writer, sheet_name="容合_final", index=False)

        #final_output.to_excel(writer, sheet_name="最終結果", index=False)

        #final.to_excel(writer, sheet_name="最終結果", index=False)
        #final_81B_assigned.to_excel(writer, sheet_name="最終結果_81B分類", index=False)
        #remaining_unassigned.to_excel(writer, sheet_name="最終結果_其他", index=False)

    # Excel 欄寬調整
    def get_display_width(text):
        width = 0
        for ch in str(text):
            width += 4 if unicodedata.east_asian_width(ch) in ('F', 'W', 'A') else 2
        return width

    wb = load_workbook(output_name)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, name="Calibri")

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.font = Font(name="Calibri")

        for col in ws.columns:
            max_len = max((get_display_width(cell.value) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    wb.save(output_name)

    print("End running")

    return output_name

# 執行 GUI
if __name__ == "__main__":
    app = ScheduleApp()
    app.mainloop()
