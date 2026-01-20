'''
初版 55-60刀次  
品號排序不夠嚴謹
2025/10/07 加入81B, 81C
'''
import os
import platform
import subprocess
import sys
import unicodedata
import json
import math
import re
import traceback
import numpy as np
import pandas as pd
import customtkinter as ctk
import tkinter as tk
import tkinter.font as tkfont
from typing import Tuple, Dict, Any, List
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta
from tkcalendar import DateEntry, Calendar
from collections import defaultdict
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

# 初始化 GUI 風格
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

global start_date, end_date

class ScheduleApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("工單排程工具")
        self.geometry("500x300")

        self.label = ctk.CTkLabel(self, text="碳帶工單排程", font=("Arial", 20))
        self.label.pack(pady=20)

        self.run_button = ctk.CTkButton(self, text="執行排程", command=self.run_schedule)
        self.run_button.pack(pady=10)

        self.output_label = ctk.CTkLabel(self, text="")
        self.output_label.pack(pady=10)

        # 綁定關閉事件
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def run_schedule(self):
        self.run_button.configure(state="disabled")
        self.output_label.configure(text="⏳ 處理中，請稍候...")
        self.update_idletasks()  # 強制更新畫面
        try:
            output_path = main() # 排程主程式
            self.output_label.configure(text="✅ 排程完成！結果已輸出至 Excel")
            self.after(500, lambda: open_file(output_path))

        except Exception as e:
            messagebox.showerror("錯誤", f"執行失敗，請確認資料格式或 Excel 是否已開啟。\n詳細錯誤：{str(e)[:200]}")
            print("🔴 發生錯誤：", e)
            traceback.print_exc()  # 🔍 印出完整錯誤堆疊
            self.output_label.configure(text="❌ 排程失敗")

        finally:
            self.run_button.configure(state="normal")

    def on_closing(self):
        # 這邊可以放你要結束前要做的事情
        self.destroy()   # 關閉視窗
        sys.exit()       # 強制結束程式

# 選擇日期區間彈出視窗
class DateFilterDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("選擇日期區間")
        self.geometry("500x360")  # 放大視窗

        self.start_date_choose = None
        self.end_date_choose = None

        # 標題文字
        ctk.CTkLabel(self, text="📅 請選擇日期區間", font=("微軟正黑體", 20, "bold")).pack(pady=(15, 10))

        # 設定共用樣式
        label_font = ("微軟正黑體", 14)
        dateentry_style = {
            "font": ("微軟正黑體", 14),
            "width": 20,
            "background": "#004080",
            "foreground": "white",
            "borderwidth": 2,
            "date_pattern": "yyyy-mm-dd"
        }

        # 起始日期
        ctk.CTkLabel(self, text="起始日期", font=label_font).pack()
        #self.start_entry = DateEntry(self, **dateentry_style)
        self.start_entry = DateEntry(self, year=2025, month=10, day=7, **dateentry_style) # test use
        self.start_entry.pack(pady=(5, 15))

        # 結束日期
        ctk.CTkLabel(self, text="結束日期", font=label_font).pack()
        #self.end_entry = DateEntry(self, **dateentry_style)
        self.end_entry = DateEntry(self, year=2025, month=10, day=31, **dateentry_style) # test use
        self.end_entry.pack(pady=(5, 15))

        # 錯誤訊息
        self.error_label = ctk.CTkLabel(self, text="", text_color="red", font=("微軟正黑體", 12))
        self.error_label.pack(pady=5)

        # 確定按鈕
        ctk.CTkButton(self, text="✅ 確定", font=("微軟正黑體", 14), command=self.confirm).pack(pady=15)

        self.grab_set()  # 鎖定視窗
        self.focus()

        self.after(100, self.confirm) # 開發使用

    def confirm(self): 
        try:
            self.start_date_choose = datetime.strptime(self.start_entry.get(), "%Y-%m-%d")
            self.end_date_choose = datetime.strptime(self.end_entry.get(), "%Y-%m-%d")
            self.destroy()
        except ValueError:
            self.error_label.configure(text="⚠️ 請選擇有效的日期")


def process_schedule_data():
    
    # 讓使用者選擇排程資料 Excel
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="請選擇排程資料 Excel 檔案",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not file_path:
        print("❌ 沒有選擇任何檔案，程式結束。")
        return  # 改用 return 而不是 exit()

    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得腳本所在目錄
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    # 基本資料路徑
    #base_path = r"E:\ribbon_schedule\\data\\基本資料-20250617-All.xlsx"
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return

    # 讀取資料
    order_df = pd.read_excel(file_path, dtype=str)
    
    base_df = pd.read_excel(base_path, dtype=str)

    # 過濾條件 1. 忽略81R, 81T  2. 狀態為Released   3. 根據選擇日期範圍去篩選開工時間資料
    order_df = order_df[~order_df["工單號碼"].str.startswith(("81R", "81T"))]
    order_df = order_df[order_df["狀態"].str.lower() == "released"] # complete released
    order_df = choose_date(order_df)

    # 數值處理
    base_df["寬度Cm"] = base_df["寬度Cm"].astype(float)
    base_df["車數"] = base_df["車數"].astype(float)
    base_df["搭1產出車數"] = base_df["搭1產出車數"].fillna(0).astype(float)
    order_df["料號"] = order_df["料號"].str.strip()
    order_df["開工數量"] = order_df["開工數量"].astype(int)
    

    # 合併資料
    merged = pd.merge(order_df, base_df, how="left", on="料號")
    merged["總長度(cm)"] = merged["寬度Cm"] * merged["車數"]
    merged["刀次"] = (merged["開工數量"] / merged["車數"]).round(2)

    # 無須配對條件
    '''
    mask_no_pair = (
        merged["工單號碼"].str.startswith(("81A", "81B")) &
        (merged["刀次"] % 1 == 0) &
        ((merged["總長度(cm)"] == 88) | (merged["總長度(cm)"] == 89) | (merged["總長度(cm)"] == 68))
    )
    '''
    mask_no_pair = (
        merged["工單號碼"].str.startswith(("81A", "81B")) &
        (merged["搭1產出車數"] == 0) &
        (merged["刀次"] % 1 == 0)
    )
    
    no_pair_df = merged[mask_no_pair].copy()
    remaining_df = merged[~mask_no_pair].copy()

    # 88cm 配對（備註配對）
    pair_rows_88 = []
    df_81c_remarks = merged[(merged["工單號碼"].str.startswith("81C")) & (merged["備註"].notna())]

    for _, c_order in df_81c_remarks.iterrows():
        pair_order_id = c_order["備註"]
        matched_rows = remaining_df[remaining_df["工單號碼"] == pair_order_id]
        if matched_rows.empty:
            continue
        a_order = matched_rows.iloc[0]
        for cars in range(1, 5):
            total_cm = a_order["總長度(cm)"] + c_order["寬度Cm"] * cars
            if abs(total_cm - 88) <= 1:
                a_copy = a_order.copy()
                c_copy = c_order.copy()
                c_copy["車數"] = cars
                c_copy["總長度(cm)"] = c_copy["寬度Cm"] * cars
                c_copy["刀次"] = 0
                pair_rows_88.extend([a_copy, c_copy])
                remaining_df = remaining_df[remaining_df["工單號碼"] != pair_order_id]
                remaining_df = remaining_df[remaining_df["工單號碼"] != c_order["工單號碼"]]
                break

    # 68cm 配對（備註配對）
    pair_rows_68 = []
    df_81c_remain = remaining_df[(remaining_df["工單號碼"].str.startswith("81C")) & (remaining_df["備註"].notna())]

    for _, c_order in df_81c_remain.iterrows():
        pair_order_id = c_order["備註"]
        matched_rows = remaining_df[remaining_df["工單號碼"] == pair_order_id]
        if matched_rows.empty:
            continue
        a_order = matched_rows.iloc[0]
        for cars in range(1, 5):
            total_cm = a_order["總長度(cm)"] + c_order["寬度Cm"] * cars
            if abs(total_cm - 68) <= 1:
                a_copy = a_order.copy()
                c_copy = c_order.copy()
                c_copy["車數"] = cars
                c_copy["總長度(cm)"] = c_copy["寬度Cm"] * cars
                c_copy["刀次"] = 0
                pair_rows_68.extend([a_copy, c_copy])
                remaining_df = remaining_df[remaining_df["工單號碼"] != pair_order_id]
                remaining_df = remaining_df[remaining_df["工單號碼"] != c_order["工單號碼"]]
                break

    # ➕ 搭1料號配對（修正版）
    pair_rows_d1 = []

    # 重新建立剩餘工單索引
    remaining_df = remaining_df.reset_index(drop=True)
    

    # 建立料號到工單列表的映射
    remaining_by_item = {}
    for _, row in remaining_df.iterrows():
        remaining_by_item.setdefault(row["料號"], []).append(row)

    for _, main_order in remaining_df.iterrows():
        main_order_id = main_order["工單號碼"]
        main_item_code = main_order["料號"]

        # 找基本資料搭1料號
        base_info = base_df[base_df["料號"] == main_item_code]
        if base_info.empty:
            continue
        d1_code = base_info.iloc[0].get("搭1料號")
        d1_car_count = base_info.iloc[0].get("搭1產出車數")
        if pd.isna(d1_code) or d1_code == "":
            continue

        d1_base_info = base_df[base_df["料號"] == d1_code]
        if d1_base_info.empty:
            continue
        d1_width = d1_base_info.iloc[0].get("寬度Cm")
        if pd.isna(d1_car_count) or pd.isna(d1_width):
            continue

        if d1_code not in remaining_by_item:
            continue

        d1_candidates = remaining_by_item[d1_code]

        for d1_order in d1_candidates:
            d1_order_id = d1_order["工單號碼"]

            d1_copy = d1_order.copy()
            d1_copy["車數"] = d1_car_count
            d1_copy["刀次"] = 0
            d1_copy["寬度Cm"] = d1_width
            d1_copy["總長度(cm)"] = d1_width * d1_car_count

            main_copy = main_order.copy()
            main_copy["備註配對工單"] = d1_order_id
            d1_copy["備註配對工單"] = main_order_id

            pair_rows_d1.extend([main_copy, d1_copy])

    # 移除已配對的副工單
    paired_d1_order_ids = set(d["工單號碼"] for d in pair_rows_d1 if "工單號碼" in d)
    remaining_df = remaining_df[~remaining_df["工單號碼"].isin(paired_d1_order_ids)].copy()

    # 合併所有配對結果
    paired_df_88 = pd.DataFrame(pair_rows_88)
    paired_df_68 = pd.DataFrame(pair_rows_68)
    paired_df_d1 = pd.DataFrame(pair_rows_d1)
    all_paired_df = pd.concat([df for df in [paired_df_88, paired_df_68, paired_df_d1] if not df.empty], ignore_index=True)

    # 格式化函式
    def format_df(df):
        df = df.rename(columns={
            "開工日期": "預計開工日",
            "預計完工日期": "預計完工日",
            "工單號碼": "工單編號",
            "料號": "品號",
            "開工數量": "預估良品數",
            "寬度Cm": "公分",
            "客戶需求日期": "客戶需求日"
        })

        # 統一欄位型別與補空值
        df["公分"] = pd.to_numeric(df["公分"], errors="coerce").round(1)
        df["預估良品數"] = pd.to_numeric(df["預估良品數"], errors="coerce").fillna(0).astype(int)
        df["車數"] = pd.to_numeric(df["車數"], errors="coerce").fillna(0).astype(int)
        df["刀次"] = pd.to_numeric(df["刀次"], errors="coerce").fillna(0).round(2)
        df["人員"] = ""
        df["餘量"] = df["預估良品數"]
        df["生產註記"] = ""
        
        # 補足缺少的欄位
        for col in ["預計開工日", "人員", "工單編號", "品號", "餘量",
                    "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]:
            if col not in df.columns:
                df[col] = ""

        return df[["預計開工日", "人員", "工單編號", "品號", "餘量",
                    "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]]


    remaining_df = format_df(remaining_df)
    no_pair_df = format_df(no_pair_df)
    if not all_paired_df.empty:
        all_paired_df = format_df(all_paired_df)

    return {
            "remaining_df": remaining_df,
            "no_pair_df": no_pair_df,
            "paired_df": all_paired_df
        }


# 無須配對工單組
# 若工單數量大且不用換料無需配對 一天最多60刀次
def split_no_pair_rows(df_no_pair, max_knife=55):
    rows = []

    for _, row in df_no_pair.iterrows():
        knife = int(row["刀次"])
        cars = float(row["車數"])
        while knife > 0:
            split_knife = min(knife, max_knife)
            knife -= split_knife

            new_row = row.copy()
            new_row["刀次"] = split_knife
            new_row["預估良品數"] = int(cars * split_knife)
            new_row["餘量"] = new_row["預估良品數"]

            rows.append(new_row)

    return pd.DataFrame(rows)


# 配對完成工單
# 先根據配對工單數量先決定好主工單刀次 剩下的再去配對其他工單 目前先他剩下的特別拿出
def split_paired_rows(df_paired, max_cut=60):
    df_paired = df_paired.copy().reset_index(drop=True)
    paired_rows = []
    remaining_rows = []

    # 保存原始數量
    original_qty_map = df_paired.set_index("工單編號")["預估良品數"].astype(float).to_dict()

    # 第一次生成主工單與子工單
    i = 0
    while i < len(df_paired):
        current_row = df_paired.iloc[i]
        current_cut = math.ceil(float(current_row["刀次"]))
        current_car = int(current_row["車數"])

        if current_cut > 0:
            main_row = current_row.copy()
            sub_rows = []

            # 搜集後面子工單（刀次 = 0）
            j = i + 1
            while j < len(df_paired) and int(df_paired.iloc[j]["刀次"]) == 0:
                sub_rows.append(df_paired.iloc[j].copy())
                j += 1

            remaining_cut = current_cut
            while remaining_cut > 0:
                split_cut = min(max_cut, remaining_cut)

                # 主工單分段
                main_piece = main_row.copy()
                main_piece["刀次"] = split_cut
                main_piece["預估良品數"] = split_cut * current_car
                main_piece["餘量"] = main_piece["預估良品數"]
                paired_rows.append(main_piece)

                # 子工單同樣拆分
                for sub_row in sub_rows:
                    sub_car = int(sub_row["車數"])
                    if sub_car == 0:
                        continue
                    sub_piece = sub_row.copy()
                    sub_piece["刀次"] = 0
                    sub_piece["預估良品數"] = split_cut * sub_car
                    sub_piece["餘量"] = sub_piece["預估良品數"]
                    paired_rows.append(sub_piece)

                remaining_cut -= split_cut

            i = j
        else:
            remaining_rows.append(current_row.copy())
            i += 1

    # 第二階段：檢查超量並修正主工單
    used_df = pd.DataFrame(paired_rows).reset_index(drop=True)
    used_summary = used_df.groupby("工單編號")["預估良品數"].sum().to_dict()

    for key, total_used in used_summary.items():
        original_qty = float(original_qty_map.get(key, 0))
        if total_used > original_qty:
            # 超過原始數量，需縮減主工單
            over_qty = total_used - original_qty

            # 找出所有這個工單作為主的段落
            mask = (used_df["工單編號"] == key) & (used_df["刀次"].astype(float) > 0)
            main_pieces = used_df.loc[mask].reset_index()  # 重置索引，確保 idx 唯一

            for idx in main_pieces.index[::-1]:  # 從最後一段開始減
                piece_qty = float(main_pieces.at[idx, "預估良品數"])
                piece_car = float(main_pieces.at[idx, "車數"])

                if over_qty <= 0:
                    break

                if piece_qty <= over_qty:
                    # 整段刪掉
                    drop_idx = main_pieces.at[idx, "index"]  # 對應原 used_df 的索引
                    used_df.drop(drop_idx, inplace=True)
                    over_qty -= piece_qty
                else:
                    # 部分減少
                    new_qty = piece_qty - over_qty
                    new_cut = math.ceil(new_qty / piece_car) if piece_car > 0 else 0
                    update_idx = main_pieces.at[idx, "index"]  # 對應原 used_df 的索引
                    used_df.at[update_idx, "預估良品數"] = new_qty
                    used_df.at[update_idx, "刀次"] = new_cut
                    used_df.at[update_idx, "餘量"] = new_qty
                    over_qty = 0

    used_df = used_df.reset_index(drop=True)

    # 生成剩餘量
    for _, row in df_paired.iterrows():
        key = row["工單編號"]
        original_qty = float(original_qty_map[key])
        used_qty = used_df[used_df["工單編號"] == key]["預估良品數"].sum()
        remain_qty = original_qty - used_qty
        if remain_qty > 0:
            remain_row = row.copy()
            remain_row["預估良品數"] = remain_qty
            remain_row["餘量"] = remain_qty
            remaining_rows.append(remain_row)

    # === 新增掃描：把刀次=0 且前一筆刀次=0 的子工單移到剩餘 ===
    final_used_rows = []
    prev_cut = None
    for idx, row in used_df.iterrows():
        current_cut = float(row.get("刀次", 0))
        if current_cut == 0 and prev_cut == 0:
            # 移到剩餘
            remaining_rows.append(row)
        else:
            final_used_rows.append(row)
        prev_cut = current_cut

    used_df = pd.DataFrame(final_used_rows).reset_index(drop=True)
    remaining_df = pd.DataFrame(remaining_rows).reset_index(drop=True)

    return used_df, remaining_df


# 延續split_paired_rows的Step 1 分配結束後再將資料做一次數量校正 
# 不會有remaining部份
def split_paired_rows_step1(df_paired, max_cut=60):

    df_paired = df_paired.copy().reset_index(drop=True)
    paired_rows = []
    remaining_rows = []

    # 第一次生成主工單與子工單
    i = 0
    while i < len(df_paired):
        current_row = df_paired.iloc[i]
        current_cut = math.ceil(float(current_row["刀次"]))
        current_car = int(current_row["車數"])

        if current_cut > 0:
            main_row = current_row.copy()
            sub_rows = []

            # 搜集後面子工單（刀次 = 0）
            j = i + 1
            while j < len(df_paired) and int(df_paired.iloc[j]["刀次"]) == 0:
                sub_rows.append(df_paired.iloc[j].copy())
                j += 1

            remaining_cut = current_cut
            while remaining_cut > 0:
                split_cut = min(max_cut, remaining_cut)

                # 主工單分段
                main_piece = main_row.copy()
                main_piece["刀次"] = split_cut
                main_piece["預估良品數"] = split_cut * current_car
                main_piece["餘量"] = main_piece["預估良品數"]
                paired_rows.append(main_piece)

                # 子工單同樣拆分
                for sub_row in sub_rows:
                    sub_car = int(sub_row["車數"])
                    if sub_car == 0:
                        continue
                    sub_piece = sub_row.copy()
                    sub_piece["刀次"] = 0
                    sub_piece["預估良品數"] = split_cut * sub_car
                    sub_piece["餘量"] = sub_piece["預估良品數"]
                    paired_rows.append(sub_piece)

                remaining_cut -= split_cut

            i = j
        else:
            remaining_rows.append(current_row.copy())
            i += 1

    return pd.DataFrame(paired_rows)


# 將剩餘可不用匹配就可以獨立的工單找出 根據基本資料 搭1料號為0或空 , 剩餘數量小於刀次就要強制多切
def process_schedule_data_second(order_df):
    
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得設定json檔案 -> 基本資料路徑
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
  
    # 基本資料路徑
    #base_path = r"E:\\ribbon_schedule\\data\\基本資料-20250617-All.xlsx"
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return

    # 讀取資料
    base_df = pd.read_excel(base_path, dtype=str)

    # 數值處理
    base_df["寬度Cm"] = base_df["寬度Cm"].astype(float)
    base_df["車數"] = base_df["車數"].astype(float)
    base_df["搭1產出車數"] = base_df["搭1產出車數"].fillna(0).astype(float)
    base_df = base_df.rename(columns={"料號": "品號"})

    order_df["品號"] = order_df["品號"].str.strip()
    order_df["預估良品數"] = order_df["預估良品數"].astype(int)
    order_df = order_df.rename(columns={"車數": "車數_error"}) #因為車數不準確

    # 合併資料
    merged = pd.merge(order_df, base_df, how="left", on="品號")
    #print(merged)
    merged["刀次"] = (merged["預估良品數"] / merged["車數"]).round(2)
    merged["總長度(cm)"] = (merged["寬度Cm"] * merged["車數"])

    # 無須配對條件
    '''
    mask_no_pair = (
        merged["工單編號"].str.startswith(("81A", "81B")) &
        (merged["刀次"] % 1 == 0) &
        ((merged["總長度(cm)"] == 88) | (merged["總長度(cm)"] == 89))
    )
    '''
    mask_no_pair = (
        (merged["工單編號"].str.startswith(("81A", "81B")) &
        (merged["搭1產出車數"] == 0) &
        (merged["刀次"] % 1 == 0)
        )
        | 
        ((merged["刀次"] % 1 == 0) & 
         (merged["預估良品數"] == (merged["車數"] * merged["刀次"]).round()) &
         ((merged["總長度(cm)"] == 88) | (merged["總長度(cm)"] == 89) | (merged["總長度(cm)"] == 68))
        )
    )
    no_pair_df = merged[mask_no_pair].copy()
    remaining_df = merged[~mask_no_pair].copy()

    # ⬇️ 拆出整刀邏輯（非整數刀次但可先出整刀）
    split_rows = []
    remaining_updates = []

    for idx, row in remaining_df.iterrows():
        try:
            car = float(row["車數"])
            cut = float(row["刀次"])
            est_qty = int(row["預估良品數"])
            width = float(row["寬度Cm"])
            output_car = float(row["搭1產出車數"])

            if cut % 1 != 0 and output_car == 0:
                #full_cut = int(cut)
                full_cut = math.ceil(cut)
                full_qty = int(car * full_cut)
                remain_qty = est_qty - full_qty

                if full_cut > 0:
                    full_row = row.copy()
                    full_row["刀次"] = full_cut
                    full_row["預估良品數"] = full_qty
                    full_row["餘量"] = full_qty
                    split_rows.append(full_row)

                    remain_row = row.copy()
                    remain_row["刀次"] = round(cut - full_cut, 2)
                    remain_row["預估良品數"] = remain_qty
                    remain_row["餘量"] = remain_qty
                    remaining_updates.append(remain_row)

                    remaining_df.drop(index=idx, inplace=True)
        except:
            continue

    if split_rows:
        no_pair_df = pd.concat([no_pair_df, pd.DataFrame(split_rows)], ignore_index=True)

    if remaining_updates:
        remaining_df = pd.concat([remaining_df, pd.DataFrame(remaining_updates)], ignore_index=True)

    # 格式化欄位
    def format_df(df):
        for col in ["預計開工日", "人員", "工單編號", "品號", "餘量",
                    "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]:
            if col not in df.columns:
                df[col] = ""
        return df[["預計開工日", "人員", "工單編號", "品號", "餘量",
                   "公分", "車數", "刀次", "預估良品數", "預計完工日", "生產註記", "客戶需求日"]]

    remaining_df = format_df(remaining_df)
    remaining_df = remaining_df[remaining_df["預估良品數"].astype(float) >= 0].copy()

    no_pair_df = format_df(no_pair_df)

    # ⬇️ 刀次 > 55 時再拆分
    expanded_rows = []

    for _, row in no_pair_df.iterrows():
        try:
            cut = float(row["刀次"])
            car = float(row["車數"])
            qty = int(row["預估良品數"])

            if cut > 55:
                full_sets = int(cut // 55)
                remaining_cut = round(cut % 55, 2)

                for _ in range(full_sets):
                    new_row = row.copy()
                    new_row["刀次"] = 55
                    new_row["預估良品數"] = int(car * 55)
                    new_row["餘量"] = new_row["預估良品數"]
                    expanded_rows.append(new_row)

                if remaining_cut > 0:
                    new_row = row.copy()
                    new_row["刀次"] = remaining_cut
                    new_row["預估良品數"] = int(car * remaining_cut)
                    new_row["餘量"] = new_row["預估良品數"]
                    expanded_rows.append(new_row)
            else:
                expanded_rows.append(row)
        except:
            expanded_rows.append(row)

    no_pair_df = pd.DataFrame(expanded_rows)

    return {
        "remaining_df": remaining_df,
        "no_pair_df": no_pair_df
    }


def prepare_final_schedule(df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining):
    # 建立主工單欄位
    df_no_pair_split["主工單編號"] = df_no_pair_split["工單編號"]
    df_no_pair_second["主工單編號"] = df_no_pair_second["工單編號"]

    df_paired_split = df_paired_split.copy()
    df_paired_split["主工單編號"] = None

    final_remaining = final_remaining.copy()
    final_remaining["主工單編號"] = None

    # 判斷主子工單
    for pos in range(len(df_paired_split)):
        row = df_paired_split.iloc[pos]
        if row["刀次"] != 0:
            df_paired_split.at[df_paired_split.index[pos], "主工單編號"] = row["工單編號"]
        else:
            prev_df = df_paired_split.iloc[:pos]
            prev_main = prev_df[prev_df["刀次"] != 0]
            if prev_main.empty:
                df_paired_split.at[df_paired_split.index[pos], "主工單編號"] = None  # 或其他預設值
            else:
                df_paired_split.at[df_paired_split.index[pos], "主工單編號"] = prev_main.iloc[-1]["工單編號"]

    final_remaining["刀次"] = pd.to_numeric(final_remaining["刀次"], errors="coerce").fillna(0).astype(int)

    for pos in range(len(final_remaining)):
        row_r = final_remaining.iloc[pos]
        cut_count = row_r["刀次"]

        if cut_count != 0:
            final_remaining.at[final_remaining.index[pos], "主工單編號"] = row_r["工單編號"]
        else:
            prev_df = final_remaining.iloc[:pos]
            prev_valid = prev_df[prev_df["刀次"] != 0]

            if prev_valid.empty:
                final_remaining.at[final_remaining.index[pos], "主工單編號"] = ""
            else:
                final_remaining.at[final_remaining.index[pos], "主工單編號"] = prev_valid.iloc[-1]["工單編號"]

    # 合併資料
    combined_df = pd.concat([df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining], ignore_index=True)
    
    # 這裡已修改為新的指派邏輯
    # 步驟 1: 找出所有屬於 '81A' 的主工單
    #unassigned_df = combined_df[combined_df["主工單編號"].str.startswith("81A", na=False)].copy()
    mask_81A = combined_df["主工單編號"].str.startswith("81A", na=False)
    mask_81B = combined_df["主工單編號"].str.startswith("81B", na=False)

    unassigned_df = combined_df[mask_81A | mask_81B].copy()
        
    # 步驟 2: 計算每個 '81A' 主工單的總刀次
    main_quantities = (
        unassigned_df
        .groupby("主工單編號")["刀次"]
        .sum()
        .reset_index()
        .sort_values(by="刀次", ascending=False)
        .reset_index(drop=True)
    )

    # 步驟 3: 將 '81A' 工單平均分配給 A159, B201, A830
    assign_A159, assign_B201, assign_A830 = [], [], []
    total_A159, total_B201, total_A830 = 0, 0, 0
    
    sorted_people = [
        {"name": "A159", "total": 0, "list": assign_A159},
        {"name": "B201", "total": 0, "list": assign_B201},
        {"name": "A830", "total": 0, "list": assign_A830}
    ]

    for _, row in main_quantities.iterrows():
        main_id = row["主工單編號"]
        qty = row["刀次"]
        
        # 找到目前刀次總數最少的人
        sorted_people.sort(key=lambda x: x["total"])
        person = sorted_people[0]
        
        person["list"].append(main_id)
        person["total"] += qty

    # 步驟 4: 指派人員
    combined_df.loc[combined_df["主工單編號"].isin(assign_A159), "人員"] = "A159"
    combined_df.loc[combined_df["主工單編號"].isin(assign_B201), "人員"] = "B201"
    combined_df.loc[combined_df["主工單編號"].isin(assign_A830), "人員"] = "A830"
    
    # 排序：主工單開工日、主工單編號、刀次
    main_orders = combined_df[combined_df["工單編號"] == combined_df["主工單編號"]]
    main_dates = main_orders.groupby("工單編號")["預計開工日"].min()
    combined_df["主工單開工日"] = combined_df["主工單編號"].map(main_dates)

    combined_df = combined_df.sort_values(by=["主工單開工日", "主工單編號", "刀次"], ascending=[True, True, False])

    return combined_df


# 先篩選完再來排
def choose_date(df: pd.DataFrame) -> pd.DataFrame:
    global start_date, end_date

    df = df.copy()
    df["日期篩選用"] = pd.to_datetime(df["開工日期"], errors="coerce")

    root = ctk.CTk()
    root.withdraw()
    dialog = DateFilterDialog(root)
    root.wait_window(dialog)

    start_date = dialog.start_date_choose
    end_date = dialog.end_date_choose

    # ✅ 若使用者未選擇日期就關閉，直接中止
    if not start_date and not end_date:
        raise ValueError("❌ 請選擇日期區間後再繼續")

    # ✅ 執行篩選
    if start_date and end_date:
        df = df[(df["日期篩選用"] >= start_date) & (df["日期篩選用"] <= end_date)]
    elif start_date:
        df = df[df["日期篩選用"] >= start_date]
    elif end_date:
        df = df[df["日期篩選用"] <= end_date]

    df = df.drop(columns=["日期篩選用"])

    return df


# 全部排完才刪選時間 或是 刪除多餘欄位
# 找到可以搭配的
'''
找品號 前面非數字的(英文部份)
1. 非B110A開頭 -> 88cm, 反之 -> 68cm
2. 搭配完剛好88cm or 68cm
3. 搭配完主工單數量要剛好的優先
4. 剩下沒搭配的刀次改為空
'''
def extract_prefix(item_code):
    m = re.match(r"^[^\d]+", item_code)
    return m.group(0) if m else ""

def find_88cm_combination(row_i, row_j, max_car=88):
    for ci in range(1, max_car+1):
        for cj in range(1, max_car+1):
            total = ci * row_i["公分"] + cj * row_j["公分"]
            if abs(total - 88) < 1e-6:
                if (row_i["預估良品數"] / ci).is_integer():
                    main_cut = int(row_i["預估良品數"] / ci)
                    child_qty = main_cut * cj
                    # 新增條件：子工單用量不得超過餘量
                    if child_qty <= row_j["預估良品數"]:
                        return ci, cj
    return None

def find_68cm_combination(row_i, row_j, max_car=68):
    for ci in range(1, max_car+1):
        for cj in range(1, max_car+1):
            total = ci * row_i["公分"] + cj * row_j["公分"]
            if abs(total - 68) < 1e-6:
                if (row_i["預估良品數"] / ci).is_integer():
                    main_cut = int(row_i["預估良品數"] / ci)
                    child_qty = main_cut * cj
                    # 新增條件：子工單用量不得超過餘量
                    if child_qty <= row_j["預估良品數"]:
                        return ci, cj
    return None

# 將可以配對的進行配對 根據 mode 去找要88cm or 68cm的品號
def remaining_cut_clean_and_repair(df_paired_split, df, mode="88cm", cut_limit=55):
    """
    mode: "88cm" → 處理非 B110A
          "68cm" → 處理 B110A
    """
    df = df.copy()
    df["公分"] = pd.to_numeric(df["公分"], errors="coerce")
    df["預估良品數"] = pd.to_numeric(df["預估良品數"], errors="coerce")
    df["品號前綴"] = df["品號"].apply(extract_prefix)
    df["品號結尾"] = df["品號"].str[-2:]
    df["is_B110A"] = df["品號"].str.startswith("B110A")

    merged_indices = set()
    leftover_rows = []

    # 判斷處理範圍
    if mode == "88cm":
        target_df = df[~df["is_B110A"]]
        find_func = find_88cm_combination
    elif mode == "68cm":
        target_df = df[df["is_B110A"]]
        find_func = find_68cm_combination
    else:
        raise ValueError("mode 必須是 '88cm' 或 '68cm'")

    # --- 核心配對流程 ---
    for prefix, group in target_df.groupby("品號前綴"):
        group = group.reset_index()
        n = len(group)

        for i in range(n):
            idx_i = group.at[i, "index"]
            if idx_i in merged_indices:
                continue

            row_i = df.loc[idx_i]

            for j in range(i + 1, n):
                idx_j = group.at[j, "index"]
                if idx_j in merged_indices:
                    continue
                row_j = df.loc[idx_j]

                if row_i["品號結尾"] != row_j["品號結尾"]:
                    continue

                res = find_func(row_i, row_j)
                if res:
                    ci, cj = res

                    main_car = ci
                    main_cut = int(row_i["預估良品數"] / ci)

                    child_car = cj
                    child_cut = 0
                    child_qty = main_cut * child_car

                    main_qty = row_i["預估良品數"] - child_qty
                    if main_qty < 0:
                        main_qty = 0

                    main_order = row_i.copy()
                    main_order["車數"] = main_car
                    main_order["刀次"] = main_cut
                    main_order["預估良品數"] = main_qty

                    child_order = row_j.copy()
                    child_order["車數"] = child_car
                    child_order["刀次"] = child_cut
                    child_order["預估良品數"] = child_qty

                    df_pair_split = split_pair_orders(main_order, child_order, cut_limit=cut_limit)
                    df_paired_split = pd.concat([df_paired_split, df_pair_split], ignore_index=True)

                    merged_indices.update([idx_i, idx_j])

                    leftover_qty = row_j["預估良品數"] - child_qty
                    if leftover_qty > 0:
                        leftover_order = row_j.copy()
                        leftover_order["預估良品數"] = leftover_qty
                        leftover_order["刀次"] = math.ceil(leftover_qty / row_j["車數"])
                        leftover_order["車數"] = row_j["車數"]
                        leftover_rows.append(leftover_order)

                    break

    # --- 未配對 ---
    leftover_indices = set(df.index) - merged_indices
    remaining_rows = df.loc[list(leftover_indices)].copy()
    remaining_rows["刀次"] = remaining_rows["刀次"].fillna("")

    if leftover_rows:
        remaining_rows = pd.concat([remaining_rows, pd.DataFrame(leftover_rows)], ignore_index=True)

    for col in ["品號前綴", "品號結尾", "is_B110A"]:
        if col in remaining_rows.columns:
            remaining_rows = remaining_rows.drop(columns=[col])
        if col in df_paired_split.columns:
            df_paired_split = df_paired_split.drop(columns=[col])

    return df_paired_split.reset_index(drop=True), remaining_rows.reset_index(drop=True)


# 一天最多55刀次 主子工單要拆分的話要一起拆分 
def split_pair_orders(main_order, child_order, cut_limit=55):
    """
    拆分主工單與子工單，並保持一一對應交錯輸出
    - main_order: 刀次 > 0
    - child_order: 刀次 = 0
    """
    results = []

    total_cut = int(main_order["刀次"])
    if total_cut <= 0:
        return [main_order.to_frame().T, child_order.to_frame().T]

    # 每刀對應的數量
    base_child_qty_per_cut = child_order["預估良品數"] / total_cut
    base_main_qty_per_cut = main_order["預估良品數"] / total_cut

    while total_cut > cut_limit:
        # 主工單
        new_main = main_order.copy()
        new_main["刀次"] = cut_limit
        new_main["預估良品數"] = int(base_main_qty_per_cut * cut_limit)

        # 子工單
        new_child = child_order.copy()
        new_child["刀次"] = 0
        new_child["預估良品數"] = int(base_child_qty_per_cut * cut_limit)

        # 成對加入
        results.append(new_main.to_frame().T)
        results.append(new_child.to_frame().T)

        total_cut -= cut_limit

    if total_cut > 0:
        new_main = main_order.copy()
        new_main["刀次"] = total_cut
        new_main["預估良品數"] = int(base_main_qty_per_cut * total_cut)

        new_child = child_order.copy()
        new_child["刀次"] = 0
        new_child["預估良品數"] = int(base_child_qty_per_cut * total_cut)

        results.append(new_main.to_frame().T)
        results.append(new_child.to_frame().T)

    # 串成 DataFrame
    return pd.concat(results, ignore_index=True)



def do_people(final_df):
    # 確保主工單開工日是 datetime
    final_df["主工單開工日"] = pd.to_datetime(final_df["主工單開工日"], errors='coerce')

    # 根據主工單開工日與主工單編號排序（主工單與子工單會排在一起）
    sorted_df = final_df.sort_values(by=["主工單開工日", "主工單編號", "刀次"], ascending=[True, True, False])

    # 移除輔助欄位並回傳
    #sorted_df = sorted_df.drop(columns=["主工單開工日", "主工單編號"])
    
    # 拆分三位人員
    A159_part = sorted_df[sorted_df["人員"] == "A159"].copy()
    A830_part = sorted_df[sorted_df["人員"] == "A830"].copy()
    B201_part = sorted_df[sorted_df["人員"] == "B201"].copy()

    return A159_part, A830_part, B201_part


def open_file(filepath):
    if platform.system() == 'Windows':
        os.startfile(filepath)
    elif platform.system() == 'Darwin':  # macOS
        subprocess.call(['open', filepath])
    else:  # Linux
        subprocess.call(['xdg-open', filepath])


#--------------------日期排程----------------------------        
def final_schedule_list(A159_part, A830_part, B201_part):
    
    global start_date, end_date

    # 彈出假期選擇視窗
    root = ctk.CTk()
    root.withdraw()
    dialog = MultiPersonLeaveDialog(root, start_date, end_date)  # date 格式
    root.wait_window(dialog)

    if dialog.result is None:
        raise Exception("❌ 使用者取消操作，請重新選擇休假日")

    break_dates = {
        "A159": set(pd.to_datetime(date).date() for date in dialog.result.get("A159", [])),
        "A830": set(pd.to_datetime(date).date() for date in dialog.result.get("A830", [])),
        "B201": set(pd.to_datetime(date).date() for date in dialog.result.get("B201", [])),
    }

    
    schedule_A159 = generate_schedule_for_person(A159_part, break_dates["A159"])
    schedule_A830 = generate_schedule_for_person(A830_part, break_dates["A830"])
    schedule_B201 = generate_schedule_for_person(B201_part, break_dates["B201"])

    start_date = start_date.date()

    
    # 容合
    if not schedule_A830.empty and schedule_A830["刀次"].notna().any() and (schedule_A830["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_A830["刀次"] > 0
        schedule_A830.loc[mask, "預估良品數"] = schedule_A830.loc[mask, "刀次"] * schedule_A830.loc[mask, "車數"]

        # 統一日期格式
        schedule_A830["實際排程日期"] = pd.to_datetime(schedule_A830["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_A830["主工單識別碼"] = schedule_A830["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_A830["排程依據日"] = schedule_A830.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list 用）
        schedule_A830["預計開工日"] = schedule_A830["排程依據日"]
        schedule_A830["預計完工日"] = schedule_A830["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_A830["原始順序"] = schedule_A830.index

        # 按照排程依據日 + 原始順序，確保主子工單一起移動且相對順序不變
        schedule_A830 = schedule_A830.sort_values(
            by=["排程依據日", "原始順序"],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_A830 = final_cal_list(schedule_A830, start_d=start_date, break_dates=break_dates["A830"])
            schedule_A830 = sort_by_customer_due_date(schedule_A830)
            schedule_A830 = final_cal_list(schedule_A830, start_d=start_date, break_dates=break_dates["A830"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_A830["預計開工日"] = pd.to_datetime(schedule_A830["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_A830["預計完工日"] = pd.to_datetime(schedule_A830["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_A830.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單編號", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("A830刀次欄位沒有資料或全部為空，跳過後續處理")

    
    # 家偉
    if not schedule_A159.empty and schedule_A159["刀次"].notna().any() and (schedule_A159["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_A159["刀次"] > 0
        schedule_A159.loc[mask, "預估良品數"] = schedule_A159.loc[mask, "刀次"] * schedule_A159.loc[mask, "車數"]

        # 統一日期格式
        schedule_A159["實際排程日期"] = pd.to_datetime(schedule_A159["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_A159["主工單識別碼"] = schedule_A159["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_A159["排程依據日"] = schedule_A159.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list 用）
        schedule_A159["預計開工日"] = schedule_A159["排程依據日"]
        schedule_A159["預計完工日"] = schedule_A159["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_A159["原始順序"] = schedule_A159.index

        # 按照排程依據日 + 原始順序，確保主子工單一起移動且相對順序不變
        schedule_A159 = schedule_A159.sort_values(
            by=["排程依據日", "原始順序"],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_A159 = final_cal_list(schedule_A159, start_d=start_date, break_dates=break_dates["A159"])
            schedule_A159 = sort_by_customer_due_date(schedule_A159)
            schedule_A159 = final_cal_list(schedule_A159, start_d=start_date, break_dates=break_dates["A159"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_A159["預計開工日"] = pd.to_datetime(schedule_A159["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_A159["預計完工日"] = pd.to_datetime(schedule_A159["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_A159.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單編號", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("A159刀次欄位沒有資料或全部為空，跳過後續處理")
    

    # 旺斌
    if not schedule_B201.empty and schedule_B201["刀次"].notna().any() and (schedule_B201["刀次"] > 0).any():
        # 計算預估良品數
        mask = schedule_B201["刀次"] > 0
        schedule_B201.loc[mask, "預估良品數"] = schedule_B201.loc[mask, "刀次"] * schedule_B201.loc[mask, "車數"]

        # 統一日期格式
        schedule_B201["實際排程日期"] = pd.to_datetime(schedule_B201["實際排程日期"].str.replace("-", "/"), errors="coerce")

        # 建立主工單識別碼（將主+子工單視為同一組）
        schedule_B201["主工單識別碼"] = schedule_B201["主工單編號"].fillna(method='ffill')

        # 取得每組主+子工單中最早的日期
        schedule_B201["排程依據日"] = schedule_B201.groupby("主工單識別碼")["實際排程日期"].transform("min")

        # 預計開工/完工都設為排程依據日（後續進 final_cal_list 用）
        schedule_B201["預計開工日"] = schedule_B201["排程依據日"]
        schedule_B201["預計完工日"] = schedule_B201["排程依據日"]

        # 保留原始順序（萬一你之後要還原）
        schedule_B201["原始順序"] = schedule_B201.index

        # 按照排程依據日 + 原始順序，確保主子工單一起移動且相對順序不變
        schedule_B201 = schedule_B201.sort_values(
            by=["排程依據日", "原始順序"],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 丟進排程邏輯（使用 datetime 版本）
        try:
            schedule_B201 = final_cal_list(schedule_B201, start_d=start_date, break_dates=break_dates["B201"])
            schedule_B201 = sort_by_customer_due_date(schedule_B201)
            schedule_B201 = final_cal_list(schedule_B201, start_d=start_date, break_dates=break_dates["B201"])
        except Exception as e:
            print("發生錯誤:", e)

        # 日期格式化回字串
        schedule_B201["預計開工日"] = pd.to_datetime(schedule_B201["預計開工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")
        schedule_B201["預計完工日"] = pd.to_datetime(schedule_B201["預計完工日"], errors="coerce").dt.strftime("%Y/%#m/%#d")

        # 清除不必要欄位
        schedule_B201.drop(columns=["實際排程日期", "排程依據日", "主工單開工日", "主工單編號", "主工單識別碼", "原始順序"], inplace=True, errors='ignore')

    else:
        print("B201刀次欄位沒有資料或全部為空，跳過後續處理")

    

    return schedule_A830, schedule_B201, schedule_A159



def generate_schedule_for_person(df: pd.DataFrame, break_dates: set, max_lookback_days: int = 60) -> pd.DataFrame:

    df = df.copy()
    df["預計開工日"] = pd.to_datetime(df["預計開工日"])
    df["預計完工日"] = pd.to_datetime(df["預計完工日"])
    df["實際排程日期"] = None

    # 每日產能限制（週一到週五）
    daily_limits = {0: 60, 1: 60, 2: 55, 3: 60, 4: 55}
    #daily_limits = {0: 55, 1: 55, 2: 55, 3: 55, 4: 55}

    # 建立主工單索引（主工單：刀次 > 0）
    last_main_idx = None
    df["主工單索引"] = None
    for idx, row in df.iterrows():
        if row["刀次"] > 0:
            last_main_idx = idx
        df.at[idx, "主工單索引"] = last_main_idx

    # 每組用「最早完工日」決定排程範圍
    df["組內最早完工日"] = df.groupby("主工單索引")["預計完工日"].transform("min")

    groups = df.groupby("主工單索引")
    sorted_groups = sorted(groups, key=lambda g: df.loc[g[0], "組內最早完工日"], reverse=True)

    # 建立排程容量表
    latest_end = df["預計完工日"].max()

    earliest_start = latest_end - timedelta(days=max_lookback_days)

    schedule_capacity = {}
    date = latest_end
    while date >= earliest_start:
        if date.weekday() < 5 and date not in break_dates:
            schedule_capacity[date] = daily_limits.get(date.weekday(), 55)
        date -= timedelta(days=1)

    for main_idx, group in sorted_groups:
        group_df = df.loc[group.index]
        earliest_end = group_df["組內最早完工日"].min()
        start_date = earliest_end - timedelta(days=max_lookback_days)

        total_doz = group_df["刀次"].sum()
        remaining = total_doz
        assigned_list = []
        current_date = earliest_end

        while remaining > 0 and current_date >= start_date:
            if current_date in schedule_capacity and schedule_capacity[current_date] > 0:
                available = schedule_capacity[current_date]
                assign = min(available, remaining)
                schedule_capacity[current_date] -= assign
                assigned_list.append((current_date.strftime("%Y-%m-%d"), assign))
                remaining -= assign
            current_date -= timedelta(days=1)

        if remaining > 0:
            raise Exception(f"排程無法完成，工單 {df.loc[main_idx, '工單編號']} 剩餘刀次 {remaining}")

        # 寫入每筆工單的排程日期
        for idx in group.index:
            if idx == main_idx:
                formatted = [f"{d}({c})" for d, c in sorted(assigned_list)]
            else:
                formatted = [d for d, _ in sorted(assigned_list)]
            df.at[idx, "實際排程日期"] = "\n".join(formatted)

    df.drop(columns=["組內最早完工日", "主工單索引"], inplace=True)

    df_new = split_schedule_dates(df)

    return df_new


def split_schedule_dates(df: pd.DataFrame) -> pd.DataFrame:
    new_rows = []
    for _, row in df.iterrows():
        sched_str = row["實際排程日期"]
        total_doz = row["刀次"]
        total_qty = row["需求數量"] if "需求數量" in row else None

        # 若不是字串或沒有 ()，表示沒有分配數量，直接保留
        if not isinstance(sched_str, str) or "(" not in sched_str or total_doz <= 60:
            # 若刀次 ≤ 60 或沒有分配數量資訊，不拆分，只改日期格式
            new_row = row.copy()
            try:
                # 防止錯誤字串，例如含括號
                date_part = str(sched_str).split("(")[0].strip()
                new_row["實際排程日期"] = pd.to_datetime(date_part).strftime("%Y/%m/%d")
            except Exception:
                new_row["實際排程日期"] = None  # 或你可以選擇保留原值
            new_rows.append(new_row)
            continue

        # 拆分多個日期與對應刀次
        lines = sched_str.strip().split("\n")

        for line in lines:
            new_row = row.copy()
            try:
                if "(" in line and ")" in line:
                    date_part = line.split("(")[0].strip()
                    doz_part = float(line.split("(")[1].replace(")", "").strip())

                    ratio = doz_part / total_doz if total_doz else 0
                    qty_part = round(total_qty * ratio) if total_qty is not None else None

                    new_row["刀次"] = doz_part
                    if qty_part is not None:
                        new_row["需求數量"] = qty_part

                    new_row["實際排程日期"] = pd.to_datetime(date_part).strftime("%Y/%m/%d")
                else:
                    # 格式只有日期無括號
                    new_row["實際排程日期"] = pd.to_datetime(line.strip()).strftime("%Y/%m/%d")
            except Exception:
                new_row["實際排程日期"] = None  # 若轉換失敗，設為 None
            new_rows.append(new_row)

    return pd.DataFrame(new_rows)


def final_cal_list(df: pd.DataFrame, start_d: datetime.date, break_dates: set) -> pd.DataFrame:
    df = df.copy()

    if isinstance(start_d, pd.Timestamp):
        start_d = start_d.date()
    elif isinstance(start_d, str):
        start_d = pd.to_datetime(start_d).date()

    daily_limits = {0: 60, 1: 60, 2: 55, 3: 60, 4: 55}
    special_limit = 60
    capacity_used = {}

    doses = pd.to_numeric(df["刀次"], errors="coerce").fillna(0).tolist()
    sku_list = df.get("品號", pd.Series([None]*len(doses))).tolist()
    new_dates = []
    current_date = start_d

    for dose, sku in zip(doses, sku_list):
        if dose == 0:
            # 子工單：繼承上一筆日期
            new_dates.append(new_dates[-1] if new_dates else current_date)
            continue

        while True:
            wd = current_date.weekday()
            if wd < 5 and current_date not in break_dates:
                used_info = capacity_used.get(current_date, {"count": 0, "sku_set": set(), "nonzero_sku_list": []})

                # 判斷當天非子工單品號（刀次>0）有哪些
                # 先從已分配當天的非子工單sku列表+目前這筆sku判斷
                current_nonzero_skus = [s for s in used_info["nonzero_sku_list"] if s is not None]
                # 加上當前正在安排的sku（dose != 0）
                if sku not in current_nonzero_skus:
                    current_nonzero_skus.append(sku)

                # 判斷當天是否只有一個品號(非子工單)
                is_single_sku = (len(set(current_nonzero_skus)) == 1)

                # 判斷是否為非加班日(週三2,週五4)，且是非假日
                is_special_day = (wd in [2, 4]) and (current_date not in break_dates)

                if is_single_sku and is_special_day:
                    limit = special_limit
                else:
                    limit = daily_limits.get(wd, 55)

                if used_info["count"] + dose <= limit:
                    used_info["count"] += dose
                    used_info["sku_set"].add(sku)
                    # 更新非子工單品號清單
                    if dose != 0 and sku is not None:
                        if sku not in used_info["nonzero_sku_list"]:
                            used_info["nonzero_sku_list"].append(sku)
                    capacity_used[current_date] = used_info
                    new_dates.append(current_date)
                    break

            current_date += timedelta(days=1)

    df.loc[:, "預計開工日"] = [d.strftime("%Y/%m/%d") for d in new_dates]
    df.loc[:, "預計完工日"] = df["預計開工日"]

    return df

class MultiPersonLeaveDialog(ctk.CTkToplevel):
    def __init__(self, parent, start_date, end_date, names=("A159", "A830", "B201")):
        super().__init__(parent)
        self.title("選擇三位人員的休假日")

        width, height = 900, 550
        self.geometry(f"{width}x{height}")
        self.center_window(width, height)

        self.leave_dates = {}
        self.names = names
        self.calendars = {}
        self.date_widgets = {}  # 可以是 label 或 dropdown

        ctk.CTkLabel(self, text="請為三位人員分別選擇休假日", font=("微軟正黑體", 20)).pack(pady=10)

        container = ctk.CTkFrame(self)
        container.pack(pady=10, padx=10, expand=True, fill="both")

        calendar_style = {
            "background": "#004080",
            "foreground": "white",
            "font": ("微軟正黑體", 14),
            "headersbackground": "#004080",
            "headersforeground": "white",
            "normalbackground": "white",
            "weekendbackground": "white",
            "disabledbackground": "#cccccc",
            "bordercolor": "#004080",
            "othermonthforeground": "#999999",
            "selectmode": "day",
            "date_pattern": "yyyy-mm-dd",
        }

        for i, name in enumerate(names):
            frame = ctk.CTkFrame(container)
            frame.grid(row=0, column=i, padx=10, pady=10)

            ctk.CTkLabel(frame, text=name, font=("微軟正黑體", 16)).pack(pady=5)

            cal = Calendar(frame, mindate=start_date, maxdate=end_date, **calendar_style)
            cal.pack(pady=5)
            self.calendars[name] = cal

            ctk.CTkButton(frame, text="加入休假日", font=("微軟正黑體", 14),
                          command=lambda n=name: self.add_date(n)).pack(pady=5)

            label = ctk.CTkLabel(frame, text="已選：無", wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            self.date_widgets[name] = label

            self.leave_dates[name] = set()

        self.error_label = ctk.CTkLabel(self, text="", text_color="red", font=("微軟正黑體", 12))
        self.error_label.pack(pady=5)

        ctk.CTkLabel(self, text="📝 若三人皆無休假，請直接按「✅ 確定」。", font=("微軟正黑體", 14), text_color="gray").pack(pady=(0, 10))

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=15)

        ctk.CTkButton(btn_frame, text="✅ 確定", font=("微軟正黑體", 14), command=self.confirm).pack(side="left", padx=20)
        #ctk.CTkButton(btn_frame, text="❌ 取消", font=("微軟正黑體", 14), command=self.cancel).pack(side="left", padx=20)

        self.result = None
        self.grab_set()

        #self.after(100, self.confirm)# 開發使用

    def center_window(self, width, height):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"+{x}+{y}")

    def add_date(self, name):
        selected = self.calendars[name].get_date()
        self.leave_dates[name].add(selected)
        self.update_widget(name)

    def remove_date(self, name, date):
        self.leave_dates[name].discard(date)
        self.update_widget(name)

    def show_delete_menu(self, event, name):
        if not self.leave_dates[name]:
            return

        font = tkfont.Font(family="微軟正黑體", size=18, weight="bold")
        menu = tk.Menu(self, tearoff=0, font=font)

        for date in sorted(self.leave_dates[name]):
            menu.add_command(
                label=f"🗑️　刪除日期：　{date}　",
                command=lambda d=date: self.remove_date(name, d)
            )

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()


    def update_widget(self, name):
        dates = sorted(list(self.leave_dates[name]))

        # 移除原本元件
        widget = self.date_widgets[name]
        widget.pack_forget()

        if len(dates) == 0:
            label = ctk.CTkLabel(widget.master, text="已選：無", wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            self.date_widgets[name] = label

        elif len(dates) <= 2:
            text = "\n".join(dates)
            label = ctk.CTkLabel(widget.master, text=text, wraplength=200,
                                 justify="left", font=("微軟正黑體", 14))
            label.pack(pady=5)
            label.bind("<Button-1>", lambda e, n=name: self.show_delete_menu(e, n))
            self.date_widgets[name] = label

        else:
            option_menu = ctk.CTkOptionMenu(widget.master, values=dates,
                                            command=lambda d: self.remove_date(name, d),
                                            font=("微軟正黑體", 14),
                                            width=180)
            option_menu.set("休假日列表(點選可刪除)")
            option_menu.pack(pady=5)
            self.date_widgets[name] = option_menu

    def confirm(self):
        self.result = self.leave_dates
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()
        
# 最後分配: 
# 1. 把剩餘工單內料號拿去已經配好的工單篩選若有與"基本資料符合"且"刀次不為0"的資料，則將取其一部份來搭配這原本剩餘工單
# 2. 排序要依照料號 -> 相同料號盡量擺在同一天生產
def final_doAssignAndSort(df_paired_split, df_no_pair_split, df_no_pair_second, remaining):
    # 先找到remaining之中剩餘工單的料號
    # 先去"基本資料"中找到其"搭一料號" 再去三人資料中搜尋是否有符合的
    # 若有符合的就拉其工單來做搭配 如無符合則將工單號留空白補上料號跟數量
    # 將完成的資料一樣排去分給B201, A830
    # 依照料號來重新排列工單 7個工作天相同料號盡量排在同一天生產

    # 基本資料:
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得腳本所在目錄
    json_path = os.path.join(script_dir, "config_ribbon.json")
    
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    # 基本資料路徑
    #base_path = r"E:\ribbon_schedule\\data\\基本資料-20250617-All.xlsx"
    base_path = config.get("base_path")
    if not base_path or not os.path.exists(base_path):
        print(f"❌ 找不到基本資料檔案，請確認路徑是否正確：\n{base_path}")
        return

    # 讀取基本資料
    base_df = pd.read_excel(base_path, dtype=str)
    base_df["搭1產出車數"] = pd.to_numeric(base_df["搭1產出車數"], errors="coerce")

    remaining_items = remaining["品號"].unique()  # 取唯一品號清單

    # 尋找可以搭配的料號
    all_staff_df = pd.concat([df_paired_split, df_no_pair_split, df_no_pair_second], ignore_index=True)

    # 建立一個空的 DataFrame 來放最終結果
    final_rows = []

    # 建立一次就好：料號 -> 寬度(cm)
    width_map = dict(zip(base_df["料號"], base_df["寬度Cm"].apply(_clean_cm)))

    for idx, row in remaining.iterrows():
        original_item = row["品號"]
        original_qty = row["預估良品數"]
        original_order = row.copy()

        row_match = base_df[base_df["料號"] == original_item]
        if row_match.empty:
            final_rows.append(original_order)
            continue

        # 讀取搭1/搭2資料
        d1_code = row_match.iloc[0]["搭1料號"]
        d1_car_count = pd.to_numeric(row_match.iloc[0]["搭1產出車數"], errors="coerce") or 0
        d2_code = row_match.iloc[0].get("搭2料號", "")
        d2_car_count = pd.to_numeric(row_match.iloc[0].get("搭2產出車數", ""), errors="coerce") or 0

        # 計算主工單刀次
        main_car_count = pd.to_numeric(original_order.get("車數", 1), errors="coerce") or 1
        if main_car_count <= 0:
            main_car_count = 1
        main_cut_count = math.ceil(pd.to_numeric(original_qty, errors="coerce") / main_car_count)
        original_order["刀次"] = main_cut_count

        # 找搭1配對
        candidates = all_staff_df[
            (all_staff_df["品號"] == d1_code) &
            (pd.to_numeric(all_staff_df["刀次"], errors="coerce") > 0)
        ]

        if not candidates.empty:
            paired_row = candidates.iloc[0].copy()

            # 主工單日期與人員同步子工單
            original_order["預計開工日"] = paired_row.get("預計開工日", original_order.get("預計開工日"))
            original_order["預計完工日"] = paired_row.get("預計完工日", original_order.get("預計完工日"))
            original_order["人員"] = paired_row.get("人員", original_order.get("人員"))

            final_rows.append(original_order)
            process_pair_order(d1_code, d1_car_count, main_cut_count, all_staff_df, original_order, final_rows, width_map)
            process_pair_order(d2_code, d2_car_count, main_cut_count, all_staff_df, original_order, final_rows, width_map)

        else:
            # 沒找到搭1配對
            final_rows.append(original_order)
            process_pair_order(d1_code, d1_car_count, main_cut_count, all_staff_df, original_order, final_rows, width_map)
            process_pair_order(d2_code, d2_car_count, main_cut_count, all_staff_df, original_order, final_rows, width_map)

    # 拆大批次、回傳結果
    final_rows_split = split_large_batches(final_rows, max_cut=55)
    final_df = pd.DataFrame(final_rows_split)

    return {"all": all_staff_df, "final_remaining": final_df}

def _clean_cm(x):
    s = str(x).lower().replace("cm", "").strip()
    try:
        v = float(s)
        return int(v) if v.is_integer() else v
    except:
        return pd.to_numeric(x, errors="coerce")


def process_pair_order(pair_code, car_count, main_cut_count, all_staff_df, original_order, final_rows, width_map):
    if not pair_code or str(pair_code).strip() == "" or pd.isna(pair_code) or car_count <= 0:
        return

    # 用 width_map 直接取
    width_value = width_map.get(str(pair_code))

    # 找現成工單
    candidates = all_staff_df[
        (all_staff_df["品號"] == pair_code) &
        (pd.to_numeric(all_staff_df["刀次"], errors="coerce") > 0)
    ]

    if not candidates.empty:
        paired_row = candidates.iloc[0].copy()
        paired_row["刀次"] = 0
        paired_row["車數"] = car_count
        paired_row["預估良品數"] = main_cut_count * car_count
        paired_row["餘量"] = paired_row["預估良品數"]
        if width_value is not None:
            paired_row["公分"] = width_value
        final_rows.append(paired_row)
    else:
        new_row = original_order.copy()
        new_row["工單編號"] = ""
        new_row["品號"] = pair_code
        new_row["車數"] = car_count
        new_row["刀次"] = 0
        new_row["預估良品數"] = main_cut_count * car_count
        new_row["餘量"] = new_row["預估良品數"]
        new_row["客戶需求日"] = ""
        if width_value is not None:
            new_row["公分"] = width_value
        final_rows.append(new_row)


def split_large_batches(rows, max_cut=55):
    split_result = []
    i = 0
    while i < len(rows):
        main_row = rows[i]
        # 預設子工單為 None
        sub_row = rows[i + 1] if (i + 1) < len(rows) and rows[i + 1].get("刀次", None) == 0 else None
            
        main_cut_raw = main_row.get("刀次", 0)
        main_cut = pd.to_numeric(main_cut_raw, errors="coerce")

        if pd.isna(main_cut):
            main_cut = 0
        else:
            main_cut = int(main_cut)

        # 若刀次沒超過限制或無子工單，直接加入
        if main_cut <= max_cut or sub_row is None:
            split_result.append(main_row)
            if sub_row is not None:
                split_result.append(sub_row)
            i += 2 if sub_row is not None else 1
            continue

        # 刀次超過 max_cut 的情況拆分
        remaining_cut = main_cut
        main_qty = main_row.get("預估良品數", 0)
        sub_qty = sub_row.get("預估良品數", 0) if sub_row is not None else 0

        while remaining_cut > 0:
            current_cut = min(max_cut, remaining_cut)
            ratio = current_cut / main_cut

            new_main = main_row.copy()
            new_main["刀次"] = current_cut
            new_main["預估良品數"] = round(main_qty * ratio)
            new_main["餘量"] = new_main["預估良品數"]
            split_result.append(new_main)

            if sub_row is not None:
                new_sub = sub_row.copy()
                new_sub["刀次"] = 0
                new_sub["預估良品數"] = round(sub_qty * ratio)
                new_sub["餘量"] = new_sub["預估良品數"]
                split_result.append(new_sub)

            remaining_cut -= current_cut

        i += 2 if sub_row is not None else 1

    return split_result

# 品號排序
# 相似品號"盡量"會擺在同一時期生產
def sort_by_customer_due_date(df: pd.DataFrame, break_dates: set = None) -> pd.DataFrame:
    df = df.copy().reset_index(drop=True)
    df["預計開工日"] = pd.to_datetime(df["預計開工日"], errors='coerce')

    if break_dates is None:
        break_dates = set()

    # 可設定的品號前綴規則 因應各種不同品號
    prefix_rules = {
        "B110": 5,
        "TTR": 4,
        "UTM": 4
    }
    default_prefix_len = 3

    def next_workdays(start_date, n):
        days = []
        d = start_date
        while len(days) <= n:  # 改成包含 start_date + 後續 n 天
            if d.weekday() < 5 and d not in break_dates:
                days.append(d)
            d += timedelta(days=1)
        return days

    result_frames = []
    idx = 0
    while idx < len(df):
        start_date = df.at[idx, "預計開工日"].date()
        workdays = next_workdays(start_date, 7)
        end_date = workdays[-1]

        mask = (df["預計開工日"].dt.date >= start_date) & (df["預計開工日"].dt.date <= end_date)
        chunk = df[mask].copy()

        if chunk.empty:
            # 沒有資料就跳到下一個日期
            next_start_idx = df.index[df["預計開工日"].dt.date > start_date].min()
            if pd.isna(next_start_idx):
                break
            idx = next_start_idx
            continue

        # 主子工單群組 (保持原順序)
        groups = []
        i = 0
        while i < len(chunk):
            row = chunk.iloc[i]
            group = [row]
            if row["刀次"] > 0:
                i += 1
                while i < len(chunk) and chunk.iloc[i]["刀次"] == 0:
                    group.append(chunk.iloc[i])
                    i += 1
            else:
                i += 1
            groups.append(group)

        # 品號分組（保留 others 在原 chunk 順序）
        grouped_by_code = defaultdict(list)
        others = []

        for g in groups:
            pn = str(g[0]["品號"]) if pd.notna(g[0]["品號"]) else None
            if pn:
                match_len = next((length for prefix, length in prefix_rules.items() if pn.startswith(prefix)), default_prefix_len)
                code = pn[:match_len]
                grouped_by_code[code].append(g)
            else:
                others.append(g)

        # 依 code 排序，每個 code 內保留原 chunk 順序
        sorted_groups = []
        for code in sorted(grouped_by_code.keys()):
            sorted_groups.extend(grouped_by_code[code])
        # others 保留在原位置
        sorted_groups.extend(others)

        sorted_chunk = pd.DataFrame([row for group in sorted_groups for row in group])
        result_frames.append(sorted_chunk)

        # 跳到下一個日期區間的第一筆
        next_start_idx = df.index[df["預計開工日"].dt.date > end_date].min()
        if pd.isna(next_start_idx):
            break
        idx = next_start_idx

    final_df = pd.concat(result_frames, ignore_index=True)
    
    # === 以上為品號分類 ===
    # === 合併同工單+品號（主工單刀次不為0）的群組 ===
    groups = []
    i = 0
    while i < len(final_df):
        row = final_df.iloc[i]
        group = [row]
        if row["刀次"] > 0:
            i += 1
            while i < len(final_df) and final_df.iloc[i]["刀次"] == 0:
                group.append(final_df.iloc[i])
                i += 1
        else:
            i += 1
        groups.append(group)

    key_to_groups = defaultdict(list)
    for idx, group in enumerate(groups):
        main = group[0]
        if pd.notna(main["工單編號"]) and pd.notna(main["品號"]) and main["刀次"] != 0:
            key = (main["工單編號"], main["品號"])
            key_to_groups[key].append((idx, group))

    used_indices = set()
    new_groups = []
    i = 0
    while i < len(groups):
        if i in used_indices:
            i += 1
            continue

        group = groups[i]
        main = group[0]
        key = (main["工單編號"], main["品號"])

        if key in key_to_groups and len(key_to_groups[key]) > 1:
            related = key_to_groups[key]
            base_idx, _ = related[0]
            if i == base_idx:
                merged_group = []
                for idx2, g in related:
                    if idx2 not in used_indices:
                        merged_group.extend(g)
                        used_indices.add(idx2)
                new_groups.append(merged_group)
            i += 1
        else:
            new_groups.append(group)
            i += 1

    final_df = pd.DataFrame([row for group in new_groups for row in group]).reset_index(drop=True)

    # === 根據每區最早的「客戶需求日」來整體排序 ===
    block = []
    temp_group = []
    for row in final_df.itertuples(index=False):
        if row.刀次 > 0 and temp_group:
            block.append(temp_group)
            temp_group = []
        temp_group.append(row)
    if temp_group:
        block.append(temp_group)

    def get_earliest_due(group):
        dates = [pd.to_datetime(r.客戶需求日, errors='coerce') for r in group if pd.notna(r.客戶需求日)]
        return min(dates) if dates else pd.Timestamp.max

    block.sort(key=get_earliest_due)

    final_df = pd.DataFrame([r._asdict() for group in block for r in group])

    # 日期欄轉格式
    final_df["預計開工日"] = pd.to_datetime(final_df["預計開工日"], errors='coerce')
    final_df["預計開工日"] = final_df["預計開工日"].dt.strftime("%Y/%#m/%#d")

    # 客戶需求日移到最前面
    cols = final_df.columns.tolist()
    if "客戶需求日" in cols:
        cols.insert(0, cols.pop(cols.index("客戶需求日")))
        final_df = final_df[cols]

    return final_df

def reorder_main_and_subs(df: pd.DataFrame) -> pd.DataFrame:
    """
    將每筆主工單與其後子工單重新排序，
    依序排列成「主工單 → 子工單 → 下一筆主工單 → 子工單」。
    子工單刀次為0，良品數 = 車數 * 對應主工單刀次。
    (已修正：使用 safe_int_conversion 確保 "車數" 和 "刀次" 轉換安全。)
    """
    df = df.copy().reset_index(drop=True)
    new_rows = []
    i = 0
    while i < len(df):
        row = df.iloc[i].copy()
        
        # 修正點 A: 使用 safe_int_conversion 確保主工單的刀次和車數安全
        current_cut = safe_int_conversion(row.get("刀次", 0))
        current_car = safe_int_conversion(row.get("車數", 0))

        if current_cut > 0:
            # 主工單
            main_row = row.copy()
            new_rows.append(main_row)
            
            # 找子工單
            j = i + 1
            while j < len(df):
                sub_row = df.iloc[j].copy()
                
                # 修正點 B: 使用 safe_int_conversion 確保子工單的刀次檢查安全
                sub_cut = safe_int_conversion(sub_row.get("刀次", 0))
                if sub_cut != 0:
                    break # 遇到下一張主工單，跳出子工單迴圈

                # 修正點 C (錯誤發生處): 使用 safe_int_conversion 確保子工單的車數安全
                sub_car = safe_int_conversion(sub_row.get("車數", 0))
                
                # 計算子工單預估良品數
                sub_row["預估良品數"] = sub_car * current_cut
                new_rows.append(sub_row)
                j += 1
                
            i = j # 將主迴圈索引跳到下一張主工單
        else:
            # 刀次=0，但沒有對應主工單 (孤立的子工單或單純刀次為0的工單)
            new_rows.append(row.copy())
            i += 1
            
    # 建立新的 DataFrame
    out_df = pd.DataFrame(new_rows)
    
    # 確保輸出欄位順序一致（可選，但保持習慣）
    if not df.empty and not out_df.empty:
        out_df = out_df[df.columns.intersection(out_df.columns).tolist() + 
                         [c for c in out_df.columns if c not in df.columns]]
    elif df.empty:
        return pd.DataFrame(columns=df.columns)
        
    return out_df.reset_index(drop=True)




def merge_same_day_orders_multi(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    df = df.copy()
    df["預計開工日"] = pd.to_datetime(df["預計開工日"])
    merged_rows = []

    i = 0
    while i < len(df):
        row = df.iloc[i]
        row_date = row["預計開工日"]

        if row["刀次"] > 0:
            main_cut = row["刀次"]
            main_car = row["車數"]
            main_id = row["工單編號"]

            # 找對應子工單
            j = i + 1
            subs = []
            while j < len(df) and df.iloc[j]["刀次"] == 0 and df.iloc[j]["預計開工日"] == row_date:
                sub_id = str(df.iloc[j]["工單編號"]).strip()
                if sub_id == "":
                    sub_id = "EMPTY_SUB"
                subs.append((sub_id, df.iloc[j]))
                j += 1

            merged = False
            if merged_rows:
                # 找前一筆主工單
                prev_idx = -1
                for k in range(len(merged_rows)-1, -1, -1):
                    if merged_rows[k]["刀次"] > 0:
                        prev_idx = k
                        break
                if prev_idx >= 0:
                    prev_main = merged_rows[prev_idx]
                    prev_main_id = prev_main["工單編號"]
                    prev_date = prev_main["預計開工日"]

                    # 找前一組子工單
                    prev_subs = []
                    m = prev_idx + 1
                    while m < len(merged_rows) and merged_rows[m]["刀次"] == 0 and merged_rows[m]["預計開工日"] == prev_date:
                        prev_sub_id = str(merged_rows[m]["工單編號"]).strip()
                        if prev_sub_id == "":
                            prev_sub_id = "EMPTY_SUB"
                        prev_subs.append((prev_sub_id, merged_rows[m]))
                        m += 1

                    # 判斷是否可合併
                    if prev_main_id == main_id and prev_date == row_date:
                        if len(subs) == 0 and len(prev_subs) == 0:
                            # 重新計算主工單
                            new_cut = prev_main["刀次"] + main_cut
                            merged_rows[prev_idx]["刀次"] = new_cut
                            merged_rows[prev_idx]["良品數"] = merged_rows[prev_idx]["車數"] * new_cut
                            merged = True
                        elif [s[0] for s in subs] == [ps[0] for ps in prev_subs]:
                            new_cut = prev_main["刀次"] + main_cut
                            merged_rows[prev_idx]["刀次"] = new_cut
                            merged_rows[prev_idx]["良品數"] = merged_rows[prev_idx]["車數"] * new_cut
                            for k, (_, sub_row) in enumerate(prev_subs):
                                merged_rows[prev_idx + 1 + k]["良品數"] = merged_rows[prev_idx + 1 + k]["車數"] * new_cut
                            merged = True

            if not merged:
                # 新增主工單
                new_main = row.copy()
                new_main["良品數"] = main_car * main_cut
                merged_rows.append(new_main)
                # 新增子工單
                for _, sub in subs:
                    new_sub = sub.copy()
                    new_sub["良品數"] = new_sub["車數"] * main_cut
                    merged_rows.append(new_sub)

            i = j
        else:
            merged_rows.append(row.copy())
            i += 1

    merged_df = pd.DataFrame(merged_rows)
    merged_df["預計開工日"] = merged_df["預計開工日"].dt.strftime("%Y-%m-%d")

    # -----------------------------
    # 🔹 統一重新計算良品數與餘量
    # -----------------------------
    recalculated = []
    i = 0
    while i < len(merged_df):
        row = merged_df.iloc[i].copy()
        if row["刀次"] > 0:  # 主工單
            main_cut = row["刀次"]
            main_car = row["車數"]

            row["預估良品數"] = main_car * main_cut
            row["餘量"] = row["預估良品數"]
            recalculated.append(row)

            # 處理子工單
            j = i + 1
            while j < len(merged_df) and merged_df.iloc[j]["刀次"] == 0 and merged_df.iloc[j]["預計開工日"] == row["預計開工日"]:
                sub = merged_df.iloc[j].copy()
                sub["預估良品數"] = sub["車數"] * main_cut
                sub["餘量"] = sub["預估良品數"]
                recalculated.append(sub)
                j += 1

            i = j
        else:
            # 孤立子工單
            row["預估良品數"] = row["車數"] * row["刀次"]
            row["餘量"] = row["預估良品數"]
            recalculated.append(row)
            i += 1

    return pd.DataFrame(recalculated)

def safe_int_conversion(value: Any, default: int = 0) -> int:
    """
    安全地將值轉換為整數。處理 None, NaN, 空字串等情況，避免 ValueError。
    (已提升至頂層，供所有函式共用。)
    """
    # 檢查是否為缺失值 (NaN, None) 或空字串
    if pd.isna(value) or value is None or str(value).strip() == "":
        return default
    try:
        # 先轉為 float 處理科學記號或奇怪格式，再轉為 int
        return int(float(value))
    except (ValueError, TypeError):
        return default

def merge_order_cutNum(A_df: pd.DataFrame, B_df: pd.DataFrame, C_df: pd.DataFrame,
                       daily_standard: int = 55) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    將每個人 (A_df, B_df, C_df) 的工單往前遞補滿每日標準刀次 daily_standard (預設55)。
    若某天原本只有一張主工單，當日上限允許到 60。
    子工單跟隨主工單切分：子良品數 = 子車數 * 當天分配刀次。
    回傳三個處理後的 DataFrame（A_df、B_df、C_df）。
    """

    def process_one(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame()

        df = df.copy().reset_index(drop=True)
        # 進行日期清理，避免 NaT 影響後續處理
        df["預計開工日"] = pd.to_datetime(df["預計開工日"], errors="coerce")
        df["預計完工日"] = pd.to_datetime(df["預計完工日"], errors="coerce")
        
        # 建立主工單與子工單結構
        orders = []
        last_main = None
        for i, row in df.iterrows():
            # 修正點 1: 安全轉換 "刀次"
            current_knives = safe_int_conversion(row.get("刀次", 0))

            # 修正點 2: 安全轉換 "車數"
            current_cars = safe_int_conversion(row.get("車數", 0))

            if current_knives > 0:
                # 主工單邏輯
                last_main = {
                    "orig_idx": i,
                    "orig_row": row.copy(),
                    "orig_date": row["預計開工日"].to_datetime64(), # 確保日期類型正確
                    "remain": current_knives, # 使用安全轉換後的值
                    "car": current_cars,      # 使用安全轉換後的值
                    "children": []
                }
                orders.append(last_main)
            else:
                # 子工單邏輯
                if last_main is not None:
                    last_main["children"].append(row.copy())
                else:
                    # 處理沒有主工單的子工單（但這種情況通常應避免）
                    last_main = {
                        "orig_idx": i,
                        "orig_row": row.copy(),
                        "orig_date": pd.to_datetime(row.get("預計開工日", pd.NaT)),
                        "remain": 0,
                        "car": current_cars, # 使用安全轉換後的值
                        "children": []
                    }
                    orders.append(last_main)

        if not orders:
            return df

        # 初始化切片：每筆主工單全部放在原始日期
        for od in orders:
            od["slices"] = {}
            if od["remain"] > 0 and pd.notna(od["orig_date"]):
                od["slices"][pd.to_datetime(od["orig_date"]).normalize()] = od["remain"]

        # 取得原始日期排序
        all_dates = sorted({d.normalize() for od in orders for d in od["slices"].keys() if pd.notna(d)})

        # 每日總刀數及主工單數量
        def daily_status(date):
            total = 0
            distinct_mains = set()
            for od in orders:
                v = od["slices"].get(date, 0)
                if v > 0:
                    total += v
                    distinct_mains.add(od["orig_idx"])
            return total, len(distinct_mains)

        # 往前補刀次
        for date in all_dates:
            _, distinct = daily_status(date)
            day_limit = 60 if distinct == 1 else daily_standard

            # 禮拜三和禮拜五固定上限 55
            if date.weekday() in [2, 4]:
                day_limit = min(day_limit, 55)

            total_now, _ = daily_status(date)
            available = day_limit - total_now
            if available <= 0:
                continue

            while available > 0:
                cand = []
                for od in orders:
                    for sdate, amt in od["slices"].items():
                        # 檢查 sdate 是否有效且大於當前日期
                        if pd.notna(sdate) and sdate.normalize() > date and amt > 0:
                            cand.append((sdate, od, amt))
                if not cand:
                    break
                cand.sort(key=lambda x: (x[0], x[1]["orig_idx"]))
                pick_date, pick_od, pick_amt = cand[0]
                move = min(pick_amt, available)
                pick_od["slices"][pick_date] -= move
                if pick_od["slices"][pick_date] <= 0:
                    del pick_od["slices"][pick_date]
                pick_od["slices"][date] = pick_od["slices"].get(date, 0) + move
                available -= move

        # 展開切片成 DataFrame
        out_rows = []
        for od in orders:
            for sdate, amt in sorted(od["slices"].items(), key=lambda x: x[0]):
                if amt > 0:
                    # 處理主工單
                    main_row = od["orig_row"].copy()
                    main_row["預計開工日"] = sdate
                    main_row["預計完工日"] = sdate
                    main_row["刀次"] = int(amt)
                    main_car = safe_int_conversion(main_row.get("車數", 0)) # 確保安全轉換
                    main_row["預估良品數"] = main_car * int(amt)
                    out_rows.append(main_row)

                    # 處理子工單
                    for ch in od["children"]:
                        ch_row = ch.copy()
                        ch_row["預計開工日"] = sdate
                        ch_row["預計完工日"] = sdate
                        ch_row["刀次"] = 0
                        
                        # 修正點 3: 安全轉換 "車數"
                        ch_car = safe_int_conversion(ch_row.get("車數", 0)) 
                        
                        ch_row["預估良品數"] = ch_car * int(amt)
                        
                        if "餘量" in ch_row.index:
                            # 由於 Pandas Series 的 index 查詢是 based on labels，這裡使用 .get 避免 KeyError
                            ch_row["餘量"] = ch_row["預估良品數"]
                        
                        out_rows.append(ch_row)

        if not out_rows:
            # 處理空列表情況
            return pd.DataFrame(columns=df.columns)

        out_df = pd.DataFrame(out_rows)
        # 確保輸出的欄位順序與輸入的 df 盡可能一致
        out_df = out_df[df.columns.intersection(out_df.columns).tolist() + 
                         [c for c in out_df.columns if c not in df.columns]]
        out_df["預計開工日"] = out_df["預計開工日"].dt.strftime("%Y-%m-%d")
        out_df["預計完工日"] = out_df["預計完工日"].dt.strftime("%Y-%m-%d")

        return out_df.reset_index(drop=True)

    A_out = process_one(A_df)
    B_out = process_one(B_df)
    C_out = process_one(C_df)

    return A_out, B_out, C_out


def Erin_use(A830_df, B201_df, A159_df):
    """
    此函式主要處理 Erin 提出的兩點需求：
    1. 根據品號，從基本資料檔中查找並補上對應的「米平方」欄位。
    2. 新增「週分組」欄位，定義每週為「週五到隔週四」，並依此排序。
    """
    print("正在執行新增需求：加入米平方與週分組...")

    # --- 步驟 1: 讀取基本資料以取得米平方 ---
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(script_dir, "config_ribbon.json")
        with open(json_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        base_path = config.get("base_path")
        if not base_path or not os.path.exists(base_path):
            print(f"⚠️ 警告：找不到基本資料檔案，無法新增米平方。路徑：{base_path}")
            return A830_df, B201_df, A159_df

        base_df = pd.read_excel(base_path, dtype=str)
        # 確保基本資料中有 '料號' 和 '米平方' 欄位
        if '料號' not in base_df.columns or '面積  M2' not in base_df.columns:
            print("⚠️ 警告：基本資料缺少 '料號' 或 '面積  M2' 欄位，無法新增。")
            return A830_df, B201_df, A159_df
            
        # 只選取需要的欄位進行合併，更高效
        m2_df = base_df[['料號', '面積  M2']].copy()
    except Exception as e:
        print(f"讀取基本資料時發生錯誤，無法新增米平方：{e}")
        return A830_df, B201_df, A159_df
    
    # --- 步驟 2 & 3: 處理每個員工的 DataFrame ---
    processed_dfs = []
    staff_dfs = [A830_df, B201_df, A159_df]

    for df in staff_dfs:
        if df is None or df.empty:
            processed_dfs.append(df)
            continue

        # 在處理米平方之前，先檢查並移除 '良品數' 欄位
        if '良品數' in df.columns:
            df = df.drop(columns=['良品數'])

        # --- 2. 加入米平方欄位 ---
        #  修改 #2：使用 left_on 和 right_on 來指定不同的欄位名稱 
        # 左邊 df 用 '品號', 右邊 m2_df 用 '料號'
        merged_df = pd.merge(df, m2_df, left_on='品號', right_on='料號', how='left')

        #  修改 #3：合併後，把多餘的 '料號' 欄位刪除 
        if '料號' in merged_df.columns:
            merged_df = merged_df.drop(columns=['料號'])

        # 1. 將米平方欄位轉為數字，若有錯則變為空值(NaN)
        merged_df['面積  M2'] = pd.to_numeric(merged_df['面積  M2'], errors='coerce')

        # 2. 進行傳統四捨五入，並轉為支援空值的整數型別
        #    (這裡使用一個小技巧：先加0.5再轉成整數，來實現傳統四捨五入)
        merged_df['面積  M2'] = np.floor(merged_df['面積  M2'] + 0.5)

        # 3. 轉成支援空值的整數型別
        merged_df['面積  M2'] = merged_df['面積  M2'].astype('Int64')

        # --- 3. 新增週分組欄位 ---
        merged_df['開工日_dt'] = pd.to_datetime(merged_df['預計開工日'], errors='coerce')

        def get_week_start_friday(date):
            if pd.isna(date):
                return None
            if date.weekday() >= 4:
                return date - pd.Timedelta(days=date.weekday() - 4)
            else:
                return date - pd.Timedelta(days=date.weekday() + 3)

        merged_df['週分組_起始日'] = merged_df['開工日_dt'].apply(get_week_start_friday)
        
        def format_week_range(start_date):
            if pd.isna(start_date):
                return ""
            end_date = start_date + pd.Timedelta(days=6)
            return f"{start_date.strftime('%Y/%m/%d')} ~ {end_date.strftime('%Y/%m/%d')}"
        
        merged_df['週分組'] = merged_df['週分組_起始日'].apply(format_week_range)

        merged_df = merged_df.rename(columns={'面積  M2': '米平方'})

        # --- 4. 重新排序並整理欄位 ---
        sorted_df = merged_df.sort_values(by=['週分組_起始日', '開工日_dt'], ascending=[True, True])
        sorted_df = sorted_df.drop(columns=['開工日_dt', '週分組_起始日'])
        
        cols = sorted_df.columns.tolist()
        week_col = cols.pop(cols.index('週分組'))
        m2_col = cols.pop(cols.index('米平方'))
        cols.insert(0, week_col)
        #cols.append(m2_col)
        cols.insert(6, m2_col)
        final_df = sorted_df[cols]

        processed_dfs.append(final_df)

    print("✅ 新增需求處理完成。")
    return processed_dfs[0], processed_dfs[1], processed_dfs[2]


def main():

    # 初階段工單分類 (輸入開始結束日期，根據config找到基本資料)
    result = process_schedule_data()  

    df_remaining = result["remaining_df"]
    df_no_pair = result["no_pair_df"]
    df_paired = result["paired_df"]

    # 無須配對的工單進行分組尚未排程
    df_no_pair_split = split_no_pair_rows(df_no_pair)

    # 配對的工單進行分組尚未排程
    df_paired_split_1, extra_remaining = split_paired_rows(df_paired)
    df_paired_split_2 = split_paired_rows_step1(df_paired_split_1)

    # 合併 extra_remaining 並去除重複的工單
    df_remaining = pd.concat([df_remaining, extra_remaining], ignore_index=True)
    df_remaining = df_remaining.drop_duplicates(subset="工單編號", keep="last")

    # 將剩餘可不用匹配就可以獨立的工單找出 根據基本資料 搭1料號為0或空 , 剩餘數量小於刀次就要強制多切
    result_second = process_schedule_data_second(df_remaining)
    df_reamining_second = result_second["remaining_df"]
    df_no_pair_second = result_second["no_pair_df"]
    
    # 沒啥作用 -> 把剩餘工單的刀次改為""
    # 新增功能 找到可以搭配的 88cm and 68cm
    df_paired_split, remaining = remaining_cut_clean_and_repair(df_paired_split_2, df_reamining_second, mode = "88cm") 
    df_paired_split, remaining = remaining_cut_clean_and_repair(df_paired_split, remaining, mode = "68cm") 

    # 剩餘工單排序 分配
    #result_final = final_doAssignAndSort(df_paired_split, df_no_pair_split, df_no_pair_second, df_reamining_second)
    result_final = final_doAssignAndSort(df_paired_split, df_no_pair_split, df_no_pair_second, remaining)
    final_remaining = result_final["final_remaining"]
    #all = result_final["all"]
    #combined_df_test = pd.concat([df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining], ignore_index=True)

    # 排版,依照預計開工時間排序,尚未加上預計完工時間,初步人員分類
    #final_output = prepare_final_schedule(df_paired_split, df_no_pair_split, df_no_pair_second)
    final_output = prepare_final_schedule(df_paired_split, df_no_pair_split, df_no_pair_second, final_remaining)

    # 工單分配
    A159_part, A830_part, B201_part = do_people(final_output)

    # 日期排程 (一天標準刀次55車，加班可以到65車，滿足出貨不需要安排加班)
    # 輸入休假 彈出視窗 選擇人員以及休假日期
    A830_part_end, B201_part_end, A159_part_end = final_schedule_list(A159_part, A830_part, B201_part)

    # 將工單往前遞補滿每日的標準刀次 
    A830_part_end, B201_part_end, A159_part_end = merge_order_cutNum(A830_part_end, B201_part_end, A159_part_end) 

    # 最後判斷每天的工單 如果有相同的主工單 那就合併~(包含其子工單也會合併) , A159家偉資料為空 
    A830_part_end_reorder = reorder_main_and_subs(A830_part_end)
    B201_part_end_reorder = reorder_main_and_subs(B201_part_end)
    A159_part_end_reorder = reorder_main_and_subs(A159_part_end)
    A830_part_end = merge_same_day_orders_multi(A830_part_end_reorder) 
    B201_part_end = merge_same_day_orders_multi(B201_part_end_reorder)  
    A159_part_end = merge_same_day_orders_multi(A159_part_end_reorder) 

    # 嘉真需求 補上米平方 & 按週分組
    A830_part_end, B201_part_end, A159_part_end = Erin_use(A830_part_end, B201_part_end, A159_part_end)

    # === 寫入 Excel ===
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 取得目前.py檔的資料夾
    output_name = os.path.join(script_dir, "排程結果_1009.xlsx")   # 存到同一層資料夾
    if os.path.exists(output_name):
        try:
            os.remove(output_name)
        except PermissionError:
            raise Exception(f"Excel 檔案 {output_name} 已開啟，請先關閉後再執行。")
    with pd.ExcelWriter(output_name, engine="openpyxl") as writer:
        #df_paired_split_1.to_excel(writer, sheet_name="初始剩餘1", index=False)
        #df_paired.to_excel(writer, sheet_name="初始剩餘", index=False)
        #extra_remaining.to_excel(writer, sheet_name="初始剩餘_剩餘", index=False)
        #df_no_pair_split.to_excel(writer, sheet_name="測試1", index=False)
        #df_no_pair_second.to_excel(writer, sheet_name="測試2", index=False)
        #remaining.to_excel(writer, sheet_name="最初剩餘工單", index=False)
        #final_remaining.to_excel(writer, sheet_name="最後剩餘工單", index=False)
        #A159_part_end_reorder.to_excel(writer, sheet_name="家偉_第一階段", index=False)
        #B201_part_end_reorder.to_excel(writer, sheet_name="旺斌_第一階段", index=False)
        #A830_part_end_reorder.to_excel(writer, sheet_name="容合_第一階段", index=False)
        #all.to_excel(writer, sheet_name="all", index=False)
        #combined_df_test.to_excel(writer, sheet_name="test_conbine", index=False)
        A159_part_end.to_excel(writer, sheet_name="家偉", index=False)
        B201_part_end.to_excel(writer, sheet_name="旺斌", index=False)
        A830_part_end.to_excel(writer, sheet_name="容合", index=False)

    # Excel 欄寬調整
    def get_display_width(text):
        width = 0
        for ch in str(text):
            width += 4 if unicodedata.east_asian_width(ch) in ('F', 'W', 'A') else 2
        return width

    wb = load_workbook(output_name)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, name="Calibri")

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.font = Font(name="Calibri")

        for col in ws.columns:
            max_len = max((get_display_width(cell.value) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    wb.save(output_name)

    print("End running")

    return output_name

'''
if __name__ == "__main__":
    main()  # 主程式
'''

# 執行 GUI
if __name__ == "__main__":
    app = ScheduleApp()
    app.mainloop()
